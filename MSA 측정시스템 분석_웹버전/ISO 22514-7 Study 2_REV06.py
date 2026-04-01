# -*- coding: utf-8 -*-
# ISO 22514-7 & VDA5 MSA Gage R&R Study 1·2·3 (개선 통합본) - Rev.07
# 변경 요약
# - 폰트 설정: OS(Windows, Mac, Linux) 자동 감지 및 한글 폰트 적용 (set_font 함수)
# - 메모리 관리: 그래프 재생성 시 기존 plt.Figure 객체 명시적 종료 강화
# - 데이터 로드: 컬럼명 공백 제거 및 데이터 타입 변환 안정성 확보

import re
import json
import os
import time
import traceback
import platform
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import numpy as np
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
from scipy.stats import norm, f, t as t_dist

import matplotlib.pyplot as plt
from matplotlib import font_manager, rc
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.ticker import FormatStrFormatter

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---- Optional tools for capture fallbacks ----
try:
    from PIL import ImageGrab, Image
except Exception:
    ImageGrab = None
    Image = None

try:
    import pyautogui
except Exception:
    pyautogui = None

NOTE_OFFSET_FRAC = 0.02
EPS = 1e-12

_save_data = {}


# ===== 1. 폰트 설정 (OS 호환성 개선) =====
def set_font():
    """운영체제에 따라 한글 폰트를 자동 설정합니다."""
    system_name = platform.system()
    try:
        if system_name == "Windows":
            font_path = "C:/Windows/Fonts/malgun.ttf"
            if os.path.exists(font_path):
                font_name = font_manager.FontProperties(fname=font_path).get_name()
                rc('font', family=font_name)
            else:
                rc('font', family="Malgun Gothic")
        elif system_name == "Darwin":  # Mac
            rc('font', family="AppleGothic")
        else:  # Linux etc.
            rc('font', family="NanumGothic")
    except Exception:
        pass
    plt.rcParams["axes.unicode_minus"] = False


set_font()

UNC_NOTES = {
    "u_CAL": "교정(표준불확도)",
    "u_BI": "치우침(Bias) 불확도",
    "u_EVR": "반복성(실험 표준편차)",
    "u_RE": "분해능/계기 해상도",
    "u_EV": "측정장치 불확도(반복성 vs RE 중 큰값)",
    "u_EVO": "반복성(ANOVA 잔차 기반)",
    "u_AV": "측정자 간 차이(재현성)",
    "u_IA": "부품×측정자 교호작용",
    "u_LIN": "선형성",
    "u_MSREST": "기타 측정시스템 잔여",
    "u_T": "환경(온도 등)",
    "u_STAB": "장기 안정성",
    "u_OBJ": "부품/측정물 영향",
    "u_GV": "게이지 자체 변동",
    "u_REST": "기타 잔여"
}


# ---------- 유틸 ----------
def get_colors(n: int):
    cmap = plt.get_cmap("tab20")
    return [cmap(i % 20) for i in range(n)]


def safe_load_df(file_path: str) -> pd.DataFrame:
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    try:
        if ext == ".csv":
            try:
                df = pd.read_csv(file_path)
            except UnicodeDecodeError:
                df = pd.read_csv(file_path, encoding="cp949")
        elif ext in [".xlsx", ".xls"]:
            df = pd.read_excel(file_path)
        else:
            raise ValueError("CSV 또는 XLSX 파일만 지원합니다.")

        if df.shape[1] < 3:
            raise ValueError("파일에는 최소 3개 컬럼(부품, 측정자, 측정값)이 있어야 합니다.")

        # 컬럼 선택 및 이름 정리 (공백 제거)
        df = df.iloc[:, :3].copy()
        df.columns = ["부품", "측정자", "측정값"]

        # 데이터 전처리 (문자열 공백 제거 등)
        for col in ["부품", "측정자"]:
            if df[col].dtype == object:
                df[col] = df[col].astype(str).str.strip()

        # 측정값 숫자 변환
        df["측정값"] = pd.to_numeric(df["측정값"], errors='coerce')
        df = df.dropna(subset=["측정값"])  # 숫자가 아닌 행 제거

        return df
    except Exception as e:
        raise ValueError(f"파일 로드 실패: {e}")


def browse_file():
    path = filedialog.askopenfilename(
        filetypes=[
            ("CSV/Excel Files", "*.csv;*.xlsx;*.xls"),
            ("CSV Files", "*.csv"),
            ("Excel Files", "*.xlsx;*.xls"),
        ]
    )
    if path:
        entry_file.delete(0, tk.END)
        entry_file.insert(0, path)


def parse_measurement_input(text_widget):
    raw = text_widget.get("1.0", tk.END)
    if not raw.strip():
        return []
    # 쉼표, 세미콜론, 공백, 줄바꿈 등으로 분리
    tokens = re.split(r'[\s,;]+', raw.strip())
    vals = []
    for tok in tokens:
        if not tok:
            continue
        try:
            vals.append(float(tok))
        except ValueError:
            continue
    return vals


def calc_u_distribution(val, method, CP=0.9545):
    if method == "균등분포":
        return val / np.sqrt(12)
    if method == "정규분포":
        return val / (2 * norm.ppf((1 + CP) / 2))
    if method == "삼각분포":
        return val / np.sqrt(24)
    if method == "U분포":
        return val / np.sqrt(8)
    return 0


def format_anova(df: pd.DataFrame) -> str:
    df_disp = df.reset_index().rename(columns={'index': '항목'})
    fmt = "{:<15s} {:>8s} {:>12s} {:>12s} {:>10s} {:>10s}\n"
    lines = [fmt.format("항목", "DF", "SS", "MS", "F값", "P값")]
    for _, row in df_disp.iterrows():
        df_str = f"{row['자유도(DF)']:8.4f}" if not pd.isna(row['자유도(DF)']) else f"{'':>8}"
        ss_str = f"{row['제곱합(SS)']:12.6f}" if not pd.isna(row['제곱합(SS)']) else f"{'':>12}"
        ms_val = row['평균제곱(MS)']
        ms_str = f"{ms_val:12.7f}" if isinstance(ms_val, (int, float, np.floating)) and not pd.isna(
            ms_val) else f"{'':>12}"
        f_val = row.get('F값', "")
        f_str = f"{f_val:10.4f}" if isinstance(f_val, (int, float, np.floating)) and not pd.isna(f_val) else f"{'':>10}"
        p_val = row.get('P값', "")
        p_str = f"{p_val:10.4f}" if isinstance(p_val, (int, float, np.floating)) and not pd.isna(p_val) else f"{'':>10}"
        lines.append(fmt.format(str(row['항목'])[:15], df_str, ss_str, ms_str, f_str, p_str))
    return "".join(lines)


# ---- 텍스트 출력 유틸 ----
def append_gage_table_to_text(df_gage: pd.DataFrame):
    text_result.insert(tk.END, "\n=== Gage  R&R 분석표 ===\n", "black_tag")
    hdr = "{:<12s} {:>12s} {:>12s} {:>10s} {:>12s}\n".format(
        "출처", "표준 편차(SD)", "연구 변동(6×SD)", "%연구 변동(%SV)", "%공차(SV/공차)"
    )
    text_result.insert(tk.END, hdr)
    for _, r in df_gage.iterrows():
        line = "{:<12s} {:>12s} {:>12s} {:>10s} {:>12s}\n".format(
            r["출처"], r["표준 편차(SD)"], r["연구 변동\n(6×SD)"], r["%연구 변동\n(%SV)"], r["%공차\n(SV/공차)"]
        )
        text_result.insert(tk.END, line)


def append_metrics_and_judgement(ndc_val, sd_part, sd_gagerr, Q_MP, C_MP,
                                 pct_sv_gagerr, pct_tol_gagerr):
    # 요약 메트릭(OK/NOT OK 라벨 포함)
    text_result.insert(tk.END, "\n", "black_tag")
    ndc_int = int(round(ndc_val)) if not np.isnan(ndc_val) else np.nan
    line = f"* NDC (구별범주의 수) = 1.41 × σ_pt / σ_G_R_R = {ndc_int}\n"
    text_result.insert(tk.END, line)
    text_result.insert(tk.END, ("  OK: 5 이상\n" if ndc_val >= 5 else "  NOT OK: 5 미만\n"),
                       "ok" if ndc_val >= 5 else "not_ok")

    if not np.isnan(Q_MP):
        text_result.insert(tk.END, f"* Q_MP(측정프로세스 성능비) = {Q_MP:0.2f}%\n")
        text_result.insert(tk.END, ("  OK: 30% 이하\n" if Q_MP <= 30 else "  NOT OK: 30% 초과\n"),
                           "ok" if Q_MP <= 30 else "not_ok")

    if not np.isnan(C_MP):
        text_result.insert(tk.END, f"* C_MP(측정프로세스 능력지수) = {C_MP:0.2f}\n")
        text_result.insert(tk.END, ("  OK: 1.33 이상 OK\n" if C_MP >= 1.33 else "  NOT OK: 1.33 미만\n"),
                           "ok" if C_MP >= 1.33 else "not_ok")

    # 판정 결과 표기
    text_result.insert(tk.END, "=== 판정 결과 ===\n", "black_tag")

    def okline(title, val, cond, unit=""):
        tag = "pass" if cond else "fail"
        if np.isnan(val):
            text_result.insert(tk.END, f"{title:<20s} NaN   -> FAIL\n", "fail")
        else:
            text_result.insert(tk.END, f"{title:<20s} {val:>6.2f}{unit} -> {'PASS' if cond else 'FAIL'}\n", tag)

    okline("1. NDC ≥ 5", ndc_val, ndc_val >= 5)
    if not np.isnan(Q_MP):
        okline("2. Q_MP ≤ 30%", Q_MP, Q_MP <= 30, unit="%")
    okline("3. %SV ≤ 30%", pct_sv_gagerr, pct_sv_gagerr <= 30, unit="%")
    okline("4. %공차 ≤ 30%", pct_tol_gagerr, pct_tol_gagerr <= 30, unit="%")


# ---------- (신규) 태그 보존 공백 압축 ----------
def _compress_blank_lines_inplace(widget: tk.Text):
    """
    연속 개행을 1개로 줄인다. 정규식/전역 검색을 쓰지 않고
    '\n\n' 위치만 찾아서 뒤쪽 개행을 제거하므로 빠르고 안전함.
    """
    idx = "1.0"
    end_limit = "end-1c"
    while True:
        start = widget.search("\n\n", idx, end_limit)  # 두 줄 연속만 찾음(비-정규식)
        if not start:
            break
        # 첫 개행 하나는 남기고, 그 뒤로 연속되는 개행만 지운다
        pos = f"{start} +1c"
        while True:
            ch = widget.get(pos, f"{pos} +1c")
            if ch != "\n":
                break
            widget.delete(pos, f"{pos} +1c")
        idx = start


# ---------- 실행 모드 동기화 ----------
def on_toggle_both():
    if var_run_both.get():
        var_run_t1.set(True)
        var_run_t2.set(True)


def on_toggle_t1_t2(_evt=None):
    if var_run_t1.get() and var_run_t2.get():
        var_run_both.set(True)
    else:
        var_run_both.set(False)


# ---------- 불확도 탭(표) ----------
def build_unc_rows(nz_items, u_mp):
    rows = []
    if u_mp and u_mp > 0:
        for k, v in nz_items:
            pct = (v / u_mp) * 100.0
            rows.append((k, v, pct, UNC_NOTES.get(k, "")))
    else:
        for k, v in nz_items:
            rows.append((k, v, 0.0, UNC_NOTES.get(k, "")))
    return rows


def populate_uncertainty_tab(nz_items, u_mp):
    for iid in tv_unc.get_children():
        tv_unc.delete(iid)
    rows = build_unc_rows(nz_items, u_mp)
    if not rows:
        return
    max_val = max(v for _, v, _, _ in rows)
    for idx, (k, v, pct, note) in enumerate(rows):
        tag = 'odd' if (idx % 2) else 'even'
        add_tags = (tag,)
        if abs(v - max_val) <= 1e-15:
            add_tags = (tag, 'maxrow')
        tv_unc.insert("", "end", values=(k, f"{v:.6f}", f"{pct:5.2f}", note), tags=add_tags)
    tv_unc.insert("", "end", values=("합성 u", f"{u_mp:.6f}", "100.00", "합성 표준불확도"), tags=("sumrow",))
    sort_tv_unc("값", reverse=True)


_tv_unc_sort_state = {"항목": False, "값": True, "비중(%)": True, "주석": False}


def sort_tv_unc(col, reverse=False):
    def _to_float(s):
        try:
            return float(str(s).replace('%', ''))
        except Exception:
            return 0.0

    items = list(tv_unc.get_children())
    data = []
    for iid in items:
        vals = tv_unc.item(iid, 'values')
        if col == "값":
            key = _to_float(vals[1])
        elif col == "비중(%)":
            key = _to_float(vals[2])
        else:
            key = vals[0 if col == "항목" else 3]
        data.append((key, iid))
    data.sort(reverse=reverse, key=lambda x: x[0])
    for _, iid in data:
        tv_unc.move(iid, '', 'end')
    _tv_unc_sort_state[col] = reverse


# ---------- 공차/USL·LSL 모드 ----------
def on_tol_mode_change(*_):
    mode = var_tol_mode.get()
    if mode == "tolerance":
        entry_TOL.config(state="normal")
        entry_USL.config(state="disabled")
        entry_LSL.config(state="disabled")
    else:
        entry_TOL.config(state="disabled")
        entry_USL.config(state="normal")
        entry_LSL.config(state="normal")


def calc_limits_from_tol(refv: float, tol: float):
    return refv + tol / 2.0, refv - tol / 2.0


def preview_cg_from_tol():
    try:
        tol = float(entry_TOL.get())
        refv = float(entry_RefV.get())
        X = np.array(parse_measurement_input(text_X))
        if len(X) < 2:
            raise ValueError("Type I 값(최소 2개)이 필요합니다.")
        sigma_repeat = np.std(X, ddof=1)
        if sigma_repeat <= 0:
            raise ValueError("표준편차가 0입니다.")
        usl, lsl = calc_limits_from_tol(refv, tol)
        mu_repeat = float(np.mean(X))
        TOL = float(usl - lsl)
        Cg = (0.2 * TOL) / (6 * sigma_repeat)
        Cgk = (0.1 * TOL - abs(mu_repeat - refv)) / (3 * sigma_repeat)
        lbl_cg_preview.config(
            text=f"[공차 기반 미리보기]  USL={usl:.6f}, LSL={lsl:.6f}  |  Cg={Cg:.4f}, Cgk={Cgk:.4f}"
        )
    except Exception as e:
        lbl_cg_preview.config(text=f"[공차 기반 미리보기] 오류: {e}")


# ---------- 분석 실행 ----------
def run_analysis():
    root.config(cursor="watch");
    root.update_idletasks()
    try:
        global _save_data

        # 메모리 관리: 이전 그래프 닫기 (명시적)
        old1 = _save_data.get("fig1")
        old2 = _save_data.get("fig2")
        if old1 is not None:
            plt.close(old1)
        if old2 is not None:
            plt.close(old2)

        run_t1 = var_run_t1.get();
        run_t2 = var_run_t2.get()
        if var_run_both.get():
            run_t1 = True;
            run_t2 = True
        if not run_t1 and not run_t2:
            messagebox.showwarning("경고", "실행 모드를 선택하세요. (Type 1, Type 2, 또는 둘 다)")
            return

        # 결과창 초기화 + 태그
        text_result.config(state="normal");
        text_result.delete("1.0", tk.END)
        for tg in ["ok", "not_ok", "black_tag", "pass", "fail", "pv_red", "pv_green", "unc_hdr", "title_tag"]:
            try:
                text_result.tag_delete(tg)
            except:
                pass
        text_result.tag_config("ok", foreground="green")
        text_result.tag_config("not_ok", foreground="red")
        text_result.tag_config("black_tag", foreground="black")
        text_result.tag_config("pass", foreground="green")
        text_result.tag_config("fail", foreground="red")
        text_result.tag_config("pv_red", foreground="red")
        text_result.tag_config("pv_green", foreground="green")
        text_result.tag_config("unc_hdr", foreground="#1f4e79", font=("Consolas", 10, "bold"))
        text_result.tag_config("title_tag", foreground="#1c3b6b", font=("Consolas", 12, "bold"))

        # 제목
        report_title = entry_title.get().strip()
        if report_title:
            text_result.insert(tk.END, f"📄 {report_title}\n", "title_tag")  # \n 한 번만

        # ===== 공통 입력 =====
        try:
            RE = float(entry_RE.get());
            RefV = float(entry_RefV.get())
        except ValueError:
            raise ValueError("RE, RefV 등의 기본 입력값이 숫자가 아닙니다.")

        # --- USL/LSL 결정 (모드에 따라) ---
        if var_tol_mode.get() == "tolerance":
            TOL_in = float(entry_TOL.get())
            USL = RefV + TOL_in / 2.0
            LSL = RefV - TOL_in / 2.0
        else:
            USL = float(entry_USL.get());
            LSL = float(entry_LSL.get())
        u_CAL = float(entry_uCAL.get());
        TOL = USL - LSL
        if USL <= LSL:
            messagebox.showerror("입력 오류", "USL > LSL 이어야 합니다.");
            return

        # Type-B 고정 입력
        try:
            u_LIN = float(entry_uLIN.get());
            u_MSREST = float(entry_uMSREST.get())
            u_T = float(entry_uT.get());
            u_STAB = float(entry_uSTAB.get())
            u_OBJ = float(entry_uOBJ.get());
            u_GV = float(entry_uGV.get())
            u_REST = float(entry_uREST.get())
        except ValueError:
            raise ValueError("Type B 불확도 입력값 중 숫자가 아닌 값이 있습니다.")

        # ---- Type 1 값(선택적) ----
        X = np.array(parse_measurement_input(text_X))
        if run_t1:
            if len(X) < 1:
                raise ValueError("Type 1 실행에는 Type I 값이 필요합니다.")
            if len(X) < 2:
                messagebox.showwarning("입력 경고", "Type I 값은 2개 이상이 권장됩니다.")
        # Type 2 단독을 허용하기 위해 기본값 처리
        if len(X) >= 1:
            n_X = len(X);
            mu_repeat = np.mean(X)
            sigma_repeat = np.std(X, ddof=1) if n_X >= 2 else 0.0
            Abias = mu_repeat - RefV;
            Abias2 = abs(Abias) * 2
            u_BI = calc_u_distribution(Abias2, combo_BI.get())
            u_EVR = sigma_repeat
        else:
            n_X = 0;
            mu_repeat = RefV;
            sigma_repeat = 0.0
            Abias = 0.0;
            u_BI = 0.0;
            u_EVR = 0.0
        u_RE = calc_u_distribution(RE, combo_RE.get())
        u_EV = max(u_EVR, u_RE)

        # Type I 합성불확도 (Type 2 단독 시 참고용)
        u_MS = np.sqrt(
            u_CAL ** 2 + u_LIN ** 2 + u_BI ** 2 + u_EV ** 2 + u_MSREST ** 2 +
            u_T ** 2 + u_STAB ** 2 + u_OBJ ** 2 + u_GV ** 2 + u_REST ** 2
        )
        U_MS = 2 * u_MS

        # 안전 초기화
        u_EVO = u_AV = u_IA = 0.0
        u_MP = np.nan;
        U_MP = np.nan;
        Q_MP = np.nan;
        C_MP = np.nan

        # --- Type1 ---
        if run_t1:
            Cg = (0.2 * TOL) / (6 * sigma_repeat) if sigma_repeat > 0 else np.nan
            Cgk = (0.1 * TOL - abs(mu_repeat - RefV)) / (3 * sigma_repeat) if sigma_repeat > 0 else np.nan
            C_MS = 0.3 * TOL / (6 * u_MS) if u_MS > 0 else np.nan
            Q_MS = 2 * U_MS / TOL * 100 if TOL > 0 else np.nan
            percent_EV = (6 * sigma_repeat / TOL) * 100 if TOL > 0 else np.nan
            RE_percent = (RE / TOL) * 100 if TOL > 0 else np.nan

            cg_ok = (not pd.isna(Cg)) and (Cg >= 1.33)
            cgk_ok = (not pd.isna(Cgk)) and (Cgk >= 1.33)
            cms_ok = (not pd.isna(C_MS)) and (C_MS >= 1.33)
            qms_ok = (not pd.isna(Q_MS)) and (Q_MS <= 15)
            ev_ok = (not pd.isna(percent_EV)) and (percent_EV <= 30)
            re_ok = (not pd.isna(RE_percent)) and (RE_percent <= 5)

            se = sigma_repeat / np.sqrt(n_X) if (n_X > 0 and sigma_repeat > 0) else np.nan
            if (not pd.isna(se)) and (se > 0):
                t_val = (mu_repeat - RefV) / se
                p_val = 2 * (1 - t_dist.cdf(abs(t_val), df=n_X - 1))
                ci_low = mu_repeat - RefV - t_dist.ppf(0.975, df=n_X - 1) * se
                ci_high = mu_repeat - RefV + t_dist.ppf(0.975, df=n_X - 1) * se
            else:
                t_val = p_val = ci_low = ci_high = np.nan

            text_result.insert(tk.END, "[Type 1 Gauge Capability 및 Bias t-test 결과]\n", "black_tag")

            def gauge_line(label, val, ok):
                if pd.isna(val):
                    text_result.insert(tk.END, f"{label:<28} {'NaN':>8}\n")
                else:
                    text_result.insert(tk.END, f"{label:<28} {val:>8.5f}   {'OK' if ok else 'NG'}\n",
                                       "ok" if ok else "not_ok")

            gauge_line("Cg (반복성 1.33 이상)", Cg, cg_ok)
            gauge_line("Cgk (1.33 이상)", Cgk, cgk_ok)
            gauge_line("C_MS (1.33 이상)", C_MS, cms_ok)
            gauge_line("Q_MS (15% 이내)", Q_MS, qms_ok)
            gauge_line("%EV (30% 이내)", percent_EV, ev_ok)
            gauge_line("RE (%) (5% 이내)", RE_percent, re_ok)

            text_result.insert(tk.END, "[치우침(Bias) t-검정 결과]\n")
            text_result.insert(tk.END, f"치우침              {Abias: .5f}\n")
            text_result.insert(tk.END, f"표준오차            {se if not pd.isna(se) else float('nan'): .5f}\n")
            text_result.insert(tk.END, f"t 통계량            {t_val if not pd.isna(t_val) else float('nan'): .3f}\n")
            if not pd.isna(p_val):
                tag = "pv_red" if p_val <= 0.05 else "pv_green"
                text_result.insert(tk.END, f"p-value             {p_val: .5f}\n", tag)
            else:
                text_result.insert(tk.END, "p-value             NaN\n")
            text_result.insert(tk.END, f"95% CI 하한         {ci_low if not pd.isna(ci_low) else float('nan'): .5f}\n")
            text_result.insert(tk.END, f"95% CI 상한         {ci_high if not pd.isna(ci_high) else float('nan'): .5f}\n")

        # === Type II ===
        ar = None;
        df_gage = None;
        fig2 = None
        if run_t2:
            file_path = entry_file.get().strip()
            if not file_path:
                messagebox.showwarning("경고", "Type 2 실행에는 CSV 또는 Excel 파일이 필요합니다.")
                return

            df = safe_load_df(file_path)

            # 카테고리형 변환 (groupby(observed=True) 대응)
            df["부품"] = df["부품"].astype("category").cat.remove_unused_categories()
            df["측정자"] = df["측정자"].astype("category").cat.remove_unused_categories()

            n_o = df["측정자"].nunique();
            n_p = df["부품"].nunique()
            counts = df.groupby(["부품", "측정자"], observed=True).size()
            if counts.empty:
                raise ValueError("데이터 그룹화 결과가 비어있습니다.")
            n_r = counts.min()

            if counts.min() < 2:
                messagebox.showwarning("설계 경고", "반복수(부품×측정자) 2 미만 구간이 있습니다.")
            if counts.min() != counts.max():
                messagebox.showwarning("설계 경고",
                                       f"불균형 데이터 감지: 반복수 min={int(counts.min())}, max={int(counts.max())}")

            if n_o == 1:
                model1 = smf.ols("측정값 ~ C(부품)", data=df).fit()
                ar = sm.stats.anova_lm(model1, typ=2)
                ar["mean_sq"] = ar["sum_sq"] / ar["df"]
                ar.loc["총변동"] = [ar["sum_sq"].sum(), ar["df"].sum(), "", "", ""]
                ar.rename(index={"C(부품)": "부품", "Residual": "반복성"}, inplace=True)
                ar.rename(columns={
                    "df": "자유도(DF)", "sum_sq": "제곱합(SS)", "mean_sq": "평균제곱(MS)",
                    "F": "F값", "PR(>F)": "P값"
                }, inplace=True)
                ar = ar.reindex(["부품", "반복성", "총변동"])[["자유도(DF)", "제곱합(SS)", "평균제곱(MS)", "F값", "P값"]]

                text_result.insert(tk.END, "=== (측정자=1, 자동측정) Gage R&R Study3 ===\n")
                text_result.insert(tk.END, format_anova(ar))

                MS_e = ar.loc["반복성", "평균제곱(MS)"]
                MS_p = ar.loc["부품", "평균제곱(MS)"]
                sd_repeat = np.sqrt(MS_e);
                sd_gagerr = sd_repeat
                sd_part = np.sqrt(max((MS_p - MS_e) / n_r, 0))
                sd_total = np.sqrt(sd_gagerr ** 2 + sd_part ** 2)

                gage_labels = ["총 Gage R&R", "반복성", "부품-대-부품", "총 변동"]
                gage_values = [sd_gagerr, sd_repeat, sd_part, sd_total]

                six_sd_repeat = 6 * sd_repeat;
                six_sd_gagerr = 6 * sd_gagerr
                six_sd_part = 6 * sd_part;
                six_sd_total = 6 * sd_total

                pct_sv_repeat = 100 * six_sd_repeat / six_sd_total
                pct_sv_gagerr = 100 * six_sd_gagerr / six_sd_total
                pct_sv_part = 100 * six_sd_part / six_sd_total

                pct_tol_repeat = 100 * six_sd_repeat / TOL
                pct_tol_gagerr = 100 * six_sd_gagerr / TOL
                pct_tol_part = 100 * six_sd_part / TOL
                pct_tol_total = 100 * six_sd_total / TOL

                rows = []
                rows.append(
                    {"출처": "총 Gage R&R", "표준 편차(SD)": f"{sd_gagerr:.7f}", "연구 변동\n(6×SD)": f"{six_sd_gagerr:.6f}",
                     "%연구 변동\n(%SV)": f"{pct_sv_gagerr:6.2f}", "%공차\n(SV/공차)": f"{pct_tol_gagerr:6.2f}"})
                rows.append({"출처": "  반복성", "표준 편차(SD)": f"{sd_repeat:.7f}", "연구 변동\n(6×SD)": f"{six_sd_repeat:.6f}",
                             "%연구 변동\n(%SV)": f"{pct_sv_repeat:6.2f}", "%공차\n(SV/공차)": f"{pct_tol_repeat:6.2f}"})
                rows.append({"출처": "부품-대-부품", "표준 편차(SD)": f"{sd_part:.7f}", "연구 변동\n(6×SD)": f"{six_sd_part:.6f}",
                             "%연구 변동\n(%SV)": f"{pct_sv_part:6.2f}", "%공차\n(SV/공차)": f"{pct_tol_part:6.2f}"})
                rows.append({"출처": "총 변동", "표준 편차(SD)": f"{sd_total:.7f}", "연구 변동\n(6×SD)": f"{six_sd_total:.6f}",
                             "%연구 변동\n(%SV)": "100.00", "%공차\n(SV/공차)": f"{pct_tol_total:6.2f}"})
                df_gage = pd.DataFrame(rows)

                # ---- 합성불확도(측정프로세스) & 메트릭 ----
                u_EVO = sd_repeat
                u_AV = 0.0
                u_IA = 0.0
                u_EV = max(u_EVR, u_RE, u_EVO)
                unc_values = [u_CAL, u_BI, u_EV, u_AV, u_IA, u_LIN, u_MSREST, u_T, u_STAB, u_OBJ, u_GV, u_REST]
                u_MP = np.sqrt(sum(v ** 2 for v in unc_values))
                U_MP = 2 * u_MP
                Q_MP = 2 * U_MP / TOL * 100 if TOL > 0 else np.nan
                C_MP = (0.4 * TOL) / (2 * U_MP) if U_MP > 0 else np.nan

                # ---- 텍스트 표 + 메트릭/판정 ----
                append_gage_table_to_text(df_gage)
                ndc = (1.41 * sd_part / sd_gagerr) if sd_gagerr > 0 else np.nan
                append_metrics_and_judgement(ndc, sd_part, sd_gagerr, Q_MP, C_MP,
                                             pct_sv_gagerr, pct_tol_gagerr)

            else:
                # 교호작용 가능
                model_full = smf.ols("측정값 ~ C(부품)*C(측정자)", data=df).fit()
                af = sm.stats.anova_lm(model_full, typ=2)
                af["mean_sq"] = af["sum_sq"] / af["df"]
                af.loc["총변동"] = [af["sum_sq"].sum(), af["df"].sum(), "", "", ""]
                af.rename(index={"C(부품)": "부품", "C(측정자)": "측정자", "C(부품):C(측정자)": "부품:측정자", "Residual": "반복성"},
                          inplace=True)
                af.rename(columns={
                    "df": "자유도(DF)", "sum_sq": "제곱합(SS)", "mean_sq": "평균제곱(MS)",
                    "F": "F값", "PR(>F)": "P값"
                }, inplace=True)

                MS_i = af.loc["부품:측정자", "평균제곱(MS)"]
                DF_i = af.loc["부품:측정자", "자유도(DF)"]
                MS_r_full = af.loc["반복성", "평균제곱(MS)"]
                DF_r = af.loc["반복성", "자유도(DF)"]

                for idx in ["부품", "측정자"]:
                    MS_f = af.loc[idx, "평균제곱(MS)"]
                    DF_f = af.loc[idx, "자유도(DF)"]
                    F_val = MS_f / MS_i if MS_i > 0 else np.nan
                    P_val = 1 - f.cdf(F_val, DF_f, DF_i) if MS_i > 0 else np.nan
                    af.loc[idx, "F값"] = F_val;
                    af.loc[idx, "P값"] = P_val

                F_int = MS_i / MS_r_full if MS_r_full > 0 else np.nan
                P_int = 1 - f.cdf(F_int, DF_i, DF_r) if MS_r_full > 0 else np.nan
                af.loc["부품:측정자", "F값"] = F_int;
                af.loc["부품:측정자", "P값"] = P_int
                af.loc["반복성", ["F값", "P값"]] = ["", ""]

                af = af.reindex(["부품", "측정자", "부품:측정자", "반복성", "총변동"]) \
                    [["자유도(DF)", "제곱합(SS)", "평균제곱(MS)", "F값", "P값"]]

                p_inter = af.loc["부품:측정자", "P값"]
                is_interaction_significant = (p_inter != "") and (not pd.isna(p_inter)) and (float(p_inter) <= 0.05)

                if is_interaction_significant:
                    text_result.insert(tk.END, "=== 교호작용 포함 이원분산분석표 Gage R&R Study2 ===\n")
                    text_result.insert(tk.END, format_anova(af))
                    ar = af.copy()
                else:
                    model_red = smf.ols("측정값 ~ C(부품)+C(측정자)", data=df).fit()
                    ar = sm.stats.anova_lm(model_red, typ=2)
                    ar["mean_sq"] = ar["sum_sq"] / ar["df"]
                    ar.loc["총변동"] = [ar["sum_sq"].sum(), ar["df"].sum(), "", "", ""]
                    ar.rename(index={"C(부품)": "부품", "C(측정자)": "측정자", "Residual": "반복성"}, inplace=True)
                    ar.rename(columns={
                        "df": "자유도(DF)", "sum_sq": "제곱합(SS)", "mean_sq": "평균제곱(MS)",
                        "F": "F값", "PR(>F)": "P값"
                    }, inplace=True)
                    ar = ar.reindex(["부품", "측정자", "반복성", "총변동"]) \
                        [["자유도(DF)", "제곱합(SS)", "평균제곱(MS)", "F값", "P값"]]
                    text_result.insert(tk.END, "=== 교호작용 없는 이원분산분석표 Gage R&R Study2 ===\n")
                    text_result.insert(tk.END, format_anova(ar))

                MS_e = ar.loc["반복성", "평균제곱(MS)"]
                MS_p = ar.loc["부품", "평균제곱(MS)"]
                MS_o = ar.loc["측정자", "평균제곱(MS)"] if "측정자" in ar.index else 0.0

                if is_interaction_significant and "부품:측정자" in ar.index:
                    MS_io = ar.loc["부품:측정자", "평균제곱(MS)"]
                    var_repeat = MS_e
                    var_operator = max((MS_o - MS_io) / (n_p * n_r), 0)
                    var_interaction = max((MS_io - MS_e) / n_r, 0)
                    var_part = max((MS_p - MS_io) / (n_o * n_r), 0)
                else:
                    MS_io = 0.0
                    var_repeat = MS_e
                    var_operator = max((MS_o - MS_e) / (n_p * n_r), 0)
                    var_interaction = 0.0
                    var_part = max((MS_p - MS_e) / (n_o * n_r), 0)

                sd_repeat = np.sqrt(var_repeat)
                sd_operator = np.sqrt(var_operator)
                sd_interaction = np.sqrt(var_interaction)
                sd_reprod = np.sqrt(sd_operator ** 2 + sd_interaction ** 2)
                sd_gagerr = np.sqrt(sd_repeat ** 2 + sd_reprod ** 2)
                sd_part = np.sqrt(var_part)
                sd_total = np.sqrt(sd_gagerr ** 2 + sd_part ** 2)

                gage_labels = ["총 Gage R&R", "반복성", "재현성"]
                gage_values = [sd_gagerr, sd_repeat, sd_reprod]
                if is_interaction_significant:
                    gage_labels += ["  측정자", "  교호작용"];
                    gage_values += [sd_operator, sd_interaction]
                gage_labels += ["부품-대-부품", "총 변동"];
                gage_values += [sd_part, sd_total]

                six_sd_repeat = 6 * sd_repeat
                six_sd_operator = 6 * sd_operator
                six_sd_interaction = 6 * sd_interaction
                six_sd_reprod = 6 * sd_reprod
                six_sd_gagerr = 6 * sd_gagerr
                six_sd_part = 6 * sd_part
                six_sd_total = 6 * sd_total

                pct_sv_repeat = 100 * six_sd_repeat / six_sd_total
                pct_sv_operator = 100 * six_sd_operator / six_sd_total if sd_operator > 0 else 0.0
                pct_sv_interaction = 100 * six_sd_interaction / six_sd_total if sd_interaction > 0 else 0.0
                pct_sv_reprod = 100 * six_sd_reprod / six_sd_total
                pct_sv_gagerr = 100 * six_sd_gagerr / six_sd_total
                pct_sv_part = 100 * six_sd_part / six_sd_total

                pct_tol_repeat = 100 * six_sd_repeat / TOL
                pct_tol_operator = 100 * six_sd_operator / TOL if sd_operator > 0 else 0.0
                pct_tol_interaction = 100 * six_sd_interaction / TOL if sd_interaction > 0 else 0.0
                pct_tol_reprod = 100 * six_sd_reprod / TOL
                pct_tol_gagerr = 100 * six_sd_gagerr / TOL
                pct_tol_part = 100 * six_sd_part / TOL
                pct_tol_total = 100 * six_sd_total / TOL

                rows = []
                rows.append(
                    {"출처": "총 Gage R&R", "표준 편차(SD)": f"{sd_gagerr:.7f}", "연구 변동\n(6×SD)": f"{six_sd_gagerr:.6f}",
                     "%연구 변동\n(%SV)": f"{pct_sv_gagerr:6.2f}", "%공차\n(SV/공차)": f"{pct_tol_gagerr:6.2f}"})
                rows.append({"출처": "  반복성", "표준 편차(SD)": f"{sd_repeat:.7f}", "연구 변동\n(6×SD)": f"{six_sd_repeat:.6f}",
                             "%연구 변동\n(%SV)": f"{pct_sv_repeat:6.2f}", "%공차\n(SV/공차)": f"{pct_tol_repeat:6.2f}"})
                rows.append({"출처": "  재현성", "표준 편차(SD)": f"{sd_reprod:.7f}", "연구 변동\n(6×SD)": f"{six_sd_reprod:.6f}",
                             "%연구 변동\n(%SV)": f"{pct_sv_reprod:6.2f}", "%공차\n(SV/공차)": f"{pct_tol_reprod:6.2f}"})
                if is_interaction_significant:
                    rows.append(
                        {"출처": "    측정자", "표준 편차(SD)": f"{sd_operator:.7f}", "연구 변동\n(6×SD)": f"{six_sd_operator:.6f}",
                         "%연구 변동\n(%SV)": f"{pct_sv_operator:6.2f}", "%공차\n(SV/공차)": f"{pct_tol_operator:6.2f}"})
                    rows.append({"출처": "    교호작용", "표준 편차(SD)": f"{sd_interaction:.7f}",
                                 "연구 변동\n(6×SD)": f"{six_sd_interaction:.6f}",
                                 "%연구 변동\n(%SV)": f"{pct_sv_interaction:6.2f}",
                                 "%공차\n(SV/공차)": f"{pct_tol_interaction:6.2f}"})
                rows.append({"출처": "부품-대-부품", "표준 편차(SD)": f"{sd_part:.7f}", "연구 변동\n(6×SD)": f"{six_sd_part:.6f}",
                             "%연구 변동\n(%SV)": f"{pct_sv_part:6.2f}", "%공차\n(SV/공차)": f"{pct_tol_part:6.2f}"})
                rows.append({"출처": "총 변동", "표준 편차(SD)": f"{sd_total:.7f}", "연구 변동\n(6×SD)": f"{six_sd_total:.6f}",
                             "%연구 변동\n(%SV)": "100.00", "%공차\n(SV/공차)": f"{pct_tol_total:6.2f}"})
                df_gage = pd.DataFrame(rows)

                # ---- 합성불확도(측정프로세스) & 메트릭 ----
                if is_interaction_significant:
                    MS_i = ar.loc["부품:측정자", "평균제곱(MS)"] if "부품:측정자" in ar.index else 0.0
                    u_IA = np.sqrt(max(MS_i - MS_e, 0) / (n_p * n_r))
                else:
                    u_IA = 0.0
                u_EVO = np.sqrt(MS_e)
                u_AV = np.sqrt(max(MS_o - MS_e, 0) / (n_p * n_r))
                u_EV = max(u_EVR, u_RE, u_EVO)
                unc_values = [u_CAL, u_BI, u_EVO, u_AV, u_IA, u_LIN, u_MSREST, u_T, u_STAB, u_OBJ, u_GV, u_REST]
                u_MP = np.sqrt(sum(v ** 2 for v in unc_values if abs(v) > EPS))
                U_MP = 2 * u_MP
                Q_MP = 2 * U_MP / TOL * 100 if TOL > 0 else np.nan
                C_MP = (0.4 * TOL) / (2 * U_MP) if U_MP > 0 else np.nan

                # ---- 텍스트 표 + 메트릭/판정 ----
                append_gage_table_to_text(df_gage)
                ndc = (1.41 * sd_part / sd_gagerr) if sd_gagerr > 0 else np.nan
                append_metrics_and_judgement(ndc, sd_part, sd_gagerr, Q_MP, C_MP,
                                             pct_sv_gagerr, pct_tol_gagerr)

        # ----------- 불확도 탭 요약 -----------
        if run_t2:
            unc_dict = {"u_CAL": u_CAL, "u_BI": u_BI, "u_EVR": u_EVR, "u_RE": u_RE, "u_EVO": u_EVO,
                        "u_AV": u_AV, "u_IA": u_IA, "u_LIN": u_LIN, "u_MSREST": u_MSREST,
                        "u_T": u_T, "u_STAB": u_STAB, "u_OBJ": u_OBJ, "u_GV": u_GV, "u_REST": u_REST}
            u_MP_show = u_MP
        else:
            unc_dict = {"u_CAL": u_CAL, "u_BI": u_BI, "u_EVR": u_EVR, "u_RE": u_RE, "u_EV": max(u_EVR, u_RE),
                        "u_LIN": u_LIN, "u_MSREST": u_MSREST, "u_T": u_T, "u_STAB": u_STAB,
                        "u_OBJ": u_OBJ, "u_GV": u_GV, "u_REST": u_REST}
            u_MP_show = u_MS

        nz_items = [(k, v) for k, v in unc_dict.items() if abs(v) > EPS]
        nz_items.sort(key=lambda x: x[1], reverse=True)
        populate_uncertainty_tab(nz_items, u_MP_show)
        text_result.insert(tk.END, "[불확도 구성 (0 값 제외) → 우측 탭에서 표 확인]\n", "unc_hdr")
        text_result.insert(tk.END, f"합성 표준불확도 u = {u_MP_show:.6f}\n")
        text_result.insert(tk.END, f"확장 불확도(95%) U = {2 * u_MP_show:.6f}\n")

        # --- 그래프(불확도) ---
        for w in frame_graph1.winfo_children(): w.destroy()
        nz_labels = [k for k, _ in nz_items];
        nz_values = [v for _, v in nz_items]
        if nz_labels:
            fig1, ax1 = plt.subplots(figsize=(6.8, 8))
            bars = ax1.barh(nz_labels, nz_values, color=get_colors(len(nz_labels)))
            ax1.set_xlabel("불확도 값");
            ax1.set_title("불확도 구성요소 (값 & 주석)")
            ax1.xaxis.set_major_formatter(FormatStrFormatter('%.6f'))
            ax1.grid(axis="x", linestyle=":", alpha=0.3)
            max_val = max(nz_values);
            ax1.set_xlim(0, max_val * 1.25)
            note_x_axes = 0.5 - NOTE_OFFSET_FRAC
            for bar, lab, val in zip(bars, nz_labels, nz_values):
                y = bar.get_y() + bar.get_height() / 2
                value_x = (val * 0.04) if val >= max_val * 0.18 else (val + max(max_val * 0.01, 1e-6))
                ax1.text(value_x, y, f"{val:.6f}", va='center', ha='left', fontsize=9)
                note = UNC_NOTES.get(lab, "")
                if note:
                    ax1.text(note_x_axes, y, note, transform=ax1.get_yaxis_transform(),
                             va='center', ha='left', fontsize=9, color='#333', clip_on=False)
            fig1.tight_layout()
        else:
            fig1 = plt.figure(figsize=(6.2, 2.4));
            plt.title("표시할 불확도 막대가 없습니다(모두 0).\n");
            plt.axis('off')
        canvas1 = FigureCanvasTkAgg(fig1, master=frame_graph1);
        canvas1.draw();
        canvas1.get_tk_widget().pack(fill="both", expand=True)

        # --- 그래프(Gage R&R) ---
        for w in frame_graph2.winfo_children(): w.destroy()
        if run_t2 and 'gage_labels' in locals():
            fig2, ax2 = plt.subplots(figsize=(5.8, 4.2))
            bars2 = ax2.bar(gage_labels, gage_values, color=get_colors(len(gage_labels)))
            ax2.set_ylabel("Std. Dev (σ)");
            ax2.set_title("Gage R&R 및 변동 요소")
            ax2.tick_params(axis='x', labelrotation=20);
            ax2.grid(axis="y", linestyle=":", alpha=0.3)
            for bar, val in zip(bars2, gage_values):
                if val > 0:
                    ax2.text(bar.get_x() + bar.get_width() / 2, val * 1.01, f"{val:.5f}", ha='center', va='bottom',
                             fontsize=9)
            ax2.margins(x=0.05);
            fig2.tight_layout()
            canvas2 = FigureCanvasTkAgg(fig2, master=frame_graph2);
            canvas2.draw();
            canvas2.get_tk_widget().pack(fill="both", expand=True)
        else:
            ttk.Label(frame_graph2, text="Type 2를 실행하면 그래프가 표시됩니다.").pack(fill="both", expand=True, pady=20)

        _save_data = {
            "anova_df": ar.copy() if isinstance(ar, pd.DataFrame) else None,
            "gage_df": df_gage.copy() if isinstance(df_gage, pd.DataFrame) else None,
            "uncertainty_dict": {k: v for k, v in nz_items},
            "u_mp": float(u_MP_show),
            "fig1": fig1,
            "fig2": fig2
        }

        # === (신규) 행간 공백 압축: 태그 보존 ===
        _compress_blank_lines_inplace(text_result)

    except Exception as e:
        messagebox.showerror("오류", f"{e}\n\n{traceback.format_exc()}")
    finally:
        root.config(cursor="")


# ---------- 캡처 ----------
def capture_window_printwindow(hwnd):
    if platform.system() != "Windows":
        return None
    try:
        import ctypes
        from ctypes import windll, wintypes, byref

        user32 = windll.user32
        gdi32 = windll.gdi32

        rect = wintypes.RECT()
        user32.GetWindowRect(hwnd, byref(rect))
        left, top, right, bottom = rect.left, rect.top, rect.right, rect.bottom
        width, height = right - left, bottom - top
        if width <= 0 or height <= 0:
            return None

        hwindc = user32.GetWindowDC(hwnd)
        srcdc = gdi32.CreateCompatibleDC(hwindc)
        bmp = gdi32.CreateCompatibleBitmap(hwindc, width, height)
        gdi32.SelectObject(srcdc, bmp)

        res = user32.PrintWindow(hwnd, srcdc, 2)
        if res != 1:
            gdi32.DeleteObject(bmp)
            gdi32.DeleteDC(srcdc)
            user32.ReleaseDC(hwnd, hwindc)
            return None

        class BITMAPINFOHEADER(ctypes.Structure):
            _fields_ = [
                ("biSize", wintypes.DWORD),
                ("biWidth", ctypes.c_long),
                ("biHeight", ctypes.c_long),
                ("biPlanes", wintypes.WORD),
                ("biBitCount", wintypes.WORD),
                ("biCompression", wintypes.DWORD),
                ("biSizeImage", wintypes.DWORD),
                ("biXPelsPerMeter", ctypes.c_long),
                ("biYPelsPerMeter", ctypes.c_long),
                ("biClrUsed", wintypes.DWORD),
                ("biClrImportant", wintypes.DWORD)
            ]

        bmi = BITMAPINFOHEADER()
        bmi.biSize = ctypes.sizeof(BITMAPINFOHEADER)
        bmi.biWidth = width
        bmi.biHeight = -height
        bmi.biPlanes = 1
        bmi.biBitCount = 32
        bmi.biCompression = 0

        buf_len = width * height * 4
        buffer = (ctypes.c_char * buf_len)()
        bits = gdi32.GetDIBits(srcdc, bmp, 0, height, buffer, ctypes.byref(bmi), 0)
        gdi32.DeleteObject(bmp)
        gdi32.DeleteDC(srcdc)
        user32.ReleaseDC(hwnd, hwindc)

        if bits == 0:
            return None

        if Image is None:
            return None
        img = Image.frombuffer('RGB', (width, height), buffer, 'raw', 'BGRA', 0, 1)
        return img
    except Exception:
        return None


def capture_and_save():
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".png",
                                                 filetypes=[("PNG Image", "*.png")],
                                                 title="화면 캡처 저장")
        if not file_path:
            return

        root.update_idletasks()
        root.lift();
        root.attributes('-topmost', True);
        root.update();
        time.sleep(0.15)

        hwnd = root.winfo_id()
        image = capture_window_printwindow(hwnd)

        if image is None and ImageGrab is not None:
            try:
                x = root.winfo_rootx();
                y = root.winfo_rooty()
                w = root.winfo_width();
                h = root.winfo_height()
                bbox = (x, y, x + w, y + h)
                image = ImageGrab.grab(bbox=bbox)
            except Exception:
                image = None

        if image is None and pyautogui is not None:
            x = root.winfo_rootx();
            y = root.winfo_rooty()
            w = root.winfo_width();
            h = root.winfo_height()
            image = pyautogui.screenshot(region=(x, y, w, h))

        root.attributes('-topmost', False)

        if image is None:
            raise RuntimeError("스크린샷 도구를 사용할 수 없습니다.")
        image.save(file_path)
        messagebox.showinfo("저장 완료", f"화면이 저장되었습니다:\n{file_path}")

    except Exception as e:
        root.attributes('-topmost', False)
        messagebox.showerror("오류", f"화면 캡처 중 오류가 발생했습니다.\n{e}\n\n{traceback.format_exc()}")


# ---------- 결과 저장(엑셀) ----------
def save_analysis_results(anova_df, gage_df, uncertainty_dict, u_mp, fig1, fig2):
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel files", ".xlsx")],
                                             title="분석 결과 저장")
    if not file_path:
        return
    try:
        wb = Workbook()
        ws_meta = wb.active;
        ws_meta.title = "메타"
        ws_meta.append(["제목", entry_title.get().strip()])
        ws_meta.append(["파일", entry_file.get().strip()])

        ws1 = wb.create_sheet(title="분산분석표(Type2)")
        if isinstance(anova_df, pd.DataFrame):
            for r in dataframe_to_rows(anova_df, index=True, header=True):
                ws1.append(r)
        else:
            ws1.append(["Type 2 미실행 또는 결과 없음"])

        ws2 = wb.create_sheet(title="Gage평가(Type2)")
        if isinstance(gage_df, pd.DataFrame):
            for r in dataframe_to_rows(gage_df, index=False, header=True):
                ws2.append(r)
        else:
            ws2.append(["Type 2 미실행 또는 결과 없음"])

        ws3 = wb.create_sheet(title="불확도구성(0제외)")
        ws3.append(["항목", "값", "비중(%)", "주석"])
        if uncertainty_dict:
            for k, v, pct, note in build_unc_rows(list(uncertainty_dict.items()), u_mp):
                ws3.append([k, v, pct, note])
        else:
            ws3.append(["표시할 항목 없음", "", "", ""])

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            try:
                for col in ws.columns:
                    values = [str(c.value) if c.value is not None else "" for c in col][:200]
                    width = min(42, max((len(s) for s in values), default=10) + 2)
                    ws.column_dimensions[col[0].column_letter].width = width
            except Exception:
                pass

        wb.save(file_path)
        base_path = os.path.splitext(file_path)[0]
        if fig1 is not None:
            fig1.savefig(base_path + "_uncertainty.png", dpi=200, bbox_inches="tight")
        if fig2 is not None:
            fig2.savefig(base_path + "_gageRR.png", dpi=200, bbox_inches="tight")
        messagebox.showinfo("저장 완료", f"엑셀/이미지 파일 저장 완료:\n{file_path}")

    except Exception as e:
        messagebox.showerror("오류", f"저장 중 오류가 발생했습니다.\n{e}\n\n{traceback.format_exc()}")


def save_results_btn():
    if not _save_data:
        messagebox.showwarning("경고", "먼저 분석을 실행하세요!")
        return
    d = _save_data
    save_analysis_results(
        d["anova_df"], d["gage_df"], d["uncertainty_dict"], d["u_mp"], d["fig1"], d["fig2"]
    )


# ---------- 설정 JSON ----------
def save_settings_json():
    try:
        data = {
            "Type1": {
                "RE": entry_RE.get(), "RefV": entry_RefV.get(),
                "USL": entry_USL.get(), "LSL": entry_LSL.get(),
                "TOL": entry_TOL.get(), "TolMode": var_tol_mode.get(),
                "u_CAL": entry_uCAL.get(),
                "BiasDist": combo_BI.get(),
                "uRE_Dist": combo_RE.get(),
                "X_values": text_X.get("1.0", tk.END).strip()
            },
            "TypeB": {
                "u_LIN": entry_uLIN.get(), "u_MSREST": entry_uMSREST.get(),
                "u_T": entry_uT.get(), "u_STAB": entry_uSTAB.get(),
                "u_OBJ": entry_uOBJ.get(), "u_GV": entry_uGV.get(),
                "u_REST": entry_uREST.get()
            },
            "Type2": {
                "file_path": entry_file.get(),
                "title": entry_title.get()
            },
            "RunMode": {"t1": var_run_t1.get(), "t2": var_run_t2.get(), "both": var_run_both.get()}
        }
        fp = filedialog.asksaveasfilename(defaultextension=".json",
                                          filetypes=[("JSON files", "*.json")],
                                          title="데이터 저장(JSON)")
        if not fp: return
        with open(fp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        messagebox.showinfo("저장 완료", f"설정이 저장되었습니다.\n{fp}")
    except Exception as e:
        messagebox.showerror("오류", f"설정 저장 중 오류\n{e}\n\n{traceback.format_exc()}")


def load_settings_json():
    try:
        fp = filedialog.askopenfilename(defaultextension=".json",
                                        filetypes=[("JSON files", "*.json")],
                                        title="데이터 불러오기(JSON)")
        # ... (이하 동일)
        if not fp: return
        with open(fp, "r", encoding="utf-8") as f:
            data = json.load(f)
        t1 = data.get("Type1", {})
        entry_RE.delete(0, tk.END);
        entry_RE.insert(0, t1.get("RE", ""))
        entry_RefV.delete(0, tk.END);
        entry_RefV.insert(0, t1.get("RefV", ""))
        entry_USL.delete(0, tk.END);
        entry_USL.insert(0, t1.get("USL", ""))
        entry_LSL.delete(0, tk.END);
        entry_LSL.insert(0, t1.get("LSL", ""))
        entry_TOL.delete(0, tk.END);
        entry_TOL.insert(0, t1.get("TOL", ""))
        entry_uCAL.delete(0, tk.END);
        entry_uCAL.insert(0, t1.get("u_CAL", ""))
        if "BiasDist" in t1: combo_BI.set(t1["BiasDist"])
        if "uRE_Dist" in t1: combo_RE.set(t1["uRE_Dist"])
        text_X.delete("1.0", tk.END);
        text_X.insert(tk.END, t1.get("X_values", ""))

        tb = data.get("TypeB", {})
        for ent, key in zip(
                [entry_uLIN, entry_uMSREST, entry_uT, entry_uSTAB, entry_uOBJ, entry_uGV, entry_uREST],
                ["u_LIN", "u_MSREST", "u_T", "u_STAB", "u_OBJ", "u_GV", "u_REST"]
        ):
            ent.delete(0, tk.END);
            ent.insert(0, tb.get(key, ""))

        t2 = data.get("Type2", {})
        entry_file.delete(0, tk.END);
        entry_file.insert(0, t2.get("file_path", ""))
        entry_title.delete(0, tk.END);
        entry_title.insert(0, t2.get("title", ""))

        rm = data.get("RunMode", {})
        if "t1" in rm: var_run_t1.set(bool(rm["t1"]))
        if "t2" in rm: var_run_t2.set(bool(rm["t2"]))
        if "both" in rm: var_run_both.set(bool(rm["both"]))

        if t1.get("TolMode") in ("usllsl", "tolerance"):
            var_tol_mode.set(t1["TolMode"])
            on_tol_mode_change()

        messagebox.showinfo("불러오기 완료", f"설정을 불러왔습니다.\n{fp}")
    except Exception as e:
        messagebox.showerror("오류", f"설정 불러오기 중 오류\n{e}\n\n{traceback.format_exc()}")


# ---------- UI ----------
try:
    import ctypes

    if platform.system() == "Windows":
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

root = tk.Tk()
root.title("ISO 22514-7 & VDA5 MSA Gage R&R Study1_2_3_김훈희 품질관리기술사 - Rev.07")
root.geometry("1520x990")

style = ttk.Style()
try:
    for theme in ("vista", "clam", "alt"):
        try:
            style.theme_use(theme);
            break
        except Exception:
            continue
except Exception:
    pass
style.configure("Treeview", rowheight=22)
style.configure("Unc.Treeview", font=("맑은 고딕", 10))
style.configure("Unc.Treeview.Heading", font=("맑은 고딕", 10, "bold"))

frame_top = ttk.Frame(root);
frame_top.pack(fill="x", padx=10, pady=5)

# --- Type I ---
frame_t1 = ttk.Labelframe(frame_top, text="Type I 불확도 입력")
frame_t1.grid(row=0, column=0, sticky="nw", padx=5, pady=5)

labels_t1 = ["RE", "RefV", "USL", "LSL", "u_CAL"]
defaults = [0.001, 6.002, 6.03, 5.97, 0.001]
entries = []
for i, (lab, defv) in enumerate(zip(labels_t1, defaults)):
    ttk.Label(frame_t1, text=lab).grid(row=i, column=0, sticky="e", padx=3, pady=2)
    e = ttk.Entry(frame_t1, width=12);
    e.insert(0, str(defv));
    e.grid(row=i, column=1, padx=3, pady=2)
    entries.append(e)
entry_RE, entry_RefV, entry_USL, entry_LSL, entry_uCAL = entries

# 공차범위 입력
ttk.Label(frame_t1, text="TOL(공차폭)").grid(row=0, column=2, sticky="e", padx=3)
entry_TOL = ttk.Entry(frame_t1, width=12)
try:
    entry_TOL.insert(0, f"{float(entry_USL.get()) - float(entry_LSL.get()):.6f}")
except Exception:
    entry_TOL.insert(0, "0.060")
entry_TOL.grid(row=0, column=3, padx=3, pady=2)

# 입력 모드 (USL/LSL vs TOL)
ttk.Label(frame_t1, text="입력 모드").grid(row=1, column=2, sticky="e", padx=3)
var_tol_mode = tk.StringVar(value="usllsl")
rb1 = ttk.Radiobutton(frame_t1, text="USL/LSL 직접입력", value="usllsl", variable=var_tol_mode, command=on_tol_mode_change)
rb2 = ttk.Radiobutton(frame_t1, text="공차범위(TOL) 입력", value="tolerance", variable=var_tol_mode,
                      command=on_tol_mode_change)
rb1.grid(row=1, column=3, sticky="w", padx=3)
rb2.grid(row=2, column=3, sticky="w", padx=3)

# 공차로 Cg/Cgk 계산 버튼 + 미리보기
btn_preview_cg = ttk.Button(frame_t1, text="공차로 Cg/Cgk 계산", command=preview_cg_from_tol)
btn_preview_cg.grid(row=3, column=3, sticky="w", padx=3, pady=2)
lbl_cg_preview = ttk.Label(frame_t1, text="", foreground="#0b5394")
lbl_cg_preview.grid(row=4, column=2, columnspan=2, sticky="w", padx=3, pady=2)

# 분포 선택
ttk.Label(frame_t1, text="Bias 분포").grid(row=5, column=0, sticky="e", padx=3)
combo_BI = ttk.Combobox(frame_t1, values=["균등분포", "정규분포", "삼각분포", "U분포"], width=10, state="readonly");
combo_BI.current(0);
combo_BI.grid(row=5, column=1)

ttk.Label(frame_t1, text="u_RE 분포").grid(row=6, column=0, sticky="e", padx=3)
combo_RE = ttk.Combobox(frame_t1, values=["균등분포", "정규분포", "삼각분포", "U분포"], width=10, state="readonly");
combo_RE.current(0);
combo_RE.grid(row=6, column=1)

# Type I 값 입력창 (Type 2 단독시 비워도 됨)
ttk.Label(frame_t1, text="Type I 값").grid(row=7, column=0, sticky="ne", padx=3)
text_X = tk.Text(frame_t1, height=6, width=72)
text_X.grid(row=7, column=1, columnspan=3, padx=3, pady=2, sticky="w")
text_X.insert(tk.END, "")  # 기본은 비워둠(단독 실행 편의)

# --- Type B 추가 불확도 ---
frame_unc = ttk.Labelframe(frame_top, text="추가 불확도 (Type B)")
frame_unc.grid(row=0, column=1, sticky="nw", padx=5, pady=5)

u_labels = ["u_LIN", "u_MSREST", "u_T", "u_STAB", "u_OBJ", "u_GV", "u_REST"]
u_entries = []
for i, lab in enumerate(u_labels):
    ttk.Label(frame_unc, text=lab).grid(row=i, column=0, sticky="e", padx=3, pady=2)
    ue = ttk.Entry(frame_unc, width=10);
    ue.insert(0, "0");
    ue.grid(row=i, column=1, padx=3, pady=2)
    u_entries.append(ue)
entry_uLIN, entry_uMSREST, entry_uT, entry_uSTAB, entry_uOBJ, entry_uGV, entry_uREST = u_entries

# --- Type II 파일 + 제목 ---
frame_t2 = ttk.Labelframe(frame_top, text="Type II CSV/Excel")
frame_t2.grid(row=0, column=2, sticky="nw", padx=5, pady=5)

entry_file = ttk.Entry(frame_t2, width=80);
entry_file.grid(row=0, column=0, padx=3, pady=2)
ttk.Button(frame_t2, text="파일 선택", command=browse_file).grid(row=0, column=1, padx=3)

ttk.Label(frame_t2, text="제목").grid(row=1, column=0, sticky="w", padx=3)
entry_title = ttk.Entry(frame_t2, width=80);
entry_title.grid(row=1, column=0, columnspan=2, sticky="we", padx=3, pady=2)

# --- 실행 모드 ---
frame_mode = ttk.Labelframe(frame_top, text="실행 모드")
frame_mode.grid(row=0, column=3, sticky="nw", padx=8, pady=5)

# 기본값: Type 2만 켬(단독 실행 기본)
var_run_t1 = tk.BooleanVar(value=False)
var_run_t2 = tk.BooleanVar(value=True)
var_run_both = tk.BooleanVar(value=False)

cb_t1 = ttk.Checkbutton(frame_mode, text="Type 1만 실행", variable=var_run_t1, command=on_toggle_t1_t2)
cb_t2 = ttk.Checkbutton(frame_mode, text="Type 2만 실행", variable=var_run_t2, command=on_toggle_t1_t2)
cb_both = ttk.Checkbutton(frame_mode, text="둘다(동시) 실행", variable=var_run_both, command=on_toggle_both)
cb_t1.grid(row=0, column=0, sticky="w", padx=4, pady=2)
cb_t2.grid(row=1, column=0, sticky="w", padx=4, pady=2)
cb_both.grid(row=2, column=0, sticky="w", padx=4, pady=2)

# --- 버튼 영역 ---
frame_btn = ttk.Frame(root);
frame_btn.pack(fill="x", pady=5)

ttk.Button(frame_btn, text="📊 분석 실행", command=run_analysis).pack(side="left", padx=18)
ttk.Button(frame_btn, text="💾 결과 저장", command=save_results_btn).pack(side="left", padx=8)
ttk.Button(frame_btn, text="🖼️ 화면 캡처 저장", command=capture_and_save).pack(side="left", padx=8)
ttk.Button(frame_btn, text="📋 결과 복사",
           command=lambda: (root.clipboard_clear(), root.clipboard_append(text_result.get("1.0", tk.END)))
           ).pack(side="left", padx=8)
ttk.Button(frame_btn, text="⚙️ 데이터 저장(JSON)", command=save_settings_json).pack(side="left", padx=8)
ttk.Button(frame_btn, text="⚙️ 데이터 불러오기(JSON)", command=load_settings_json).pack(side="left", padx=8)

# --- 아래 Panedwindow ---
frame_bottom = ttk.Panedwindow(root, orient=tk.HORIZONTAL)
frame_bottom.pack(fill="both", expand=True, padx=10, pady=5)

frame_graph1 = ttk.Labelframe(frame_bottom, text="불확도 시각화(0 제외)", width=600)
frame_bottom.add(frame_graph1, weight=1)

frame_graph2 = ttk.Labelframe(frame_bottom, text="Gage R&R 시각화", width=500)
frame_bottom.add(frame_graph2, weight=1)

frame_output = ttk.Labelframe(frame_bottom, text="결과/표시")
frame_bottom.add(frame_output, weight=2)

notebook = ttk.Notebook(frame_output)
tab_result = ttk.Frame(notebook)
tab_uncert = ttk.Frame(notebook)
notebook.add(tab_result, text="분석 결과")
notebook.add(tab_uncert, text="불확도 구성")
notebook.pack(fill="both", expand=True)

# 분석결과(텍스트) 탭
xscrollbar = tk.Scrollbar(tab_result, orient="horizontal")
yscrollbar = tk.Scrollbar(tab_result, orient="vertical")
text_result = tk.Text(tab_result, font=("Consolas", 10), wrap="none",
                      xscrollcommand=xscrollbar.set, yscrollcommand=yscrollbar.set)
text_result.pack(fill="both", expand=True, padx=5, pady=5, side="left")
xscrollbar.config(command=text_result.xview);
xscrollbar.pack(side="bottom", fill="x")
yscrollbar.config(command=text_result.yview);
yscrollbar.pack(side="right", fill="y")

# 불확도 구성 탭(표)
tv_unc = ttk.Treeview(tab_uncert, columns=("항목", "값", "비중(%)", "주석"), show="headings", style="Unc.Treeview")
for col, w, anchor in [("항목", 110, "center"), ("값", 130, "e"), ("비중(%)", 90, "e"), ("주석", 460, "w")]:
    tv_unc.heading(col, text=col, command=lambda c=col: sort_tv_unc(c, reverse=not _tv_unc_sort_state.get(c, False)))
    tv_unc.column(col, width=w, anchor=anchor, stretch=True)

_tv_unc_tags = {
    'odd': {'background': '#fafafa'},
    'even': {'background': '#ffffff'},
    'maxrow': {'background': '#fff6bf'},
    'sumrow': {'font': ('맑은 고딕', 10, 'bold')}
}
for tag, opts in _tv_unc_tags.items():
    tv_unc.tag_configure(tag, **opts)

ys2 = tk.Scrollbar(tab_uncert, orient="vertical", command=tv_unc.yview)
xs2 = tk.Scrollbar(tab_uncert, orient="horizontal", command=tv_unc.xview)
tv_unc.configure(xscrollcommand=xs2.set, yscrollcommand=ys2.set)
tv_unc.pack(fill="both", expand=True, padx=5, pady=(5, 0))
ys2.pack(side="right", fill="y")
xs2.pack(side="bottom", fill="x")

# 상태바
frame_status = ttk.Frame(root, padding=5)
frame_status.pack(fill="x", side="bottom")
ttk.Label(
    frame_status,
    text="🏅 품질관리기술사 김훈희",
    font=("맑은 고딕", 10, "italic"),
    foreground="#555555"
).pack(side="left")

# 초기 모드 상태 반영
on_tol_mode_change()

root.mainloop()