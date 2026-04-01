import os
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import json
import re
import datetime
import html
import importlib.util
import sys
from pathlib import Path

APP_ROOT = Path(__file__).resolve().parent
LOGIC_ROOT = APP_ROOT / "logic"
if str(APP_ROOT) not in sys.path:
    sys.path.insert(0, str(APP_ROOT))


def _ensure_local_logic_package():
    logic_init = LOGIC_ROOT / "__init__.py"
    if not logic_init.exists():
        return

    desired_root = str(LOGIC_ROOT.resolve())
    existing = sys.modules.get("logic")
    existing_paths = [str(Path(path).resolve()) for path in getattr(existing, "__path__", [])] if existing else []
    if desired_root in existing_paths:
        return

    for name, module in list(sys.modules.items()):
        if name != "logic" and not name.startswith("logic."):
            continue
        module_file = str(getattr(module, "__file__", "") or "")
        module_paths = [str(Path(path).resolve()) for path in getattr(module, "__path__", [])]
        if desired_root in module_paths or module_file.startswith(desired_root):
            continue
        sys.modules.pop(name, None)

    spec = importlib.util.spec_from_file_location(
        "logic",
        logic_init,
        submodule_search_locations=[str(LOGIC_ROOT)],
    )
    if spec is None or spec.loader is None:
        return

    module = importlib.util.module_from_spec(spec)
    sys.modules["logic"] = module
    spec.loader.exec_module(module)


_ensure_local_logic_package()

from logic.common import (
    calc_u_distribution,
    load_and_normalize_data,
    normalize_study1_values,
    normalize_study2_df,
    normalize_study4_df,
    normalize_study5_df,
    normalize_study6_df,
    detect_outliers,
    check_duplicate_measurements,
)
from logic.study1_core import run_study1_analysis
from logic.study2_core import run_study2_analysis
from logic.study4_core import run_study4_analysis
from logic.study5_core import run_study5_analysis
from logic.study6_core import run_study6_analysis, calc_fleiss_kappa
from logic.report import (
    create_study1_pdf,
    create_study2_pdf,
    create_study4_pdf,
    create_study5_pdf,
    create_study6_pdf,
)

AI_FEATURE_AVAILABLE = True
AI_FEATURE_IMPORT_ERROR = ""

try:
    from logic.ai_payloads import (
        build_study1_payload as build_ai_study1_payload,
        build_study23_payload as build_ai_study23_payload,
        build_study4_payload as build_ai_study4_payload,
        build_combined_payload as build_ai_combined_payload,
    )
    from logic.ai_explainer import build_ai_explanation
except Exception as exc:
    try:
        _ensure_local_logic_package()
        from logic.ai_payloads import (
            build_study1_payload as build_ai_study1_payload,
            build_study23_payload as build_ai_study23_payload,
            build_study4_payload as build_ai_study4_payload,
            build_combined_payload as build_ai_combined_payload,
        )
        from logic.ai_explainer import build_ai_explanation
    except Exception as inner_exc:
        AI_FEATURE_AVAILABLE = False
        AI_FEATURE_IMPORT_ERROR = f"{exc} | retry: {inner_exc}"

        def build_ai_study1_payload(*args, **kwargs):
            return {}

        def build_ai_study23_payload(*args, **kwargs):
            return {}

        def build_ai_study4_payload(*args, **kwargs):
            return {}

        def build_ai_combined_payload(*args, **kwargs):
            return {}

    def build_ai_explanation(payload, mode="summary", settings=None):
        return {
            "headline": "",
            "summary": [],
            "interpretation": [],
            "actions": [],
            "lecture_points": [],
            "highlights": [],
            "report_comment": "",
            "mode": mode,
            "source": "disabled",
            "error": AI_FEATURE_IMPORT_ERROR,
        }

st.set_page_config(page_title="MSA AI Studio", layout="wide", page_icon="🔬")

# --- 🚫 브라우저 자동번역 방지 (부모 프레임에 translate=no 적용) ---
components.html("""
<script>
(function(){
    var p = window.parent || window;
    var doc = p.document;
    var h = doc.documentElement;
    h.setAttribute('translate', 'no');
    h.classList.add('notranslate');
    h.lang = 'ko';
    if (!doc.querySelector('meta[name="google"]')) {
        var m = doc.createElement('meta');
        m.name = 'google'; m.content = 'notranslate';
        doc.head && doc.head.appendChild(m);
    }
    // MutationObserver: 번역된 Study 용어 즉시 원상복구
    var terms = {
        '연구 가이드':'Study Guide', '학습 가이드':'Study Guide',
        '연구-1':'Study-1', '연구 1':'Study-1',
        '연구-2':'Study-2', '스터디-2':'Study-2',
        '연구-3':'Study-3', '스터디-3':'Study-3',
        '연구-4':'Study-4', '스터디-4':'Study-4',
        '연구-5':'Study-5', '스터디-5':'Study-5',
        '연구-6':'Study-6', '스터디-6':'Study-6',
        '연구-2/3':'Study-2/3', '스터디-2/3':'Study-2/3'
    };
    function fixNode(node) {
        if (node.nodeType === 3) {
            var t = node.textContent;
            for (var k in terms) { t = t.split(k).join(terms[k]); }
            if (t !== node.textContent) node.textContent = t;
        } else if (node.nodeType === 1) {
            node.childNodes.forEach(fixNode);
        }
    }
    var obs = new MutationObserver(function(muts) {
        muts.forEach(function(m) {
            m.addedNodes.forEach(fixNode);
            if (m.type === 'characterData') fixNode(m.target);
        });
    });
    if (doc.body) {
        obs.observe(doc.body, { childList:true, subtree:true, characterData:true });
        fixNode(doc.body);
    }

    // AG Grid 컨텍스트 메뉴 한글화
    var menuMap = {
        'Format': '형식', 'Autosize': '자동 너비',
        'Pin column': '열 고정', 'Pin Column': '열 고정',
        'Hide column': '열 숨기기', 'Hide Column': '열 숨기기',
        'Reset columns': '열 초기화', 'Reset Columns': '열 초기화',
        'Ascending': '오름차순', 'Descending': '내림차순',
        'No sort': '정렬 없음',
    };
    function translateMenuNode(node) {
        if (node.nodeType === 3) {
            var t = node.textContent.trim();
            if (menuMap[t]) node.textContent = menuMap[t];
        } else if (node.nodeType === 1) {
            var cl = node.className || '';
            if (typeof cl === 'string' && (cl.indexOf('ag-menu') !== -1 || cl.indexOf('ag-popup') !== -1 || cl.indexOf('ag-column-drop') !== -1)) {
                node.querySelectorAll('*').forEach(function(el) {
                    el.childNodes.forEach(function(c) { if (c.nodeType === 3) { var t2 = c.textContent.trim(); if (menuMap[t2]) c.textContent = menuMap[t2]; } });
                });
            }
            node.childNodes.forEach(translateMenuNode);
        }
    }
    var menuObs = new MutationObserver(function(muts) {
        muts.forEach(function(m) { m.addedNodes.forEach(translateMenuNode); });
    });
    if (doc.body) {
        menuObs.observe(doc.body, { childList:true, subtree:true });
    }
})();
</script>
""", height=0)

# --- 🎨 글로벌 스타일 ---
st.markdown("""
<style>
.main .block-container { padding-top: 1.5rem; }
.status-ok-box {
    color: #2ecc71; font-weight: bold; font-size: 1.2rem;
    border: 2px solid #2ecc71; padding: 10px; border-radius: 8px;
    background-color: rgba(46,204,113,0.05); text-align: center; margin-bottom: 15px;
}
.status-ng-box {
    color: #e74c3c; font-weight: bold; font-size: 1.2rem;
    border: 2px solid #e74c3c; padding: 10px; border-radius: 8px;
    background-color: rgba(231,76,60,0.05); text-align: center; margin-bottom: 15px;
}
.status-warn-box {
    color: #8a6d1d; font-weight: bold; font-size: 1.2rem;
    border: 2px solid #f0c36d; padding: 10px; border-radius: 8px;
    background-color: rgba(240,195,109,0.15); text-align: center; margin-bottom: 15px;
}
.stNumberInput label { font-weight: bold; font-size: 0.85rem; }
.stTabs [data-baseweb="tab-list"] { gap: 10px; }
.stTabs [data-baseweb="tab"] {
    height: 40px; background-color: #f8f9fa;
    border-radius: 5px; padding: 10px;
}

/* ── 파일 업로더 → 버튼 스타일 통일 ── */
[data-testid="stFileUploaderDropzone"] {
    min-height: 0 !important;
    padding: 0 !important;
    border: 1px solid rgba(49,51,63,0.2) !important;
    border-radius: 8px !important;
    background: transparent !important;
    box-shadow: none !important;
}
/* 드래그앤드롭 안내 텍스트만 숨김 */
[data-testid="stFileUploaderDropzoneInstructions"] {
    display: none !important;
}
[data-testid="stFileUploaderDropzone"] > div,
[data-testid="stFileUploaderDropzone"] > section {
    padding: 0 !important;
    min-height: 0 !important;
    gap: 0 !important;
    border: none !important;
    background: transparent !important;
}
/* Browse 버튼 한글화 및 스타일 */
[data-testid="stFileUploaderDropzone"] button {
    width: 100% !important;
    min-height: 38px !important;
    border-radius: 8px !important;
    font-size: 0 !important;
}
[data-testid="stFileUploaderDropzone"] button::after {
    content: "파일 선택";
    font-size: 14px !important;
}
/* 업로드 완료 파일명 표시 */
[data-testid="stFileUploaderFile"] {
    padding: 2px 4px !important;
    font-size: 0.75rem !important;
}
.study1-panel-title {
    color: #5d86be;
    font-size: 1.55rem;
    font-weight: 700;
    margin-bottom: 0.75rem;
}
.study1-note-list {
    border: 1px solid #cdd6e3;
    border-radius: 8px;
    background: #fbfcfe;
    padding: 0.8rem 0.95rem;
    margin-bottom: 0.8rem;
}
.study1-note-item {
    color: #a36f00;
    font-size: 0.98rem;
    line-height: 1.55;
    margin: 0.12rem 0;
}
.study1-note-bullet {
    color: #b88300;
    font-weight: 700;
    margin-right: 0.4rem;
}
.study1-input-label {
    color: #5d86be;
    font-size: 1.05rem;
    font-weight: 600;
    margin: 0.2rem 0 0.35rem 0;
}
.study23-panel-title {
    color: #496b99;
    font-size: 1.05rem;
    font-weight: 700;
    margin-bottom: 0.45rem;
}
/* ── 툴팁 ? 아이콘 (msa-help-icon 동일 디자인) ── */
.tip-icon {
    display: inline-flex; align-items: center; justify-content: center;
    width: 1.05rem; height: 1.05rem; border-radius: 999px;
    background: #edf3fb; color: #5a7daa;
    border: 1px solid #cad7e8;
    font-size: 0.72rem; font-weight: 700;
    cursor: help; position: relative; margin-left: 8px;
    vertical-align: middle; line-height: 1;
}
.tip-icon::after {
    content: attr(data-tip);
    position: absolute; bottom: 130%; left: 50%;
    transform: translateX(-50%);
    background: rgba(27,38,59,0.96); color: #fff;
    padding: 6px 12px; border-radius: 6px;
    font-size: 12px; white-space: nowrap;
    pointer-events: none; opacity: 0;
    transition: opacity 0.15s;
    z-index: 9999;
}
.tip-icon:hover::after { opacity: 1; }
/* 판정 기준 HTML 테이블 */
.ref-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
.ref-table th {
    background: #f0f2f6; text-align: left;
    padding: 6px 10px; border-bottom: 2px solid #ddd; white-space: nowrap;
}
.ref-table td { padding: 5px 10px; border-bottom: 1px solid #eee; }
.msa-title-row {
    display: flex;
    align-items: center;
    gap: 0.45rem;
    margin-bottom: 0.45rem;
}
.msa-title-row .study23-panel-title,
.msa-title-row .study1-panel-title,
.msa-title-row .study1-input-label {
    margin-bottom: 0;
}
.msa-help-icon {
    position: relative;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 1.05rem;
    height: 1.05rem;
    border-radius: 999px;
    background: #edf3fb;
    color: #5a7daa;
    border: 1px solid #cad7e8;
    font-size: 0.72rem;
    font-weight: 700;
    cursor: help;
    line-height: 1;
}
.msa-help-icon:focus {
    outline: 2px solid rgba(73, 107, 153, 0.35);
    outline-offset: 1px;
}
.msa-help-bubble {
    visibility: hidden;
    opacity: 0;
    pointer-events: none;
    position: absolute;
    top: calc(100% + 8px);
    left: 50%;
    transform: translateX(-50%) translateY(-4px);
    width: min(280px, 42vw);
    background: rgba(27, 38, 59, 0.96);
    color: #f8fbff;
    border-radius: 10px;
    padding: 0.65rem 0.75rem;
    box-shadow: 0 12px 24px rgba(15, 23, 42, 0.18);
    font-size: 0.78rem;
    line-height: 1.45;
    text-align: left;
    white-space: normal;
    z-index: 9999;
    transition: opacity 0.16s ease, transform 0.16s ease;
}
.msa-help-icon:hover .msa-help-bubble,
.msa-help-icon:focus .msa-help-bubble,
.msa-help-icon:focus-within .msa-help-bubble {
    visibility: visible;
    opacity: 1;
    transform: translateX(-50%) translateY(0);
}
.study23-mini-title {
    color: #6a7f9d;
    font-size: 0.82rem;
    font-weight: 700;
    margin-bottom: 0.2rem;
    text-transform: uppercase;
}
.study23-report-note {
    color: #6b7788;
    font-size: 0.86rem;
}
.study23-metric-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 0.65rem;
    margin: 0.25rem 0 0.85rem 0;
}
.study23-metric-card {
    border: 1px solid #d9e1ec;
    border-radius: 12px;
    padding: 0.8rem 0.9rem;
    background: #f8fbff;
}
.study23-metric-card.is-ok {
    border-color: rgba(46, 204, 113, 0.45);
    background: rgba(46, 204, 113, 0.08);
}
.study23-metric-card.is-ng {
    border-color: rgba(231, 76, 60, 0.42);
    background: rgba(231, 76, 60, 0.08);
}
.study23-metric-card.is-warn {
    border-color: rgba(240, 173, 78, 0.45);
    background: rgba(240, 173, 78, 0.10);
}
.study23-metric-head {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 0.5rem;
    margin-bottom: 0.3rem;
}
.study23-metric-name {
    color: #334155;
    font-size: 0.82rem;
    font-weight: 700;
}
.study23-metric-status {
    border-radius: 999px;
    padding: 0.12rem 0.45rem;
    font-size: 0.73rem;
    font-weight: 700;
}
.study23-metric-status.is-ok {
    background: rgba(46, 204, 113, 0.16);
    color: #1f8f52;
}
.study23-metric-status.is-ng {
    background: rgba(231, 76, 60, 0.14);
    color: #c0392b;
}
.study23-metric-status.is-warn {
    background: rgba(240, 173, 78, 0.18);
    color: #8a6d1d;
}
.study23-metric-value {
    color: #18212f;
    font-size: 1.18rem;
    font-weight: 800;
    line-height: 1.1;
}
.study23-metric-criterion {
    color: #64748b;
    font-size: 0.76rem;
    margin-top: 0.25rem;
}
.consultant-badge {
    position: relative;
    width: 100%;
    margin-top: 0.9rem;
    padding: 0.9rem 1.1rem 0.9rem 1.1rem;
    border-radius: 10px;
    border: none;
    border-top: 4px solid #c9a227;
    background: linear-gradient(160deg, #0d1c38, #1a3260);
    box-shadow: 0 12px 32px rgba(0, 0, 0, 0.38);
    box-sizing: border-box;
}
.consultant-badge-role {
    color: #c9a227;
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.06em;
}
.consultant-badge-name {
    color: #ffffff;
    font-size: 1.25rem;
    font-weight: 800;
    line-height: 1.2;
    margin-top: 0.2rem;
}
.consultant-badge-caption {
    color: rgba(201, 162, 39, 0.8);
    font-size: 0.76rem;
    margin-top: 0.38rem;
    padding-top: 0.38rem;
    border-top: 1px solid rgba(201, 162, 39, 0.25);
}
.stTextArea textarea {
    font-family: Consolas, "Courier New", monospace !important;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────
# Excel 보고서 생성 함수
# ─────────────────────────────────────────
def _get_report_metadata(store=None):
    source = st.session_state if store is None else store
    gage_no = str(source.get("msa_gage_no", "")).strip()
    metadata = {}
    if gage_no:
        metadata["gage_no"] = gage_no
    return metadata


def _report_metadata_items(metadata=None):
    metadata = _get_report_metadata() if metadata is None else metadata
    items = []
    gage_no = str(metadata.get("gage_no", "")).strip()
    if gage_no:
        items.append(("Gauge No.", gage_no))
    return items


def _merge_report_metadata(params=None, store=None):
    merged = dict(params or {})
    for label, value in _report_metadata_items(_get_report_metadata(store)):
        merged[label] = value
    return merged


AI_MODE_LABELS = {
    "summary": "AI 핵심 요약",
    "practical": "AI 상세 해설",
    "report_comment": "AI 보고서 코멘트",
}

AI_AUDIENCE_OPTIONS = {
    "초급": "beginner",
    "실무": "practical",
}

AI_LENGTH_OPTIONS = {
    "짧게": "short",
    "보통": "normal",
    "자세히": "long",
}

AI_RESPONSE_SCHEMA_VERSION = "v3"


def _env_flag(name, default="0"):
    raw = str(os.getenv(name, default)).strip().lower()
    return raw not in {"", "0", "false", "no", "off"}


def _resolve_ai_config_path() -> Path:
    # Allow explicit override for portable deployments.
    override = str(os.getenv("MSA_AI_CONFIG_PATH", "")).strip()
    if override:
        return Path(override).expanduser()

    local_appdata = str(os.getenv("LOCALAPPDATA", "")).strip()
    if local_appdata:
        return Path(local_appdata) / "MSA-AI-Studio" / "ai_config.json"

    return Path.home() / ".msa_ai_studio" / "ai_config.json"


_AI_CONFIG_PATH = _resolve_ai_config_path()
_AI_CONFIG_LEGACY_PATH = APP_ROOT / ".ai_config.json"


def _normalize_ai_config(data: dict | None) -> dict:
    cfg = dict(data or {})

    provider_api_keys_raw = cfg.get("provider_api_keys", {})
    provider_api_keys: dict[str, str] = {}
    if isinstance(provider_api_keys_raw, dict):
        for provider_name, key_value in provider_api_keys_raw.items():
            provider_norm = str(provider_name or "").strip().lower()
            key_norm = str(key_value or "").strip()
            if provider_norm and key_norm:
                provider_api_keys[provider_norm] = key_norm

    legacy_provider = str(cfg.get("provider", "")).strip().lower()
    legacy_key = str(cfg.get("api_key", "")).strip()
    if legacy_provider and legacy_key and legacy_provider not in provider_api_keys:
        provider_api_keys[legacy_provider] = legacy_key

    cfg["provider"] = legacy_provider
    cfg["api_key"] = legacy_key
    cfg["provider_api_keys"] = provider_api_keys
    return cfg


def _load_ai_config() -> dict:
    import json

    candidate_paths = []
    if _AI_CONFIG_PATH.exists():
        candidate_paths.append(_AI_CONFIG_PATH)
    if _AI_CONFIG_LEGACY_PATH.exists() and _AI_CONFIG_LEGACY_PATH != _AI_CONFIG_PATH:
        candidate_paths.append(_AI_CONFIG_LEGACY_PATH)

    for path in candidate_paths:
        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return _normalize_ai_config(data)
        except Exception:
            continue

    return _normalize_ai_config({})


def _save_ai_config(config: dict) -> None:
    try:
        import json
        normalized = _normalize_ai_config(config)
        _AI_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with _AI_CONFIG_PATH.open("w", encoding="utf-8") as f:
            json.dump(normalized, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _build_ai_provider_config(
    base_config: dict,
    provider: str,
    api_key: str,
    model: str,
    clear_key: bool = False,
) -> dict:
    cfg = _normalize_ai_config(base_config)
    provider_norm = str(provider or "").strip().lower()
    api_key_norm = str(api_key or "").strip()
    provider_api_keys = dict(cfg.get("provider_api_keys", {}))

    if provider_norm:
        if clear_key:
            provider_api_keys.pop(provider_norm, None)
        elif api_key_norm:
            provider_api_keys[provider_norm] = api_key_norm

    cfg.update(
        {
            "provider": provider_norm,
            "api_key": api_key_norm,
            "model": model,
            "provider_api_keys": provider_api_keys,
        }
    )
    return cfg


def _ensure_ai_state():
    ss = st.session_state
    ss.setdefault(
        "ai_settings",
        {
            "enabled": _env_flag("AI_FEATURE_ENABLED", "0"),
            "provider": str(os.getenv("AI_PROVIDER", "")).strip(),
            "api_key": str(os.getenv("ANTHROPIC_API_KEY", os.getenv("OPENAI_API_KEY", ""))).strip(),
            "model": "",
            "language": "ko",
        },
    )
    ss.setdefault("ai_outputs", {})
    ss.setdefault("ai_report_comments", {})
    ss.setdefault("ai_selected_modes", {})
    ss.setdefault("ai_last_error", {})


def _ai_cache_key(payload, mode, audience, length):
    payload_hash = str(payload.get("payload_hash", "nohash"))
    return f"{AI_RESPONSE_SCHEMA_VERSION}:{payload_hash}:{mode}:{audience}:{length}"


def _build_ai_response(payload, study_key, mode, audience, length):
    _ensure_ai_state()
    settings = dict(st.session_state.get("ai_settings", {}))
    settings.update(
        {
            "audience": audience,
            "length": length,
            "language": settings.get("language", "ko"),
        }
    )

    payload_for_ai = dict(payload)
    payload_for_ai["audience"] = audience
    payload_for_ai["length"] = length
    payload_for_ai["language"] = settings.get("language", "ko")

    cache_key = _ai_cache_key(payload_for_ai, mode, audience, length)
    cached = st.session_state["ai_outputs"].get(study_key, {}).get(cache_key)
    if cached is not None:
        return cached

    response = build_ai_explanation(payload_for_ai, mode=mode, settings=settings)
    response["mode"] = mode
    response["audience"] = audience
    response["length"] = length
    response["cache_key"] = cache_key
    response["payload_hash"] = payload_for_ai.get("payload_hash")

    study_outputs = st.session_state["ai_outputs"].setdefault(study_key, {})
    study_outputs[cache_key] = response
    if response.get("report_comment"):
        st.session_state["ai_report_comments"][study_key] = {
            "content": response.get("report_comment", ""),
            "mode": mode,
            "audience": audience,
            "length": length,
            "payload_hash": payload_for_ai.get("payload_hash"),
        }
    if response.get("error"):
        st.session_state["ai_last_error"][study_key] = response["error"]
    return response


def _render_ai_output_block(response):
    source = str(response.get("source", "fallback")).strip().lower()
    source_label = "외부 AI" if source == "llm" else "기본 해설 엔진"
    st.caption(f"{source_label} 결과입니다. 최종 판단은 원본 결과표 기준입니다.")

    standard = str(response.get("standard", "")).strip()
    if standard:
        st.caption(f"해설 기준 모드: {standard}")

    audience = str(response.get("audience", "")).strip().lower()
    audience_label = {
        "beginner": "초급 입문형 설명",
        "practical": "실무 적용형 설명",
    }.get(audience, "")
    if audience_label:
        st.caption(f"설명 관점: {audience_label}")

    reference_note = str(response.get("reference_note", "")).strip()
    if reference_note:
        st.caption(reference_note)

    headline = str(response.get("headline", "")).strip()
    if headline:
        st.markdown(f"**{headline}**")

    def _ai_highlight_kind(status):
        normalized = str(status or "").strip().lower()
        if any(token in normalized for token in ["ok", "적합", "완료"]):
            return "ok"
        if any(token in normalized for token in ["warn", "주의", "조건부"]):
            return "warn"
        if any(token in normalized for token in ["ng", "부적합", "미실행", "불안정"]):
            return "ng"
        return ""

    def _ai_highlight_card_html(card):
        card_state = f"is-{_ai_highlight_kind(card.get('status'))}" if _ai_highlight_kind(card.get("status")) else ""
        return f"""
        <div class="study23-metric-card {card_state}">
            <div class="study23-metric-head">
                <div class="study23-metric-name">{html.escape(str(card.get('metric_name', '-')))}</div>
                <div class="study23-metric-status {card_state}">{html.escape(str(card.get('status', '-') or '-'))}</div>
            </div>
            <div class="study23-metric-value">{html.escape(str(card.get('metric_value', '-')))}</div>
            <div class="study23-metric-criterion">기준: {html.escape(str(card.get('criterion', '-') or '-'))}</div>
        </div>
        """

    highlights = [item for item in (response.get("highlights") or []) if item]
    if highlights:
        st.markdown("**분석 하이라이트**")
        for idx in range(0, len(highlights), 3):
            group = highlights[idx:idx + 3]
            cols = st.columns(len(group), gap="small")
            for col, card in zip(cols, group):
                with col:
                    st.markdown(_ai_highlight_card_html(card), unsafe_allow_html=True)

    summary_lines = [str(line).strip() for line in response.get("summary", []) if str(line).strip()]
    interpretation_lines = [str(line).strip() for line in response.get("interpretation", []) if str(line).strip()]
    all_key_points = summary_lines + interpretation_lines
    if all_key_points:
        st.markdown("**핵심 키포인트**")
        for line in all_key_points:
            st.markdown(f"- {line}")

    actions = [str(line).strip() for line in response.get("actions", []) if str(line).strip()]
    if actions:
        st.markdown("**권장 액션**")
        for line in actions:
            st.markdown(f"- {line}")

    report_comment = str(response.get("report_comment", "")).strip()
    if report_comment:
        with st.expander("보고서 코멘트", expanded=False):
            st.markdown(report_comment)

    error_text = str(response.get("error", "")).strip()
    if error_text:
        st.caption(f"AI 호출 오류가 있어 기본 해설로 대체되었습니다: {error_text}")


def _render_ai_explainer_card(study_key, payload, title="AI 해설"):
    _ensure_ai_state()
    if not AI_FEATURE_AVAILABLE:
        st.info(f"AI 해설 모듈을 불러오지 못해 기본 앱만 실행합니다: {AI_FEATURE_IMPORT_ERROR}")
        return
    if not payload:
        st.info("분석 실행 후 AI 해설을 사용할 수 있습니다.")
        return

    overall = str(payload.get("judgement_summary", {}).get("overall", "")).strip()
    if overall in {"", "미실행"}:
        st.info("분석 실행 후 AI 해설을 사용할 수 있습니다.")
        return

    selected_modes = st.session_state["ai_selected_modes"]
    default_mode = selected_modes.get(study_key, "summary")
    if default_mode not in AI_MODE_LABELS:
        default_mode = "summary"
        selected_modes[study_key] = default_mode

    with st.container(border=True):
        st.markdown(f"<div class='study23-panel-title'>{title}</div>", unsafe_allow_html=True)
        st.caption("AI 해설은 보조 설명이며 최종 판단은 기존 결과표 기준입니다.")

        option_cols = st.columns([1, 1, 1.15], gap="small")
        audience_label = option_cols[0].selectbox(
            "대상 독자",
            list(AI_AUDIENCE_OPTIONS.keys()),
            index=0,
            key=f"{study_key}_ai_audience",
        )
        length_label = option_cols[1].selectbox(
            "설명 길이",
            list(AI_LENGTH_OPTIONS.keys()),
            index=1,
            key=f"{study_key}_ai_length",
        )
        provider_active = bool(st.session_state["ai_settings"].get("enabled")) and bool(st.session_state["ai_settings"].get("provider"))
        option_cols[2].caption(
            "외부 AI 연결 상태: 사용"
            if provider_active
            else "외부 AI 미연결. 기본 해설 엔진으로 동작합니다."
        )

        triggered = st.button(
            "AI 핵심 요약",
            key=f"{study_key}_summary_btn",
            use_container_width=True,
        )
        if triggered:
            selected_modes[study_key] = "summary"

        audience = AI_AUDIENCE_OPTIONS[audience_label]
        length = AI_LENGTH_OPTIONS[length_label]
        cache_key = _ai_cache_key(payload, "summary", audience, length)
        cached = st.session_state["ai_outputs"].get(study_key, {}).get(cache_key)

        if triggered:
            cached = _build_ai_response(payload, study_key, "summary", audience, length)

        if cached is None:
            st.info("버튼을 눌러 AI 핵심 요약을 생성하세요.")
            return

        _render_ai_output_block(cached)


def create_study1_excel(df_summary, raw_X, params_dict, standard="ISO 22514-7"):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    is_iso_export = "ISO 22514-7" in str(standard)
    standard_label = "ISO 22514-7" if is_iso_export else "AIAG MSA 4th Edition"
    title_text = (
        "MSA Study-1 분석 보고서 (ISO 22514-7)"
        if is_iso_export
        else "MSA Study-1 분석 보고서 (AIAG MSA 4th Edition)"
    )

    export_df = df_summary.copy()
    if not is_iso_export and "항목" in export_df.columns:
        export_df = export_df[
            ~export_df["항목"].astype(str).str.contains(
                r"C_MS|Q_MS|^u_|^U_MS$|^u_MS_REST$|^Abias$",
                regex=True,
                na=False,
            )
        ].reset_index(drop=True)

    if is_iso_export:
        export_params = dict(params_dict)
    else:
        allowed = {"Standard", "RE (분해능)", "RefV (참조값)", "USL (상한)", "LSL (하한)", "TOL"}
        allowed.add("Gauge No.")
        export_params = {k: v for k, v in params_dict.items() if k in allowed}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Study-1 결과"

    # 타이틀
    ws['A1'] = title_text
    ws['A1'].font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f"분석일시: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(name='맑은 고딕', size=10, italic=True)
    ws.merge_cells('A2:D2')

    # 입력 파라미터
    r = 4
    ws.cell(r, 1, '입력 파라미터').font = Font(name='맑은 고딕', bold=True, size=11)
    r += 1
    for k, v in export_params.items():
        ws.cell(r, 1, k).font = Font(name='맑은 고딕')
        ws.cell(r, 2, v).font = Font(name='맑은 고딕')
        r += 1

    r += 1
    ws.cell(r, 1, '분석 결과').font = Font(name='맑은 고딕', bold=True, size=11)
    r += 1
    for col, h in enumerate(['항목', '값', '기준', '판정'], 1):
        c = ws.cell(r, col, h)
        c.font = Font(name='맑은 고딕', bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        c.alignment = Alignment(horizontal='center')
    r += 1

    for _, row in export_df.iterrows():
        ws.cell(r, 1, str(row['항목'])).font = Font(name='맑은 고딕')
        val = row['값']
        vc = ws.cell(r, 2, float(val) if isinstance(val, (int, float, np.integer, np.floating)) else str(val))
        vc.font = Font(name='맑은 고딕')
        if isinstance(val, (int, float, np.integer, np.floating)):
            vc.number_format = '0.000'
        ws.cell(r, 3, str(row.get('기준', ''))).font = Font(name='맑은 고딕')
        판정 = str(row['판정'])
        pc = ws.cell(r, 4, 판정)
        pc.font = Font(name='맑은 고딕', bold=True)
        if 판정 == 'OK':
            pc.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        elif 판정 == 'NG':
            pc.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        r += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10

    # 원시 데이터 시트
    ws2 = wb.create_sheet('측정 원시 데이터')
    ws2['A1'] = f"Standard: {standard_label}"
    ws2['A1'].font = Font(name='맑은 고딕', bold=True)
    ws2.merge_cells('A1:B1')
    ws2['A2'] = '번호'
    ws2['B2'] = '측정값'
    ws2['A2'].font = Font(name='맑은 고딕', bold=True)
    ws2['B2'].font = Font(name='맑은 고딕', bold=True)
    for i, x in enumerate(raw_X, 1):
        ws2.cell(i + 2, 1, i)
        ws2.cell(i + 2, 2, float(x))
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_study6_excel(
    res_df: pd.DataFrame,
    fk: dict,
    raw_df: pd.DataFrame,
    metadata: dict | None = None,
) -> bytes | None:
    """Study-6 Attribute 분석 결과를 openpyxl Excel 보고서로 반환."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    # ── 색상 팔레트 ──
    C_TITLE   = "4472C4"   # 진파랑
    C_SUB     = "5B9BD5"   # 중파랑
    C_HEAD    = "D9E1F2"   # 연파랑
    C_OK      = "E2EFDA"   # 연초록
    C_WARN    = "FFF2CC"   # 연노랑
    C_NG      = "FCE4D6"   # 연빨강
    C_WHITE   = "FFFFFF"
    FONT_NAME = "맑은 고딕"

    def hfill(color): return PatternFill("solid", fgColor=color)
    def hfont(bold=False, color="000000", size=10):
        return Font(name=FONT_NAME, bold=bold, color=color, size=size)
    def halign(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    thin = Side(style="thin", color="BFBFBF")
    def hborder(): return Border(left=thin, right=thin, top=thin, bottom=thin)

    def grade_fill(val: str) -> PatternFill | None:
        if val in ("적합", "우수"):   return hfill(C_OK)
        if val in ("조건부 채택", "보통"): return hfill(C_WARN)
        if val in ("부적합", "미흡"): return hfill(C_NG)
        return None

    def write_header_row(ws, row, headers, col_start=1, bg=C_TITLE, fg="FFFFFF"):
        for c, h in enumerate(headers, col_start):
            cell = ws.cell(row, c, h)
            cell.font      = hfont(bold=True, color=fg)
            cell.fill      = hfill(bg)
            cell.alignment = halign()
            cell.border    = hborder()

    def write_data_cell(ws, row, col, value, bold=False, fill=None, num_fmt=None, align="center"):
        cell = ws.cell(row, col, value)
        cell.font      = hfont(bold=bold)
        cell.alignment = halign(h=align)
        cell.border    = hborder()
        if fill:  cell.fill = fill
        if num_fmt: cell.number_format = num_fmt
        return cell

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════
    # Sheet 1 — 분석 결과 요약
    # ══════════════════════════════════════
    ws = wb.active
    ws.title = "분석 결과"

    # 제목
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value     = "MSA Study-6 : Attribute 계수형 합치도 분석 보고서 (AIAG MSA 4th)"
    title_cell.font      = hfont(bold=True, color=C_WHITE, size=13)
    title_cell.fill      = hfill(C_TITLE)
    title_cell.alignment = halign()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"].value     = f"분석 일시: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font      = hfont(size=9)
    ws["A2"].alignment = halign(h="right")

    # ── 요약 지표 ──
    metadata_items = _report_metadata_items(metadata)

    r = 4
    if metadata_items:
        ws.merge_cells(f"A{r}:J{r}")
        ws.cell(r, 1).value = "[ Report Info ]"
        ws.cell(r, 1).font = hfont(bold=True, size=11, color=C_WHITE)
        ws.cell(r, 1).fill = hfill(C_SUB)
        ws.cell(r, 1).alignment = halign(h="left")
        r += 1
        write_header_row(ws, r, ["Item", "Value"], bg=C_HEAD, fg="000000")
        r += 1
        for label, value in metadata_items:
            write_data_cell(ws, r, 1, label, bold=True, align="left")
            write_data_cell(ws, r, 2, value, align="left")
            r += 1
        r += 1

    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(r, 1).value     = "[ 요약 지표 ]"
    ws.cell(r, 1).font      = hfont(bold=True, size=11, color=C_WHITE)
    ws.cell(r, 1).fill      = hfill(C_SUB)
    ws.cell(r, 1).alignment = halign(h="left")
    r += 1

    summary_items = [
        ("평균 유효성(E)",  round(float(res_df["유효성(E)"].mean()), 4), "≥ 0.90: 적합"),
        ("평균 P(FA)",     round(float(res_df["P(FA)"].mean()), 4),     "< 0.05: 적합"),
        ("평균 P(miss)",   round(float(res_df["P(miss)"].mean()), 4),   "< 0.02: 적합"),
        ("평균 Cohen's κ", round(float(res_df["Cohen's κ"].mean()), 4), "≥ 0.75: 우수 (AIAG)"),
    ]
    write_header_row(ws, r, ["지표", "값", "기준"], bg=C_HEAD, fg="000000")
    r += 1
    for label, val, crit in summary_items:
        write_data_cell(ws, r, 1, label, align="left")
        write_data_cell(ws, r, 2, val, num_fmt="0.0000")
        write_data_cell(ws, r, 3, crit, align="left")
        r += 1

    # ── 평가자별 상세 결과 ──
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(r, 1).value     = "[ 평가자별 분석 결과 ]"
    ws.cell(r, 1).font      = hfont(bold=True, size=11, color=C_WHITE)
    ws.cell(r, 1).fill      = hfill(C_SUB)
    ws.cell(r, 1).alignment = halign(h="left")
    r += 1

    detail_headers = ["평가자", "유효성(E)", "유효성 판정", "P(FA)", "P(FA) 판정",
                       "P(miss)", "P(miss) 판정", "Cohen's κ", "AIAG 판정(κ)", "종합 판정"]
    write_header_row(ws, r, detail_headers)
    r += 1

    for _, row_data in res_df.iterrows():
        verdict_grade_cols = {"유효성 판정", "P(FA) 판정", "P(miss) 판정", "종합 판정"}
        kappa_grade_cols   = {"AIAG 판정(κ)"}
        for ci, col in enumerate(detail_headers, 1):
            val = row_data.get(col, "")
            fill = None
            bold = False
            if col in verdict_grade_cols or col in kappa_grade_cols:
                fill = grade_fill(str(val))
                bold = True
            num_fmt = "0.000" if col in {"유효성(E)", "P(FA)", "P(miss)", "Cohen's κ"} else None
            write_data_cell(ws, r, ci, val, bold=bold, fill=fill, num_fmt=num_fmt)
        r += 1

    # ── 계산식 상세 ──
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(r, 1).value     = "[ 계산식 상세 ]"
    ws.cell(r, 1).font      = hfont(bold=True, size=11, color=C_WHITE)
    ws.cell(r, 1).fill      = hfill(C_SUB)
    ws.cell(r, 1).alignment = halign(h="left")
    r += 1

    calc_headers = ["평가자", "유효성 계산", "P(FA) 계산", "P(miss) 계산", "AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"]
    write_header_row(ws, r, calc_headers, bg=C_HEAD, fg="000000")
    r += 1
    for _, row_data in res_df.iterrows():
        for ci, col in enumerate(calc_headers, 1):
            val = row_data.get(col, "")
            fill = grade_fill(str(val)) if col in {"AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"} else None
            write_data_cell(ws, r, ci, str(val), fill=fill, align="left" if "계산" in str(col) else "center")
        r += 1

    # ── Fleiss' Kappa ──
    r += 1
    ws.merge_cells(f"A{r}:J{r}")
    ws.cell(r, 1).value     = "[ Fleiss' Kappa — 평가자 간 상호 일치도 ]"
    ws.cell(r, 1).font      = hfont(bold=True, size=11, color=C_WHITE)
    ws.cell(r, 1).fill      = hfill(C_SUB)
    ws.cell(r, 1).alignment = halign(h="left")
    r += 1

    fk_val = fk.get("fleiss_kappa", float("nan"))
    if not (fk_val != fk_val):  # not NaN
        fk_headers = ["Fleiss' κ", "AIAG 판정", "미니탭 판정", "평가자 수", "부품 수"]
        fk_values  = [round(fk_val, 4), fk["aiag_grade"], fk["minitab_grade"],
                      fk["n_raters"], fk["n_subjects"]]
        write_header_row(ws, r, fk_headers, bg=C_HEAD, fg="000000")
        r += 1
        for ci, (h, v) in enumerate(zip(fk_headers, fk_values), 1):
            fill = grade_fill(str(v)) if h in {"AIAG 판정", "미니탭 판정"} else None
            write_data_cell(ws, r, ci, v, fill=fill,
                            num_fmt="0.0000" if h == "Fleiss' κ" else None)
        r += 1
        if "p_bar" in fk:
            r += 1
            ws.cell(r, 1).value = (f"P̄(관측 일치) = {fk['p_bar']:.4f}  │  "
                                   f"P̄ₑ(기대 일치) = {fk['p_e_bar']:.4f}  │  "
                                   f"p₀ = {fk['p0']:.4f}  │  p₁ = {fk['p1']:.4f}")
            ws.cell(r, 1).font = hfont(size=9)
            ws.merge_cells(f"A{r}:J{r}")

    # 열 너비 조정
    col_widths = [14, 10, 12, 10, 12, 10, 12, 12, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════
    # Sheet 2 — 원시 데이터
    # ══════════════════════════════════════
    ws2 = wb.create_sheet("원시 데이터")
    ws2.merge_cells(f"A1:{get_column_letter(len(raw_df.columns))}1")
    ws2["A1"].value     = "입력 원시 데이터"
    ws2["A1"].font      = hfont(bold=True, color=C_WHITE)
    ws2["A1"].fill      = hfill(C_TITLE)
    ws2["A1"].alignment = halign()

    write_header_row(ws2, 2, list(raw_df.columns), bg=C_HEAD, fg="000000")
    for ri, row_data in raw_df.iterrows():
        for ci, val in enumerate(row_data, 1):
            cell = ws2.cell(ri + 3, ci, val)
            cell.font      = hfont()
            cell.alignment = halign()
            cell.border    = hborder()
    for ci, col in enumerate(raw_df.columns, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = max(12, len(str(col)) + 4)

    # ══════════════════════════════════════
    # Sheet 3 — 판정 기준표
    # ══════════════════════════════════════
    ws3 = wb.create_sheet("판정 기준")
    ws3.merge_cells("A1:D1")
    ws3["A1"].value     = "Study-6 판정 기준표 (AIAG MSA 4th Edition)"
    ws3["A1"].font      = hfont(bold=True, color=C_WHITE, size=12)
    ws3["A1"].fill      = hfill(C_TITLE)
    ws3["A1"].alignment = halign()
    ws3.row_dimensions[1].height = 24

    r3 = 3
    ws3.cell(r3, 1).value = "[ 유효성 / P(FA) / P(miss) 기준 ]"
    ws3.cell(r3, 1).font  = hfont(bold=True, size=10, color=C_WHITE)
    ws3.cell(r3, 1).fill  = hfill(C_SUB)
    ws3.merge_cells(f"A{r3}:D{r3}")
    r3 += 1
    acc_headers = ["판정", "유효성 E", "허위 경보율 P(FA)", "누락률 P(miss)"]
    write_header_row(ws3, r3, acc_headers, bg=C_HEAD, fg="000000")
    r3 += 1
    acc_rows = [
        ("적합",       "0.90 ~ 1.00", "0 ~ 0.05",    "0 ~ 0.02"),
        ("조건부 채택", "0.80 ~ 0.90", "0.05 ~ 0.10", "0.02 ~ 0.05"),
        ("부적합",     "0.80 이하",   "0.10 이상",   "0.05 이상"),
    ]
    for label, e, fa, miss in acc_rows:
        fill = grade_fill(label)
        for ci, v in enumerate([label, e, fa, miss], 1):
            write_data_cell(ws3, r3, ci, v, bold=(ci == 1), fill=fill if ci == 1 else None)
        r3 += 1

    r3 += 1
    ws3.cell(r3, 1).value = "[ Kappa 해석 기준 ]"
    ws3.cell(r3, 1).font  = hfont(bold=True, size=10, color=C_WHITE)
    ws3.cell(r3, 1).fill  = hfill(C_SUB)
    ws3.merge_cells(f"A{r3}:D{r3}")
    r3 += 1
    write_header_row(ws3, r3, ["평가지침", "Cohen's κ (AIAG)", "Cohen's κ (미니탭)", "Fleiss' κ"], bg=C_HEAD, fg="000000")
    r3 += 1
    kappa_rows = [
        ("우수", "≥ 0.75", "≥ 0.90", "≥ 0.75"),
        ("보통", "0.40 ~ 0.75", "0.70 ~ 0.90", "0.40 ~ 0.75"),
        ("미흡", "< 0.40",  "< 0.70",  "< 0.40"),
    ]
    for label, aiag, minitab, fleiss in kappa_rows:
        fill = grade_fill(label)
        for ci, v in enumerate([label, aiag, minitab, fleiss], 1):
            write_data_cell(ws3, r3, ci, v, bold=(ci == 1), fill=fill if ci == 1 else None)
        r3 += 1

    for ci, w in enumerate([14, 18, 20, 18], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_study23_excel(
    params: dict,
    type1_result: dict | None,
    type2_result: dict | None,
    report_text: str = "",
    raw_type1_values=None,
    raw_type2_df: pd.DataFrame | None = None,
    is_expert: bool = False,
) -> bytes | None:
    """Build a Study-2/3 Excel report from the current analysis results."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    C = {
        "title": "4472C4",
        "sub": "5B9BD5",
        "head": "D9E1F2",
        "ok": "E2EFDA",
        "warn": "FFF2CC",
        "ng": "FCE4D6",
        "white": "FFFFFF",
    }
    FN = "Malgun Gothic"
    thin = Side(style="thin", color="BFBFBF")

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def font(bold=False, color="000000", size=10):
        return Font(name=FN, bold=bold, color=color, size=size)

    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def border():
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def grade_fill(value):
        kind = _study23_status_kind(value)
        if kind == "ok":
            return fill(C["ok"])
        if kind == "warn":
            return fill(C["warn"])
        if kind == "ng":
            return fill(C["ng"])
        return None

    def sanitize_excel_text(value):
        text = str(value)
        return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)

    def excel_value(value):
        if isinstance(value, (np.floating, float)):
            return None if pd.isna(value) else float(value)
        if isinstance(value, (np.bool_, bool)):
            return "TRUE" if bool(value) else "FALSE"
        if isinstance(value, (np.integer, int)):
            return int(value)
        if pd.isna(value):
            return None
        return sanitize_excel_text(value)

    def title_row(ws, row, text, span, color=C["title"], size=12):
        cell = ws.cell(row, 1, text)
        cell.font = font(bold=True, color=C["white"], size=size)
        cell.fill = fill(color)
        cell.alignment = align(h="left")
        ws.row_dimensions[row].height = 22

    def subhead(ws, row, text, span, color=C["sub"]):
        cell = ws.cell(row, 1, text)
        cell.font = font(bold=True, color=C["white"])
        cell.fill = fill(color)
        cell.alignment = align(h="left")

    def head_row(ws, row, headers, col_start=1):
        for ci, header in enumerate(headers, col_start):
            cell = ws.cell(row, ci, header)
            cell.font = font(bold=True)
            cell.fill = fill(C["head"])
            cell.alignment = align()
            cell.border = border()

    def data_cell(ws, row, col, value, bold=False, gfill=None, num_fmt=None, halign="center"):
        excelized = excel_value(value)
        cell = ws.cell(row, col)
        cell.value = excelized
        if isinstance(excelized, str):
            cell.data_type = "s"
        cell.font = font(bold=bold)
        cell.alignment = align(h=halign)
        cell.border = border()
        if gfill:
            cell.fill = gfill
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    def df_to_sheet(ws, df, start_row, grade_cols=None, num_cols=None, align_left_cols=None, cell_fill_fn=None):
        if df is None or df.empty:
            return start_row
        grade_cols = grade_cols or set()
        num_cols = num_cols or set()
        align_left_cols = align_left_cols or set()
        head_row(ws, start_row, list(df.columns))
        row = start_row + 1
        for _, row_data in df.iterrows():
            for ci, col in enumerate(df.columns, 1):
                value = row_data[col]
                fill_color = cell_fill_fn(col, value, row_data) if callable(cell_fill_fn) else None
                data_cell(
                    ws,
                    row,
                    ci,
                    value,
                    bold=((col in grade_cols and grade_fill(str(value)) is not None) or fill_color is not None),
                    gfill=fill_color if fill_color is not None else (grade_fill(str(value)) if col in grade_cols else None),
                    num_fmt="0.0000" if col in num_cols else None,
                    halign="left" if col in align_left_cols else "center",
                )
            row += 1
        return row

    def auto_width(ws, min_width=12, max_width=36):
        for idx, col_cells in enumerate(ws.iter_cols(1, ws.max_column), 1):
            width = min_width
            for cell in col_cells:
                if cell.value is None:
                    continue
                width = max(width, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(idx)].width = min(width, max_width)

    standard = str(params.get("standard", "ISO 22514-7" if is_expert else "AIAG")).strip()
    is_iso_export = "ISO 22514-7" in standard
    standard_label = "ISO 22514-7" if is_iso_export else "AIAG MSA 4th Edition"
    sheet_prefix = "ISO22514-7" if is_iso_export else "AIAG"
    workbook_title = (
        "MSA Study-2/3 Report (ISO 22514-7)"
        if is_iso_export
        else "MSA Study-2/3 Report (AIAG MSA 4th Edition)"
    )

    def compact_text(value):
        return sanitize_excel_text(value).strip()

    def compact_df(df, rename_map=None, keep_cols=None):
        if df is None or df.empty:
            return pd.DataFrame()
        out = df.copy()
        if keep_cols is not None:
            cols = [col for col in keep_cols if col in out.columns]
            out = out[cols]
        if rename_map:
            out = out.rename(columns={col: rename_map.get(col, col) for col in out.columns})
        return out

    def compact_keywords(lines):
        keywords = []
        for raw_line in lines:
            line = compact_text(raw_line)
            if not line:
                continue
            if line.startswith("==="):
                keywords.append(line.strip("= ").strip())
            elif len(keywords) < 3:
                keywords.append(line)
            if len(keywords) >= 4:
                break
        return keywords

    def study23_variation_fill(column, value, _row_data=None):
        if column not in {"Tol%", "SV%"}:
            return None
        if not isinstance(value, (int, float, np.integer, np.floating)) or pd.isna(value):
            return None
        numeric = float(value)
        if numeric <= 10:
            return fill(C["ok"])
        if numeric <= 30:
            return fill(C["warn"])
        return fill(C["ng"])

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = f"{sheet_prefix} Summary"

    title_row(ws_summary, 1, workbook_title, 4, size=13)
    ws_summary["A2"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A2"].font = font(size=9)
    ws_summary["A2"].alignment = align(h="left")

    row = 3
    summary_rows = [
        ("Std", standard_label),
        (
            "Run",
            "Type I + Type II" if (type1_result is not None and type2_result is not None)
            else ("Type I" if type1_result is not None else ("Type II" if type2_result is not None else "-")),
        ),
        ("Status", type2_result.get("overall_status") if type2_result is not None else "-"),
        (
            "Inter",
            "-" if type2_result is None else ("Y" if bool(type2_result.get("interaction_significant")) else "N"),
        ),
        ("Source", params.get("type2_source", "-")),
    ]
    head_row(ws_summary, row, ["Key", "Value"])
    row += 1
    for item, value in summary_rows:
        data_cell(ws_summary, row, 1, item, bold=True, halign="left")
        fill_color = grade_fill(str(value)) if item == "Status" else None
        data_cell(ws_summary, row, 2, value, bold=fill_color is not None, gfill=fill_color, halign="left")
        row += 1

    if is_iso_export:
        param_rows = [
            ("Gauge No.", params.get("Gauge No.")),
            ("RE", params.get("re")),
            ("RefV", params.get("refv")),
            ("USL", params.get("usl")),
            ("LSL", params.get("lsl")),
            ("TOL", params.get("tol")),
            ("u_CAL", params.get("u_cal")),
            ("Bias Dist.", params.get("bias_dist")),
            ("u_RE Dist.", params.get("re_dist")),
            ("u_LIN", params.get("u_lin")),
            ("u_MSREST", params.get("u_msrest")),
            ("u_T", params.get("u_t")),
            ("u_STAB", params.get("u_stab")),
            ("u_OBJ", params.get("u_obj")),
            ("u_GV", params.get("u_gv")),
            ("u_REST", params.get("u_rest")),
        ]
    else:
        param_rows = [
            ("Gauge No.", params.get("Gauge No.")),
            ("RE", params.get("re")),
            ("RefV", params.get("refv")),
            ("USL", params.get("usl")),
            ("LSL", params.get("lsl")),
            ("TOL", params.get("tol")),
        ]
    param_rows = [(k, v) for k, v in param_rows if v not in (None, "")]
    if param_rows:
        row += 1
        head_row(ws_summary, row, ["Param", "Value"])
        row += 1
        for key, value in param_rows:
            num_fmt = "0.0000" if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool) else None
            data_cell(ws_summary, row, 1, key, bold=True, halign="left")
            data_cell(ws_summary, row, 2, value, halign="left", num_fmt=num_fmt)
            row += 1

    visible_judgement = pd.DataFrame()
    if type2_result is not None:
        visible_judgement = _study23_visible_judgement(type2_result, is_iso_export).copy()
        if is_iso_export and not pd.isna(type2_result.get("c_mp", np.nan)):
            c_mp_status = "적합" if float(type2_result["c_mp"]) >= 1.33 else "부적합"
            visible_judgement = pd.concat(
                [
                    visible_judgement,
                    pd.DataFrame(
                        [{"항목": "C_MP", "값": type2_result["c_mp"], "기준": ">= 1.33", "판정": c_mp_status}]
                    ),
                ],
                ignore_index=True,
            )
        if not visible_judgement.empty:
            row += 1
            metric_df = compact_df(
                visible_judgement,
                rename_map={"항목": "Metric", "값": "Value", "기준": "Crit", "판정": "Judge"},
                keep_cols=["항목", "값", "기준", "판정"],
            )
            row = df_to_sheet(
                ws_summary,
                metric_df,
                row,
                grade_cols={"Judge"},
                num_cols={c for c in metric_df.columns if pd.api.types.is_numeric_dtype(metric_df[c])},
                align_left_cols={"Metric", "Crit"},
            )

        warnings = type2_result.get("warnings", [])
        if warnings:
            row += 1
            head_row(ws_summary, row, ["Warn"])
            row += 1
            for warning in warnings:
                data_cell(ws_summary, row, 1, warning, halign="left")
                row += 1
    keywords = compact_keywords(report_text.splitlines()) if report_text else []
    if keywords:
        row += 1
        head_row(ws_summary, row, ["Keyword"])
        row += 1
        for keyword in keywords:
            data_cell(ws_summary, row, 1, keyword, halign="left")
            row += 1

    if type1_result is not None:
        summary_df = type1_result.get("summary")
    else:
        summary_df = None

    ws_core = wb.create_sheet(f"{sheet_prefix} Core")
    title_row(ws_core, 1, f"{sheet_prefix} Core", 5)
    row = 2

    if summary_df is not None:
        summary_df = type1_result.get("summary")
        if summary_df is not None and not summary_df.empty and not is_iso_export and "항목" in summary_df.columns:
            summary_df = summary_df[~summary_df["항목"].astype(str).str.contains(r"u_|Abias|불확도|정의유래", regex=True, na=False)]
        if summary_df is not None and not summary_df.empty:
            summary_df = compact_df(
                summary_df,
                rename_map={"항목": "Metric", "값": "Value", "기준": "Crit", "판정": "Judge"},
                keep_cols=["항목", "값", "기준", "판정"],
            )
            row = df_to_sheet(
                ws_core,
                summary_df,
                row,
                grade_cols={"Judge"},
                num_cols={col for col in summary_df.columns if pd.api.types.is_numeric_dtype(summary_df[col])},
                align_left_cols={"Metric", "Crit"},
            )
            row += 1

    if type2_result is not None:
        gage_df = type2_result.get("gage")
        if gage_df is not None and not gage_df.empty:
            gage_df = compact_df(
                gage_df,
                rename_map={"출처": "Source", "6*SD": "6SD", "%SV": "SV%", "%공차": "Tol%"},
                keep_cols=["출처", "SD", "6*SD", "%SV", "%공차"],
            )
            row = df_to_sheet(
                ws_core,
                gage_df,
                row,
                num_cols={col for col in gage_df.columns if pd.api.types.is_numeric_dtype(gage_df[col])},
                align_left_cols={"Source"},
                cell_fill_fn=study23_variation_fill,
            )
            row += 1

        anova_df = type2_result.get("anova")
        if anova_df is not None and not anova_df.empty:
            anova_export = anova_df.reset_index()
            if "index" in anova_export.columns:
                anova_export = anova_export.rename(columns={"index": "항목"})
            anova_export = compact_df(
                anova_export,
                rename_map={
                    "항목": "Term",
                    "자유도(DF)": "DF",
                    "제곱합(SS)": "SS",
                    "평균제곱(MS)": "MS",
                    "F값": "F",
                    "P값": "P",
                },
            )
            row = df_to_sheet(
                ws_core,
                anova_export,
                row,
                num_cols={col for col in anova_export.columns if pd.api.types.is_numeric_dtype(anova_export[col])},
                align_left_cols={"Term"},
            )
            row += 1

        if is_iso_export:
            u_elements = type2_result.get("u_elements") or {}
            non_zero_u = [
                {"U": key, "Value": value}
                for key, value in u_elements.items()
                if abs(value) > 1e-12
            ]
            if non_zero_u:
                u_df = pd.DataFrame(non_zero_u).sort_values("Value", ascending=False).reset_index(drop=True)
                row = df_to_sheet(
                    ws_core,
                    u_df,
                    row,
                    num_cols={"Value"},
                    align_left_cols={"U"},
                )
                row += 1
                head_row(ws_core, row, ["U Key", "Value"])
                row += 1
                data_cell(ws_core, row, 1, "u_MP", bold=True, halign="left")
                data_cell(ws_core, row, 2, type2_result.get("u_mp"), num_fmt="0.0000")
                row += 1
                data_cell(ws_core, row, 1, "U_MP", bold=True, halign="left")
                data_cell(ws_core, row, 2, type2_result.get("u_mp_expanded"), num_fmt="0.0000")
    auto_width(ws_core, min_width=9, max_width=20)

    has_raw = (raw_type1_values is not None and len(raw_type1_values) > 0) or (raw_type2_df is not None and not raw_type2_df.empty)
    if has_raw:
        ws_raw = wb.create_sheet(f"{sheet_prefix} Raw")
        title_row(ws_raw, 1, f"{sheet_prefix} Raw", 3)
        raw_row = 2
        if raw_type1_values is not None and len(raw_type1_values) > 0:
            head_row(ws_raw, raw_row, ["T1 No", "T1 Val"])
            raw_row += 1
            for idx, value in enumerate(raw_type1_values, 1):
                data_cell(ws_raw, raw_row, 1, idx)
                data_cell(ws_raw, raw_row, 2, value, num_fmt="0.0000")
                raw_row += 1
            raw_row += 1

        if raw_type2_df is not None and not raw_type2_df.empty:
            raw_export = raw_type2_df.copy()
            raw_export = compact_df(
                raw_export,
                rename_map={"부품": "Part", "측정자": "Op", "측정값": "Value", "Operator": "Op"},
            )
            raw_row = df_to_sheet(
                ws_raw,
                raw_export,
                raw_row,
                num_cols={col for col in raw_export.columns if pd.api.types.is_numeric_dtype(raw_export[col])},
                align_left_cols={col for col in raw_export.columns if not pd.api.types.is_numeric_dtype(raw_export[col])},
            )
        auto_width(ws_raw, min_width=9, max_width=18)

    auto_width(ws_summary, min_width=9, max_width=22)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_study4_excel(
    params: dict,
    result: dict | None,
    raw_linearity_df: pd.DataFrame | None = None,
    raw_grr_df: pd.DataFrame | None = None,
    standard: str = "ISO 22514-7",
) -> bytes | None:
    """Build a Study-4 Excel report from the current analysis results."""
    if result is None:
        return None

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    C = {
        "title": "4472C4",
        "sub": "5B9BD5",
        "head": "D9E1F2",
        "ok": "E2EFDA",
        "warn": "FFF2CC",
        "ng": "FCE4D6",
        "white": "FFFFFF",
    }
    FN = "Malgun Gothic"
    thin = Side(style="thin", color="BFBFBF")

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def font(bold=False, color="000000", size=10):
        return Font(name=FN, bold=bold, color=color, size=size)

    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def border():
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def sanitize_excel_text(value):
        text = str(value)
        return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)

    def excel_value(value):
        if isinstance(value, (np.floating, float)):
            return None if pd.isna(value) else float(value)
        if isinstance(value, (np.integer, int)):
            return int(value)
        if isinstance(value, (np.bool_, bool)):
            return "TRUE" if bool(value) else "FALSE"
        if pd.isna(value):
            return None
        return sanitize_excel_text(value)

    def status_fill(value):
        if isinstance(value, (np.bool_, bool)):
            return fill(C["ok"] if bool(value) else C["ng"])
        raw = str(value).strip()
        upper = raw.upper()
        if raw in {"합격", "양호", "적합", "Stable"} or upper in {"OK", "PASS", "Y", "TRUE"}:
            return fill(C["ok"])
        if raw in {"조건부", "보통"} or upper in {"WARN", "WARNING"}:
            return fill(C["warn"])
        if raw in {"부적합", "불합격", "Unstable"} or upper in {"NG", "FAIL", "N", "FALSE"}:
            return fill(C["ng"])
        return None

    def title_row(ws, row, text, color=C["title"], size=12):
        cell = ws.cell(row, 1, text)
        cell.font = font(bold=True, color=C["white"], size=size)
        cell.fill = fill(color)
        cell.alignment = align(h="left")
        ws.row_dimensions[row].height = 22

    def section_row(ws, row, text, color=C["sub"]):
        cell = ws.cell(row, 1, text)
        cell.font = font(bold=True, color=C["white"])
        cell.fill = fill(color)
        cell.alignment = align(h="left")

    def head_row(ws, row, headers):
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(row, ci, header)
            cell.font = font(bold=True)
            cell.fill = fill(C["head"])
            cell.alignment = align()
            cell.border = border()

    def data_cell(ws, row, col, value, bold=False, gfill=None, num_fmt=None, halign="center"):
        excelized = excel_value(value)
        cell = ws.cell(row, col)
        cell.value = excelized
        if isinstance(excelized, str):
            cell.data_type = "s"
        cell.font = font(bold=bold)
        cell.alignment = align(h=halign)
        cell.border = border()
        if gfill:
            cell.fill = gfill
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    def compact_df(df, rename_map=None, keep_cols=None):
        if df is None or df.empty:
            return pd.DataFrame()
        out = df.copy()
        if keep_cols is not None:
            cols = [col for col in keep_cols if col in out.columns]
            out = out[cols]
        if rename_map:
            out = out.rename(columns={col: rename_map.get(col, col) for col in out.columns})
        return out

    def df_to_sheet(ws, df, start_row, grade_cols=None, num_cols=None, align_left_cols=None):
        if df is None or df.empty:
            return start_row
        grade_cols = grade_cols or set()
        num_cols = num_cols or set()
        align_left_cols = align_left_cols or set()
        head_row(ws, start_row, list(df.columns))
        row = start_row + 1
        for _, row_data in df.iterrows():
            for ci, col in enumerate(df.columns, 1):
                value = row_data[col]
                gfill = status_fill(value) if col in grade_cols else None
                data_cell(
                    ws,
                    row,
                    ci,
                    value,
                    bold=(gfill is not None),
                    gfill=gfill,
                    num_fmt="0.0000" if col in num_cols else None,
                    halign="left" if col in align_left_cols else "center",
                )
            row += 1
        return row

    def write_section(ws, row, label, df, grade_cols=None, num_cols=None, align_left_cols=None):
        if df is None or df.empty:
            return row
        section_row(ws, row, label)
        row += 1
        row = df_to_sheet(
            ws,
            df,
            row,
            grade_cols=grade_cols,
            num_cols=num_cols,
            align_left_cols=align_left_cols,
        )
        return row + 1

    def auto_width(ws, min_width=10, max_width=24):
        for idx, col_cells in enumerate(ws.iter_cols(1, ws.max_column), 1):
            width = min_width
            for cell in col_cells:
                if cell.value is None:
                    continue
                width = max(width, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(idx)].width = min(width, max_width)

    standard_text = str(standard or params.get("standard") or "ISO 22514-7").strip()
    is_iso_export = "ISO 22514-7" in standard_text
    standard_label = "ISO 22514-7" if is_iso_export else "AIAG MSA 4th Edition"
    sheet_prefix = "ISO22514-7" if is_iso_export else "AIAG"
    workbook_title = (
        "MSA Study-4 Report (ISO 22514-7)"
        if is_iso_export
        else "MSA Study-4 Report (AIAG MSA 4th Edition)"
    )

    overall_pass = bool(result.get("linearity_pass", False))
    if is_iso_export:
        overall_pass = overall_pass and bool(result.get("process_pass", True))

    if raw_linearity_df is None or raw_linearity_df.empty:
        raw_linearity_df = result.get("raw_with_bias")
    if raw_linearity_df is not None and not raw_linearity_df.empty:
        raw_linearity_df = raw_linearity_df.copy()
        if "Bias" not in raw_linearity_df.columns and {"Reference", "Value"}.issubset(raw_linearity_df.columns):
            raw_linearity_df["Bias"] = raw_linearity_df["Value"] - raw_linearity_df["Reference"]

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = f"{sheet_prefix} Summary"

    title_row(ws_summary, 1, workbook_title, size=13)
    ws_summary["A2"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary["A2"].font = font(size=9)
    ws_summary["A2"].alignment = align(h="left")

    row = 3
    summary_rows = [
        ("Std", standard_label),
        ("Status", "OK" if overall_pass else "NG"),
        ("Linear", "OK" if bool(result.get("linearity_pass", False)) else "NG"),
        ("Process", "-" if not is_iso_export or "gage_result" not in result else ("OK" if bool(result.get("process_pass", False)) else "NG")),
        ("GRR", "Y" if "gage_result" in result else "N"),
    ]
    head_row(ws_summary, row, ["Key", "Value"])
    row += 1
    for key, value in summary_rows:
        value_fill = status_fill(value)
        data_cell(ws_summary, row, 1, key, bold=True, halign="left")
        data_cell(ws_summary, row, 2, value, bold=value_fill is not None, gfill=value_fill, halign="left")
        row += 1

    if is_iso_export:
        param_rows = [
            ("Gauge No.", params.get("Gauge No.")),
            ("RE", params.get("re")),
            ("u_CAL", params.get("u_cal")),
            ("USL", params.get("usl")),
            ("LSL", params.get("lsl")),
            ("TOL", params.get("tol")),
        ]
    else:
        param_rows = [
            ("Gauge No.", params.get("Gauge No.")),
            ("RE", params.get("re")),
            ("USL", params.get("usl")),
            ("LSL", params.get("lsl")),
            ("TOL", params.get("tol")),
        ]
    param_rows = [(key, value) for key, value in param_rows if value not in (None, "")]
    if param_rows:
        row += 1
        head_row(ws_summary, row, ["Param", "Value"])
        row += 1
        for key, value in param_rows:
            data_cell(ws_summary, row, 1, key, bold=True, halign="left")
            data_cell(
                ws_summary,
                row,
                2,
                value,
                num_fmt="0.0000" if isinstance(value, (int, float, np.integer, np.floating)) else None,
                halign="left",
            )
            row += 1

    if is_iso_export:
        metric_frames = [
            compact_df(
                result.get("linearity_capability"),
                rename_map={"항목": "Metric", "값": "Value", "기준": "Crit", "판정": "Judge"},
                keep_cols=["항목", "값", "기준", "판정"],
            )
        ]
        if "capability_summary" in result:
            metric_frames.append(
                compact_df(
                    result.get("capability_summary"),
                    rename_map={"항목": "Metric", "값": "Value", "기준": "Crit", "판정": "Judge"},
                    keep_cols=["항목", "값", "기준", "판정"],
                )
            )
        metric_frames = [df for df in metric_frames if df is not None and not df.empty]
        summary_metrics = pd.concat(metric_frames, ignore_index=True) if metric_frames else pd.DataFrame()
    else:
        aiag_summary = result.get("summary")
        if aiag_summary is not None and not aiag_summary.empty:
            aiag_summary = aiag_summary[
                aiag_summary["항목"].isin(["Intercept", "Slope", "S", "R-squared"])
            ].reset_index(drop=True)
        summary_metrics = compact_df(
            aiag_summary,
            rename_map={"항목": "Metric", "값": "Value", "기준": "Crit", "판정": "Judge"},
            keep_cols=["항목", "값", "기준", "판정"],
        )

    if not summary_metrics.empty:
        row += 1
        row = df_to_sheet(
            ws_summary,
            summary_metrics,
            row,
            grade_cols={"Judge"},
            num_cols={col for col in summary_metrics.columns if pd.api.types.is_numeric_dtype(summary_metrics[col])},
            align_left_cols={"Metric", "Crit"},
        )

    ws_core = wb.create_sheet(f"{sheet_prefix} Core")
    title_row(ws_core, 1, f"{sheet_prefix} Core")
    row = 2

    if is_iso_export:
        sections = [
            (
                "Linearity Capability",
                compact_df(
                    result.get("linearity_capability"),
                    rename_map={"항목": "Metric", "값": "Value", "기준": "Crit", "판정": "Judge"},
                    keep_cols=["항목", "값", "기준", "판정"],
                ),
                {"Judge"},
                {"Metric", "Crit"},
            ),
            (
                "Process Capability",
                compact_df(
                    result.get("capability_summary"),
                    rename_map={"항목": "Metric", "값": "Value", "기준": "Crit", "판정": "Judge"},
                    keep_cols=["항목", "값", "기준", "판정"],
                ),
                {"Judge"},
                {"Metric", "Crit"},
            ),
            (
                "Uncertainty",
                compact_df(
                    result.get("uncertainty_components"),
                    rename_map={"구분": "Group", "코드": "Code", "값": "Value"},
                    keep_cols=["구분", "코드", "값"],
                ),
                set(),
                {"Group", "Code"},
            ),
            (
                "Resolution",
                compact_df(
                    result.get("resolution_summary"),
                    rename_map={"항목": "Metric", "값": "Value"},
                    keep_cols=["항목", "값"],
                ),
                set(),
                {"Metric"},
            ),
            (
                "Gage",
                compact_df(
                    result.get("gage_result", {}).get("gage") if isinstance(result.get("gage_result"), dict) else None,
                    rename_map={"출처": "Source", "6*SD": "6SD", "%SV": "SV%", "%공차": "Tol%"},
                    keep_cols=["출처", "SD", "6*SD", "%SV", "%공차"],
                ),
                set(),
                {"Source"},
            ),
            (
                "Process U",
                compact_df(
                    result.get("process_components"),
                    rename_map={"항목": "U", "값": "Value"},
                    keep_cols=["항목", "값"],
                ),
                set(),
                {"U"},
            ),
            (
                "Regression",
                compact_df(
                    result.get("regression_table"),
                    rename_map={"항목": "Term"},
                ),
                set(),
                {"Term"},
            ),
            (
                "ANOVA",
                compact_df(result.get("anova_table")),
                set(),
                {"Source"},
            ),
        ]
    else:
        sections = [
            (
                "Coefficients",
                compact_df(
                    result.get("aiag_linearity", {}).get("coefficients") if isinstance(result.get("aiag_linearity"), dict) else None,
                    rename_map={"예측 변수": "Term", "계수": "Coef", "SE 계수": "SE", "P": "P"},
                    keep_cols=["예측 변수", "계수", "SE 계수", "P"],
                ),
                set(),
                {"Term"},
            ),
            (
                "Bias",
                compact_df(
                    result.get("aiag_linearity", {}).get("bias_summary") if isinstance(result.get("aiag_linearity"), dict) else None,
                    rename_map={"기준": "Ref", "치우침": "Bias", "P": "P"},
                    keep_cols=["기준", "치우침", "P"],
                ),
                set(),
                {"Ref"},
            ),
            (
                "Regression",
                compact_df(
                    result.get("regression_table"),
                    rename_map={"항목": "Term"},
                ),
                set(),
                {"Term"},
            ),
            (
                "ANOVA",
                compact_df(result.get("anova_table")),
                set(),
                {"Source"},
            ),
        ]

    for label, df, grade_cols, align_left_cols in sections:
        if df is None or df.empty:
            continue
        row = write_section(
            ws_core,
            row,
            label,
            df,
            grade_cols=grade_cols,
            num_cols={col for col in df.columns if pd.api.types.is_numeric_dtype(df[col])},
            align_left_cols=align_left_cols,
        )

    if (
        raw_linearity_df is not None and not raw_linearity_df.empty
    ) or (
        raw_grr_df is not None and not raw_grr_df.empty
    ):
        ws_raw = wb.create_sheet(f"{sheet_prefix} Raw")
        title_row(ws_raw, 1, f"{sheet_prefix} Raw")
        row = 2
        if raw_linearity_df is not None and not raw_linearity_df.empty:
            linearity_export = compact_df(
                raw_linearity_df,
                rename_map={"Value": "Data"},
                keep_cols=["Reference", "Value", "Bias"],
            )
            row = write_section(
                ws_raw,
                row,
                "Linearity Raw",
                linearity_export,
                num_cols={col for col in linearity_export.columns if pd.api.types.is_numeric_dtype(linearity_export[col])},
            )
        if raw_grr_df is not None and not raw_grr_df.empty:
            grr_export = compact_df(
                raw_grr_df,
                rename_map={"Operator": "Op", "Value": "Gdata"},
                keep_cols=["Part", "Operator", "Value"],
            )
            row = write_section(
                ws_raw,
                row,
                "Gage R&R Raw",
                grr_export,
                num_cols={col for col in grr_export.columns if pd.api.types.is_numeric_dtype(grr_export[col])},
                align_left_cols={"Part", "Op"},
            )
        auto_width(ws_raw, min_width=10, max_width=18)

    auto_width(ws_summary, min_width=10, max_width=20)
    auto_width(ws_core, min_width=10, max_width=22)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def create_all_studies_excel(ss) -> bytes | None:
    """Study-1~6 통합 Excel 보고서 생성. ss = st.session_state"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    # ── 공통 스타일 헬퍼 ──────────────────────────────
    C = {
        "title": "4472C4", "sub": "5B9BD5", "head": "D9E1F2",
        "ok": "E2EFDA",    "warn": "FFF2CC", "ng": "FCE4D6", "white": "FFFFFF",
    }
    FN = "맑은 고딕"
    thin = Side(style="thin", color="BFBFBF")

    def fill(c):          return PatternFill("solid", fgColor=c)
    def font(bold=False, color="000000", size=10):
        return Font(name=FN, bold=bold, color=color, size=size)
    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def border():
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    def grade_fill(val):
        v = str(val)
        if v in ("적합", "우수", "안정 (Stable)", "Stable"): return fill(C["ok"])
        if v in ("조건부 채택", "보통"):  return fill(C["warn"])
        if v in ("부적합", "미흡", "불안정 (Unstable)", "Unstable"): return fill(C["ng"])
        return None

    def pick_dataframe(*candidates):
        """Prefer a populated DataFrame without triggering pandas truth-value errors."""
        first_df = None
        for candidate in candidates:
            if isinstance(candidate, pd.DataFrame):
                if first_df is None:
                    first_df = candidate
                if not candidate.empty:
                    return candidate
        return first_df

    def study23_metric_fill(metric, value):
        if not isinstance(value, (int, float, np.integer, np.floating)) or pd.isna(value):
            return None
        numeric = float(value)
        if metric == "NDC":
            return fill(C["ok"]) if numeric >= 5 else fill(C["ng"])
        if metric == "Q_MP (%)":
            if numeric <= 15:
                return fill(C["ok"])
            if numeric <= 30:
                return fill(C["warn"])
            return fill(C["ng"])
        if metric == "C_MP":
            return fill(C["ok"]) if numeric >= 1.33 else fill(C["ng"])
        return None

    def study23_variation_fill(column, value, _row_data=None):
        if column not in {"%공차", "%SV", "Tol%", "SV%"}:
            return None
        if not isinstance(value, (int, float, np.integer, np.floating)) or pd.isna(value):
            return None
        numeric = float(value)
        if numeric <= 10:
            return fill(C["ok"])
        if numeric <= 30:
            return fill(C["warn"])
        return fill(C["ng"])

    def judged_row_fill(judge_cols, highlight_cols):
        judge_cols = tuple(judge_cols)
        highlight_cols = set(highlight_cols)

        def _fill(column, _value, row_data):
            verdict = None
            for judge_col in judge_cols:
                if judge_col in row_data.index:
                    verdict = row_data[judge_col]
                    break
            if verdict is None:
                return None
            return grade_fill(str(verdict)) if column in highlight_cols else None

        return _fill

    def study4_coeff_fill(_column, _value, row_data):
        coef_col = "계수" if "계수" in row_data.index else ("Coef" if "Coef" in row_data.index else None)
        p_col = "P" if "P" in row_data.index else None
        term_col = "예측 변수" if "예측 변수" in row_data.index else ("Term" if "Term" in row_data.index else None)
        if term_col is not None:
            term_text = str(row_data[term_col])
        else:
            term_text = str(row_data.iloc[0]) if len(row_data.index) else ""

        verdict_fill = None
        row_pos = row_data.name if isinstance(row_data.name, (int, np.integer)) else None

        if ("Intercept" in term_text or "절편" in term_text or row_pos == 0) and p_col is not None:
            p_value = pd.to_numeric(pd.Series([row_data[p_col]]), errors="coerce").iloc[0]
            if not pd.isna(p_value):
                verdict_fill = fill(C["ok"]) if float(p_value) > 0.05 else fill(C["ng"])
        elif ("Slope" in term_text or "기울기" in term_text or "Reference" in term_text or row_pos == 1) and coef_col is not None:
            coef_value = pd.to_numeric(pd.Series([row_data[coef_col]]), errors="coerce").iloc[0]
            if not pd.isna(coef_value):
                verdict_fill = fill(C["ok"]) if abs(float(coef_value) - 1.0) <= 0.05 else fill(C["ng"])
        return verdict_fill

    def study4_bias_fill(_column, _value, row_data):
        p_col = "P" if "P" in row_data.index else None
        if p_col is None:
            return None
        p_value = pd.to_numeric(pd.Series([row_data[p_col]]), errors="coerce").iloc[0]
        if pd.isna(p_value):
            return None
        return fill(C["ok"]) if float(p_value) > 0.05 else fill(C["ng"])

    def title_row(ws, row, text, span, color=C["title"], size=12):
        ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
        c = ws.cell(row, 1, text)
        c.font = font(bold=True, color=C["white"], size=size)
        c.fill = fill(color); c.alignment = align()
        ws.row_dimensions[row].height = 22

    def subhead(ws, row, text, span, color=C["sub"]):
        ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
        c = ws.cell(row, 1, text)
        c.font = font(bold=True, color=C["white"], size=10)
        c.fill = fill(color); c.alignment = align(h="left")

    def head_row(ws, row, headers, col_start=1):
        for ci, h in enumerate(headers, col_start):
            c = ws.cell(row, ci, h)
            c.font = font(bold=True); c.fill = fill(C["head"])
            c.alignment = align(); c.border = border()

    def data_cell(ws, row, col, val, bold=False, gfill=None, num_fmt=None, halign="center"):
        c = ws.cell(row, col, val)
        c.font = font(bold=bold); c.alignment = align(h=halign)
        c.border = border()
        if gfill: c.fill = gfill
        if num_fmt: c.number_format = num_fmt
        return c

    def df_to_sheet(ws, df, start_row, grade_cols=None, num_cols=None, align_left_cols=None, cell_fill_fn=None):
        head_row(ws, start_row, list(df.columns))
        grade_cols  = grade_cols  or set()
        num_cols    = num_cols    or set()
        align_left_cols = align_left_cols or set()
        r = start_row + 1
        for _, row_data in df.iterrows():
            for ci, col in enumerate(df.columns, 1):
                val = row_data[col]
                gf = cell_fill_fn(col, val, row_data) if callable(cell_fill_fn) else (grade_fill(str(val)) if col in grade_cols else None)
                nf = "0.0000" if col in num_cols else None
                ha = "left" if col in align_left_cols else "center"
                data_cell(ws, r, ci, val, bold=(gf is not None),
                          gfill=gf, num_fmt=nf, halign=ha)
            r += 1
        return r

    def autowidth(ws, widths):
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    # ── 데이터 수집 및 분석 재실행 ─────────────────────
    study_status = {}

    # Study-1
    s1_result = ss.get("s1_result")
    study_status["Study-1 (Bias)"] = "완료" if s1_result is not None else "미실행"

    # Study-2/3
    s23_t1 = ss.get("s23_type1_result")
    s23_t2 = ss.get("s23_type2_result")
    study_status["Study-2/3 (Precision)"] = "완료" if (s23_t1 or s23_t2) else "미실행"

    # Study-4 – 저장된 파라미터로 재분석
    s4_result = None
    s4_df_raw = ss.get("s4_df")
    if s4_df_raw is not None and not (isinstance(s4_df_raw, pd.DataFrame) and s4_df_raw.empty):
        try:
            _s4_cfg = _get_study4_reference_sample()
            s4_result = run_study4_analysis(
                normalize_study4_df(s4_df_raw),
                float(ss.get("s4_usl", _s4_cfg["USL"])),
                float(ss.get("s4_lsl", _s4_cfg["LSL"])),
                float(ss.get("s4_ucal", _s4_cfg["u_CAL"])),
                float(ss.get("s4_re",  _s4_cfg["RE"])),
                grr_df=normalize_study2_df(ss["s4_grr_df"]) if ss.get("s4_grr_df") is not None else None,
            )
            study_status["Study-4 (Linearity)"] = "완료"
        except Exception:
            study_status["Study-4 (Linearity)"] = "오류"
    else:
        study_status["Study-4 (Linearity)"] = "미실행"

    # Study-5 – 재분석
    s5_result = None
    s5_df_raw = ss.get("s5_df")
    if s5_df_raw is not None and not (isinstance(s5_df_raw, pd.DataFrame) and s5_df_raw.empty):
        try:
            s5_result = run_study5_analysis(normalize_study5_df(s5_df_raw))
            study_status["Study-5 (Stability)"] = "완료"
        except Exception:
            study_status["Study-5 (Stability)"] = "오류"
    else:
        study_status["Study-5 (Stability)"] = "미실행"

    # Study-6 – 재분석
    s6_res_df = None; s6_fk = None; s6_raw = None
    s6_df_raw = ss.get("s6_df")
    if s6_df_raw is not None and not (isinstance(s6_df_raw, pd.DataFrame) and s6_df_raw.empty):
        try:
            s6_raw = normalize_study6_df(s6_df_raw)
            s6_res_df = run_study6_analysis(s6_raw).rename(
                columns={"Kappa": "Cohen's κ", "AIAG 판정": "AIAG 판정(κ)", "미니탭 판정": "미니탭 판정(κ)"}
            )
            s6_fk = calc_fleiss_kappa(s6_raw)
            study_status["Study-6 (Attribute)"] = "완료"
        except Exception:
            study_status["Study-6 (Attribute)"] = "오류"
    else:
        study_status["Study-6 (Attribute)"] = "미실행"

    # ══════════════════════════════════════════════════
    # Sheet 0 : 통합 요약 (Cover)
    # ══════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    ws0 = wb.active; ws0.title = "통합 요약"
    title_row(ws0, 1, "MSA 통합 분석 보고서 (Study 1~6)  |  AIAG MSA 4th Edition", 4, size=13)
    ws0.cell(2, 1).value = f"작성 일시: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws0.cell(2, 1).font  = font(size=9); ws0.merge_cells("A2:D2")
    ws0.cell(2, 1).alignment = align(h="right")

    metadata_items = _report_metadata_items(_get_report_metadata(ss))

    r0 = 4
    if metadata_items:
        head_row(ws0, r0, ["Item", "Value"])
        r0 += 1
        for label, value in metadata_items:
            data_cell(ws0, r0, 1, label, bold=True, halign="left")
            data_cell(ws0, r0, 2, value, halign="left")
            r0 += 1
        r0 += 1
    head_row(ws0, r0, ["Study", "분석 항목", "상태", "비고"])
    r0 += 1
    study_desc = {
        "Study-1 (Bias)":         "측정기 치우침 (Bias / 교정 능력)",
        "Study-2/3 (Precision)":  "반복성·재현성 (Gage R&R / 불확도)",
        "Study-4 (Linearity)":    "선형성 (Linearity Study)",
        "Study-5 (Stability)":    "안정성 관리도 (X-bar & R Chart)",
        "Study-6 (Attribute)":    "계수형 합치도 (Attribute Gauge R&R)",
    }
    for study, status in study_status.items():
        gf = fill(C["ok"]) if status == "완료" else (fill(C["warn"]) if status == "오류" else None)
        data_cell(ws0, r0, 1, study, bold=True)
        data_cell(ws0, r0, 2, study_desc.get(study, ""), halign="left")
        data_cell(ws0, r0, 3, status, bold=True, gfill=gf)
        data_cell(ws0, r0, 4, "해당 시트 참조" if status == "완료" else "분석을 먼저 실행하세요")
        r0 += 1
    autowidth(ws0, [22, 36, 10, 24])

    # ══════════════════════════════════════════════════
    # Sheet 1 : Study-1
    # ══════════════════════════════════════════════════
    if s1_result is not None:
        ws1 = wb.create_sheet("Study-1 Bias")
        title_row(ws1, 1, "Study-1 : 치우침 및 측정 능력 (Bias / Calibration Capability)", 5)
        subhead(ws1, 3, "[ 분석 결과 ]", 5)
        grade_c = {col for col in s1_result.columns if "판정" in str(col)}
        num_c   = {col for col in s1_result.columns if s1_result[col].dtype in [float, "float64"]}
        df_to_sheet(ws1, s1_result, 4, grade_cols=grade_c, num_cols=num_c,
                    align_left_cols={"항목"},
                    cell_fill_fn=judged_row_fill({"판정"}, {"값", "기준", "판정"}))
        autowidth(ws1, [28, 14, 14, 14, 14])

    # ══════════════════════════════════════════════════
    # Sheet 2 : Study-2/3
    # ══════════════════════════════════════════════════
    if s23_t1 or s23_t2:
        ws2 = wb.create_sheet("Study-2_3 Precision")
        title_row(ws2, 1, "Study-2/3 : 반복성·재현성 (Gage R&R / Measurement Uncertainty)", 7)
        r2 = 3
        for label, tres in [("Type I (반복성)", s23_t1), ("Type II (Gage R&R)", s23_t2)]:
            if tres is None:
                continue
            subhead(ws2, r2, f"[ {label} ]", 7); r2 += 1
            # 핵심 지표
            metrics = [
                ("NDC",   tres.get("ndc",  "—")),
                ("Q_MP (%)", tres.get("q_mp", "—")),
                ("C_MP",  tres.get("c_mp", "—")),
            ]
            head_row(ws2, r2, ["지표", "값", "판정 기준"])
            r2 += 1
            crit = {"NDC": "≥ 5 : 우수", "Q_MP (%)": "≤ 15 % : 우수", "C_MP": "≥ 1.33 : 우수"}
            for name, val in metrics:
                v = round(float(val), 4) if isinstance(val, (int, float)) and not isinstance(val, bool) else val
                metric_fill = study23_metric_fill(name, v)
                data_cell(ws2, r2, 1, name, bold=True)
                data_cell(ws2, r2, 2, v, bold=(metric_fill is not None), gfill=metric_fill, num_fmt="0.0000" if isinstance(v, float) else None)
                data_cell(ws2, r2, 3, crit.get(name, ""), bold=(metric_fill is not None), gfill=metric_fill, halign="left")
                r2 += 1
            r2 += 1
            # Gage 분산 테이블
            gage_df = tres.get("gage")
            if gage_df is not None and not gage_df.empty:
                subhead(ws2, r2, "Gage 분산 분석표", 7, color=C["head"]); r2 += 1
                num_c2 = {c for c in gage_df.columns if gage_df[c].dtype in [float, "float64"]}
                r2 = df_to_sheet(ws2, gage_df, r2, num_cols=num_c2, align_left_cols={"출처"}, cell_fill_fn=study23_variation_fill)
            r2 += 2
        autowidth(ws2, [24, 12, 12, 12, 12, 12, 12])

    # ══════════════════════════════════════════════════
    # Sheet 3 : Study-4
    # ══════════════════════════════════════════════════
    if s4_result is not None:
        ws4 = wb.create_sheet("Study-4 Linearity")
        title_row(ws4, 1, "Study-4 : 선형성 (Linearity Study)", 6)
        r4 = 3
        aiag = s4_result.get("aiag_linearity", {})
        # 종합 판정
        lpass = "적합" if s4_result.get("linearity_pass") else "부적합"
        subhead(ws4, r4, "[ 종합 판정 ]", 6); r4 += 1
        data_cell(ws4, r4, 1, "선형성 판정", bold=True)
        data_cell(ws4, r4, 2, lpass, bold=True, gfill=grade_fill(lpass))
        ws4.merge_cells(f"B{r4}:F{r4}"); r4 += 2

        aiag_judge = _study4_aiag_judgement(s4_result)
        if aiag_judge is not None and not aiag_judge.empty:
            subhead(ws4, r4, "[ 판정 기준 ]", 6); r4 += 1
            r4 = df_to_sheet(ws4, aiag_judge, r4,
                             grade_cols={"판정"},
                             num_cols={c for c in aiag_judge.columns if aiag_judge[c].dtype in [float, "float64"]},
                             align_left_cols={"항목", "기준"},
                             cell_fill_fn=judged_row_fill({"판정"}, {"값", "기준", "판정"}))
            r4 += 1

        # 회귀 계수
        coeff = pick_dataframe(aiag.get("coefficients"), aiag.get("coeff_table"))
        if coeff is not None and not coeff.empty:
            subhead(ws4, r4, "[ 회귀 계수 ]", 6); r4 += 1
            r4 = df_to_sheet(ws4, coeff, r4,
                             num_cols={c for c in coeff.columns if coeff[c].dtype in [float, "float64"]},
                             align_left_cols={"예측 변수"},
                             cell_fill_fn=study4_coeff_fill)
            r4 += 1
        # 치우침 요약
        bias_sum = aiag.get("bias_summary")
        if bias_sum is not None and not bias_sum.empty:
            subhead(ws4, r4, "[ 기준값별 치우침 요약 ]", 6); r4 += 1
            r4 = df_to_sheet(ws4, bias_sum, r4,
                             num_cols={c for c in bias_sum.columns if bias_sum[c].dtype in [float, "float64"]},
                             align_left_cols={"기준"},
                             cell_fill_fn=study4_bias_fill)
        autowidth(ws4, [18, 14, 14, 14, 14, 14])

    # ══════════════════════════════════════════════════
    # Sheet 4 : Study-5
    # ══════════════════════════════════════════════════
    if s5_result is not None:
        ws5 = wb.create_sheet("Study-5 Stability")
        title_row(ws5, 1, "Study-5 : 안정성 (Stability — X-bar & R 관리도)", 4)
        stable = "안정 (Stable)" if s5_result.get("is_stable") else "불안정 (Unstable)"
        r5 = 3
        subhead(ws5, r5, "[ 종합 판정 ]", 4); r5 += 1
        data_cell(ws5, r5, 1, "안정성 판정", bold=True)
        data_cell(ws5, r5, 2, stable, bold=True, gfill=grade_fill(stable))
        ws5.merge_cells(f"B{r5}:D{r5}"); r5 += 2

        summ = s5_result.get("summary")
        if summ is not None and not summ.empty:
            subhead(ws5, r5, "[ 관리도 통계량 ]", 4); r5 += 1
            r5 = df_to_sheet(ws5, summ, r5,
                             num_cols={c for c in summ.columns if summ[c].dtype in [float, "float64"]},
                             align_left_cols={"항목"})
        # 이상점 정보
        out_x = s5_result.get("out_x", [])
        out_r = s5_result.get("out_r", [])
        x_bars = s5_result.get("x_bars", [])
        ranges = s5_result.get("ranges", [])
        r5 += 1
        subhead(ws5, r5, "[ 부분군별 데이터 ]", 4); r5 += 1
        head_row(ws5, r5, ["부분군", "X-bar", "Range", "X-bar 이상", "Range 이상"])
        r5 += 1
        for i, (xb, rng, ox, orng) in enumerate(zip(x_bars, ranges, out_x, out_r), 1):
            gf_x   = fill(C["ng"]) if ox   else None
            gf_r   = fill(C["ng"]) if orng else None
            data_cell(ws5, r5, 1, i)
            data_cell(ws5, r5, 2, round(float(xb),  4), num_fmt="0.0000", gfill=gf_x)
            data_cell(ws5, r5, 3, round(float(rng), 4), num_fmt="0.0000", gfill=gf_r)
            data_cell(ws5, r5, 4, "이상" if ox   else "정상", gfill=gf_x)
            data_cell(ws5, r5, 5, "이상" if orng else "정상", gfill=gf_r)
            r5 += 1
        autowidth(ws5, [10, 12, 12, 12, 12])

    # ══════════════════════════════════════════════════
    # Sheet 5 : Study-6
    # ══════════════════════════════════════════════════
    if s6_res_df is not None:
        create_study6_excel(s6_res_df, s6_fk, s6_raw)   # 기존 함수로 시트 내용 참조 불가 — 별도 시트 직접 작성
        ws6 = wb.create_sheet("Study-6 Attribute")
        title_row(ws6, 1, "Study-6 : 계수형 합치도 (Attribute Gauge R&R)", 8)
        r6 = 3
        # 평가자별 결과
        subhead(ws6, r6, "[ 평가자별 분석 결과 ]", 8); r6 += 1
        display_c = ["평가자", "유효성(E)", "P(FA)", "P(miss)", "Cohen's κ",
                     "AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"]
        grade_c6 = {"AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"}
        num_c6   = {"유효성(E)", "P(FA)", "P(miss)", "Cohen's κ"}
        r6 = df_to_sheet(ws6, s6_res_df[display_c], r6, grade_cols=grade_c6, num_cols=num_c6,
                         cell_fill_fn=judged_row_fill({"종합 판정"}, {"유효성(E)", "P(FA)", "P(miss)", "Cohen's κ", "AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"}))
        r6 += 1
        # Fleiss kappa
        if s6_fk and not (s6_fk.get("fleiss_kappa") != s6_fk.get("fleiss_kappa")):
            subhead(ws6, r6, "[ Fleiss' κ — 평가자 간 상호 일치도 ]", 8); r6 += 1
            fk_headers = ["Fleiss' κ", "AIAG 판정", "미니탭 판정", "평가자 수", "부품 수"]
            fk_values  = [round(s6_fk["fleiss_kappa"], 4), s6_fk["aiag_grade"],
                          s6_fk["minitab_grade"], s6_fk["n_raters"], s6_fk["n_subjects"]]
            head_row(ws6, r6, fk_headers); r6 += 1
            for ci, (h, v) in enumerate(zip(fk_headers, fk_values), 1):
                gf = grade_fill(str(v)) if h in {"AIAG 판정", "미니탭 판정"} else None
                data_cell(ws6, r6, ci, v, gfill=gf,
                          num_fmt="0.0000" if h == "Fleiss' κ" else None)
            r6 += 2
        # 계산식 상세
        subhead(ws6, r6, "[ 계산식 상세 ]", 8); r6 += 1
        calc_c = ["평가자", "유효성 계산", "P(FA) 계산", "P(miss) 계산", "AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"]
        r6 = df_to_sheet(ws6, s6_res_df[calc_c], r6,
                         grade_cols={"AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"},
                         align_left_cols={"유효성 계산", "P(FA) 계산", "P(miss) 계산"},
                         cell_fill_fn=judged_row_fill({"종합 판정"}, {"AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"}))
        autowidth(ws6, [14, 14, 10, 10, 12, 14, 14, 12])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def render_combined_report(level):
    st.header("📋 통합 MSA 보고서 (Study 1~6)")
    st.info("각 Study 분석을 실행한 후 이 페이지에서 전체 결과를 하나의 Excel 파일로 다운로드할 수 있습니다.")

    # 각 Study 상태 미리보기
    status_data = []
    checks = {
        "Study-1 (Bias)":        (lambda ss: ss.get("s1_result") is not None, "치우침 분석"),
        "Study-2/3 (Precision)": (
            lambda ss: ss.get("s23_type1_result") is not None or ss.get("s23_type2_result") is not None,
            "반복성·재현성",
        ),
        "Study-4 (Linearity)":   (lambda ss: ss.get("s4_df") is not None, "선형성 분석"),
        "Study-5 (Stability)":   (lambda ss: ss.get("s5_df") is not None, "안정성 관리도"),
        "Study-6 (Attribute)":   (lambda ss: ss.get("s6_df") is not None, "계수형 합치도"),
    }
    for study, (predicate, desc) in checks.items():
        has = bool(predicate(st.session_state))
        status_data.append({"Study": study, "항목": desc, "상태": "✅ 완료" if has else "⬜ 미실행"})

    st.dataframe(pd.DataFrame(status_data), use_container_width=True, hide_index=True)

    completed = sum(1 for d in status_data if "완료" in d["상태"])
    st.markdown(f"**{completed} / {len(status_data)}** 개 Study 완료")

    if completed == 0:
        st.warning("최소 1개의 Study 분석을 먼저 실행하세요.")
        return

    combined_ai_payload = build_ai_combined_payload(st.session_state)
    _render_ai_explainer_card("combined", combined_ai_payload, title="통합 AI 요약")

    if st.button("📊 통합 Excel 보고서 생성", type="primary", use_container_width=True):
        with st.spinner("보고서 생성 중..."):
            excel_data = create_all_studies_excel(st.session_state)
        if excel_data:
            fname = f"MSA_통합보고서_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                "💾 Excel 다운로드", excel_data, fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.error("openpyxl 패키지가 필요합니다. pip install openpyxl")


# ─────────────────────────────────────────
# 메인 / 사이드바
# ─────────────────────────────────────────
def _uploaded_file_signature(uploaded_file):
    if uploaded_file is None:
        return None
    return (uploaded_file.name, getattr(uploaded_file, "size", None))


def _store_uploaded_data_once(uploaded_file, signature_key, state_key, normalizer):
    if uploaded_file is None:
        return
    signature = _uploaded_file_signature(uploaded_file)
    if st.session_state.get(signature_key) == signature:
        return
    st.session_state[state_key] = load_and_normalize_data(uploaded_file, normalizer)
    st.session_state[signature_key] = signature


def _store_uploaded_study1_values(uploaded_file):
    if uploaded_file is None:
        return
    signature = _uploaded_file_signature(uploaded_file)
    if st.session_state.get("s1_upload_sig") == signature:
        return
    values = load_and_normalize_data(uploaded_file, normalize_study1_values)
    st.session_state["s1_data"] = ", ".join(map(str, values))
    st.session_state["s1_upload_sig"] = signature


_STUDY23_TOOLTIP_CATALOG = {
    "panel.type1": {
        "AIAG": "Type I 반복측정값과 기본 규격을 입력하는 영역입니다. AIAG 모드에서는 핵심 입력만 사용합니다.",
        "ISO 22514-7": "Type I 측정기 능력 계산에 필요한 반복측정값과 기본 불확도 입력 영역입니다.",
    },
    "panel.tol_mode": {
        "AIAG": "USL/LSL을 직접 넣거나 공차폭(TOL)만 넣어 계산할 수 있습니다.",
        "ISO 22514-7": "규격 상한·하한 또는 공차폭을 기준으로 Study-2/3 계산 공차를 정의합니다.",
    },
    "panel.typeb": {
        "AIAG": "AIAG 모드에서는 사용하지 않는 ISO 보조 불확도 입력입니다.",
        "ISO 22514-7": "ISO 22514-7 계산에 반영할 추가 Type B 불확도 항목입니다.",
    },
    "panel.file": {
        "AIAG": "Type II Gage R&R 분석용 데이터를 업로드하거나 Figure 6 샘플을 사용할 수 있습니다.",
        "ISO 22514-7": "Type II Gage R&R 분석용 원시 데이터 소스입니다. 업로드 파일 또는 Figure 6 샘플을 사용합니다.",
    },
    "panel.run_scope": {
        "AIAG": "Type 1, Type 2, 동시 실행 중 하나를 선택합니다. 일반 업로드 분석은 Type 2만 실행이 보통입니다.",
        "ISO 22514-7": "Type I, Type II, 동시 실행 중 분석 범위를 선택합니다.",
    },
    "panel.unc_chart": {
        "AIAG": "AIAG 모드에서는 이 불확도 상세 차트를 표시하지 않습니다.",
        "ISO 22514-7": "0이 아닌 불확도 구성요소를 크기 순으로 비교하는 차트입니다.",
    },
    "panel.gage_chart": {
        "AIAG": "총 Gage R&R, 반복성, 재현성, 부품 간 변동을 한눈에 비교합니다.",
        "ISO 22514-7": "Gage R&R 주요 변동요소의 표준편차를 비교하는 차트입니다.",
    },
    "panel.result": {
        "AIAG": "핵심 판정값과 분석 요약을 확인하는 영역입니다.",
        "ISO 22514-7": "전체 판정, 핵심 지표, 분석 리포트를 확인하는 영역입니다.",
    },
    "field.re": {
        "AIAG": "측정기의 최소 눈금 또는 분해능입니다.",
        "ISO 22514-7": "측정시스템 분해능에 해당하는 입력값으로 u_RE 계산의 기준이 됩니다.",
    },
    "field.refv": {
        "AIAG": "기준이 되는 참조값입니다.",
        "ISO 22514-7": "Type I 반복측정의 기준 참조값입니다.",
    },
    "field.usl": {
        "AIAG": "규격 상한값입니다.",
        "ISO 22514-7": "허용 규격의 상한으로 공차 계산에 사용됩니다.",
    },
    "field.lsl": {
        "AIAG": "규격 하한값입니다.",
        "ISO 22514-7": "허용 규격의 하한으로 공차 계산에 사용됩니다.",
    },
    "field.ucal": {
        "AIAG": "AIAG 화면에는 노출하지 않지만 내부 계산 기본값은 유지합니다.",
        "ISO 22514-7": "교정성적서 기반 표준 불확도 입력값입니다.",
    },
    "field.bias_dist": {
        "AIAG": "AIAG 모드에서는 기본 분포를 내부적으로 사용합니다.",
        "ISO 22514-7": "Bias 항목을 표준불확도로 환산할 때 적용할 분포 형태입니다.",
    },
    "field.re_dist": {
        "AIAG": "AIAG 모드에서는 기본 분포를 내부적으로 사용합니다.",
        "ISO 22514-7": "분해능 RE 값을 표준불확도로 환산할 때 적용할 분포 형태입니다.",
    },
    "field.type1_values": {
        "AIAG": "한 기준값을 반복 측정한 값을 입력합니다. 콤마 또는 줄바꿈으로 구분할 수 있습니다.",
        "ISO 22514-7": "Type I 반복측정값입니다. 콤마 또는 줄바꿈으로 구분해 입력합니다.",
    },
    "field.tol_mode": {
        "AIAG": "규격 상하한을 직접 넣거나 공차폭만 넣어 계산할 수 있습니다.",
        "ISO 22514-7": "공차 입력 방식을 선택합니다. 직접입력 또는 TOL 기반 계산이 가능합니다.",
    },
    "field.tol": {
        "AIAG": "USL-LSL 대신 사용할 공차폭입니다.",
        "ISO 22514-7": "중심값 기준으로 상하한을 계산할 때 사용하는 전체 공차폭입니다.",
    },
    "field.file_upload": {
        "AIAG": "파일 형식은 [부품, 측정자, 측정값] 3개 열입니다.",
        "ISO 22514-7": "파일 형식은 [부품, 측정자, 측정값] 3개 열이며 CSV/Excel을 지원합니다.",
    },
    "field.file_name": {
        "AIAG": "현재 분석에 연결된 파일 또는 샘플 데이터 이름입니다.",
        "ISO 22514-7": "현재 Type II 분석에 사용될 데이터 소스 이름입니다.",
    },
    "field.run_scope": {
        "AIAG": "Type 1, Type 2, 동시 실행 중 하나를 선택합니다.",
        "ISO 22514-7": "Type I과 Type II 중 어떤 분석을 실행할지 선택합니다.",
    },
    "action.run": {
        "AIAG": "선택한 실행 범위로 분석을 시작합니다.",
        "ISO 22514-7": "현재 입력값과 데이터로 Study-2/3 계산을 실행합니다.",
    },
    "action.save_report": {
        "AIAG": "현재 결과를 AIAG 기준 Excel 파일로 저장합니다.",
        "ISO 22514-7": "현재 분석 결과를 ISO 22514-7 기준 Excel 파일로 저장합니다.",
    },
    "action.save_json": {
        "AIAG": "현재 결과 데이터를 JSON 형식으로 저장합니다.",
        "ISO 22514-7": "현재 분석 결과와 계산 데이터를 JSON으로 저장합니다.",
    },
    "action.print": {
        "AIAG": "현재 화면을 인쇄 또는 PDF로 저장합니다.",
        "ISO 22514-7": "현재 화면을 인쇄하거나 PDF로 저장합니다.",
    },
    "action.sample": {
        "AIAG": "Figure 6 예제값을 불러와 바로 Type 2 분석에 사용할 수 있습니다.",
        "ISO 22514-7": "문헌 Figure 6 기반 샘플 데이터를 불러옵니다.",
    },
    "metric.overall": {
        "AIAG": "전체 판정은 NDC, %SV, %공차 기준을 종합해 합격/조건부/불합격으로 결정합니다.",
        "ISO 22514-7": "전체 판정은 핵심 판단 지표를 종합해 합격/조건부/불합격으로 결정합니다.",
    },
    "metric.ndc": {
        "AIAG": "측정시스템이 구분 가능한 범주의 수입니다. 일반적으로 5 이상이면 사용 가능합니다.",
        "ISO 22514-7": "NDC(Number of Distinct Categories)는 측정시스템이 구분 가능한 실질 범주 수입니다.",
    },
    "metric.q_mp": {
        "AIAG": "측정 프로세스 확장불확도가 공차에서 차지하는 비율입니다. 작을수록 좋습니다.",
        "ISO 22514-7": "Q_MP는 2U_MP/TOL 비율로, 측정 프로세스 확장불확도의 공차 대비 비율입니다.",
    },
    "metric.pct_sv": {
        "AIAG": "Gage R&R 변동이 전체 변동에서 차지하는 비율입니다.",
        "ISO 22514-7": "%SV는 Gage R&R 표준편차가 총 변동 대비 차지하는 비율입니다.",
    },
    "metric.pct_tol": {
        "AIAG": "Gage R&R가 공차 폭에서 차지하는 비율입니다.",
        "ISO 22514-7": "%공차는 Gage R&R 6시그마 폭이 공차 대비 차지하는 비율입니다.",
    },
    "metric.c_mp": {
        "AIAG": "측정 프로세스 능력 지표입니다. 값이 클수록 좋습니다.",
        "ISO 22514-7": "C_MP는 측정 프로세스 능력 지표로, 측정 불확도 대비 공차 여유를 나타냅니다.",
    },
}


def _study23_tooltip(level, key):
    mode_key = "ISO 22514-7" if _is_iso_mode(level) else "AIAG"
    entry = _STUDY23_TOOLTIP_CATALOG.get(key, "")
    if isinstance(entry, dict):
        return entry.get(mode_key) or entry.get("AIAG") or entry.get("ISO 22514-7") or ""
    return str(entry)


def _study23_panel_title(title, tooltip_text=""):
    if not tooltip_text:
        st.markdown(f"<div class='study23-panel-title'>{html.escape(title)}</div>", unsafe_allow_html=True)
        return
    safe_title = html.escape(title)
    safe_tooltip = html.escape(tooltip_text).replace("\n", "<br>")
    st.markdown(
        f"""
        <div class="study23-title-row">
            <div class="study23-panel-title">{safe_title}</div>
            <span class="study23-help-icon" tabindex="0">i
                <span class="study23-help-bubble">{safe_tooltip}</span>
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _study23_judgement_value(type2_result, item):
    if type2_result is None:
        return np.nan, ""
    judgement = type2_result.get("judgement")
    if judgement is None or judgement.empty:
        return np.nan, ""
    matched = judgement.loc[judgement["항목"] == item]
    if matched.empty:
        return np.nan, ""
    row = matched.iloc[0]
    return row.get("값", np.nan), str(row.get("판정", ""))


def _study23_status_kind(status):
    label = str(status).strip()
    if label in {"OK", "양호", "적합"}:
        return "ok"
    if label in {"WARN", "조건부", "조건부 채택"}:
        return "warn"
    if label in {"NG", "개선 필요", "부적합", "미달"}:
        return "ng"
    return ""


def _study23_overall_status(judgement, fallback_pass=False):
    if judgement is None or judgement.empty:
        return "합격" if fallback_pass else "불합격"

    severities = []
    for verdict in judgement["판정"]:
        kind = _study23_status_kind(verdict)
        severities.append({"ok": 0, "warn": 1, "ng": 2}.get(kind, 0))

    max_severity = max(severities, default=0)
    if max_severity == 0:
        return "합격"
    if max_severity == 1:
        return "조건부 채택"
    return "불합격"


def _study23_visible_judgement(type2_result, is_expert):
    if type2_result is None:
        return pd.DataFrame(columns=["항목", "값", "기준", "판정"])
    judgement = type2_result.get("judgement")
    if judgement is None or judgement.empty:
        return pd.DataFrame(columns=["항목", "값", "기준", "판정"])

    display = judgement.copy()
    if not is_expert:
        display = display[~display["항목"].isin(["Q_MP"])]
    return display.reset_index(drop=True)


def _study23_metric_cards(judgement, type2_result=None, is_expert=False):
    if judgement is None:
        return []
    if judgement is None or judgement.empty:
        cards = []
    else:
        cards = []
        for _, row in judgement.iterrows():
            status = str(row.get("판정", ""))
            status_kind = _study23_status_kind(status)
            cards.append({
                "metric_name": str(row.get("항목", "")),
                "metric_value": str(_format_display_value(row.get("값", np.nan))),
                "criterion": str(row.get("기준", "-")),
                "status": status,
                "card_state": f"is-{status_kind}" if status_kind else "",
            })

    if is_expert and type2_result is not None:
        c_mp = type2_result.get("c_mp", np.nan)
        if not pd.isna(c_mp):
            status = "적합" if float(c_mp) >= 1.33 else "부적합"
            status_kind = _study23_status_kind(status)
            cards.append({
                "metric_name": "C_MP",
                "metric_value": str(_format_display_value(c_mp)),
                "criterion": ">= 1.33",
                "status": status,
                "card_state": f"is-{status_kind}" if status_kind else "",
            })

    return cards


def _study23_metric_card_html(card):
    return f"""
    <div class="study23-metric-card {card['card_state']}">
        <div class="study23-metric-head">
            <div class="study23-metric-name">{html.escape(card['metric_name'])}</div>
            <div class="study23-metric-status {card['card_state']}">{html.escape(card['status'] or '-')}</div>
        </div>
        <div class="study23-metric-value">{html.escape(card['metric_value'])}</div>
        <div class="study23-metric-criterion">기준: {html.escape(card['criterion'])}</div>
    </div>
    """


def _style_study23_judgement(df):
    if df is None or df.empty:
        return df

    def _row_style(row):
        kind = _study23_status_kind(row.get("판정", ""))
        if kind == "ok":
            return ["background-color: rgba(46, 204, 113, 0.10); color: #1f5131; font-weight: 600;"] * len(row)
        if kind == "warn":
            return ["background-color: rgba(240, 173, 78, 0.12); color: #8a6d1d; font-weight: 600;"] * len(row)
        if kind == "ng":
            return ["background-color: rgba(231, 76, 60, 0.10); color: #7a1e18; font-weight: 600;"] * len(row)
        return [""] * len(row)

    return (
        df.style
        .format(lambda value: _format_display_value(value))
        .apply(_row_style, axis=1)
        .hide(axis="index")
    )


def _style_study23_gage(df):
    if df is None or df.empty:
        return df

    def _row_style(row):
        source = str(row.get("출처", ""))
        if source == "총 Gage R&R":
            return ["background-color: rgba(63, 127, 186, 0.10); font-weight: 700;"] * len(row)
        if source in {"반복성", "재현성", "측정자", "교호작용"}:
            return ["background-color: rgba(243, 164, 59, 0.10);"] * len(row)
        if source == "부품-대-부품":
            return ["background-color: rgba(90, 169, 107, 0.10);"] * len(row)
        if source == "총 변동":
            return ["background-color: rgba(148, 163, 184, 0.10); font-weight: 700;"] * len(row)
        return [""] * len(row)

    return (
        df.style
        .format(lambda value: _format_display_value(value))
        .apply(_row_style, axis=1)
        .hide(axis="index")
    )


def _style_study23_anova(df):
    if df is None or df.empty:
        return df

    def _row_style(row):
        label = str(row.get("항목", row.get("index", "")))
        if label in {"부품", "측정자", "교호작용"}:
            return ["background-color: rgba(63, 127, 186, 0.08);"] * len(row)
        if label in {"반복성", "총변동"}:
            return ["background-color: rgba(148, 163, 184, 0.08);"] * len(row)
        return [""] * len(row)

    return (
        df.style
        .format(lambda value: _format_display_value(value))
        .apply(_row_style, axis=1)
        .hide(axis="index")
    )


def _parse_measurement_text(raw_text):
    return np.array([float(x) for x in re.findall(r"[-+]?\d*\.?\d+", raw_text or "")], dtype=float)


def _summary_value(df_summary, label, default=np.nan):
    matched = df_summary.loc[df_summary["항목"] == label, "값"]
    if matched.empty:
        return default
    value = matched.iloc[0]
    return float(value) if isinstance(value, (int, float, np.integer, np.floating)) else value


DISPLAY_DECIMALS = 3


def _truncate_decimal(value, digits=DISPLAY_DECIMALS):
    if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
        factor = 10 ** digits
        return np.trunc(float(value) * factor) / factor
    return value


def _format_display_value(value, digits=DISPLAY_DECIMALS, suffix=""):
    if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
        return f"{float(value):.{digits}f}{suffix}"
    return value


def _format_display_df(df, digits=DISPLAY_DECIMALS):
    if df is None:
        return df
    display = df.copy()
    for col in display.columns:
        display[col] = display[col].map(lambda value: _format_display_value(value, digits))
    return display


def _style_s6_result(df):
    """Study-6 결과/계산식 테이블: 기준 초과 셀 빨강/노랑 하이라이트 (AIAG MSA 4판)"""
    RED    = "background-color: #fde8e8; color: #c0392b; font-weight: 700;"
    YELLOW = "background-color: #fff8e1; color: #856404;"
    GREEN  = "color: #27ae60; font-weight: 700;"
    res = pd.DataFrame("", index=df.index, columns=df.columns)

    # 수치 임계값 컬럼
    thresholds = {
        "유효성(E)":  lambda v: RED if v < 0.8 else (YELLOW if v < 0.9 else ""),
        "P(FA)":      lambda v: RED if v >= 0.10 else (YELLOW if v >= 0.05 else ""),
        "P(miss)":    lambda v: RED if v >= 0.05 else (YELLOW if v >= 0.02 else ""),
    }
    # "44/45=0.978" 형식 계산 텍스트 → 마지막 숫자 추출 후 동일 임계값 적용
    calc_thresholds = {
        "유효성 계산": thresholds["유효성(E)"],
        "P(FA) 계산":  thresholds["P(FA)"],
        "P(miss) 계산":thresholds["P(miss)"],
    }
    # 판정 텍스트 컬럼
    verdict_map = {
        "부적합": RED, "조건부 채택": YELLOW, "적합": GREEN,
    }
    kappa_map = {
        "우수": GREEN, "보통": YELLOW, "미흡": RED,
    }
    verdict_cols  = {"종합 판정", "유효성 판정", "P(FA) 판정", "P(miss) 판정"}
    kappa_cols    = {"AIAG 판정", "미니탭 판정", "AIAG 판정(κ)", "미니탭 판정(κ)"}

    for col in df.columns:
        if col in thresholds:
            for i, v in enumerate(df[col]):
                try:
                    res.at[df.index[i], col] = thresholds[col](float(v))
                except (TypeError, ValueError):
                    pass
        elif col in calc_thresholds:
            fn = calc_thresholds[col]
            for i, v in enumerate(df[col]):
                nums = re.findall(r"[-+]?\d*\.?\d+", str(v))
                if nums:
                    try:
                        res.at[df.index[i], col] = fn(float(nums[-1]))
                    except (TypeError, ValueError):
                        pass
        elif col in verdict_cols:
            for i, v in enumerate(df[col]):
                res.at[df.index[i], col] = verdict_map.get(str(v), "")
        elif col in kappa_cols:
            for i, v in enumerate(df[col]):
                res.at[df.index[i], col] = kappa_map.get(str(v), "")

    return (
        df.style
        .format(lambda v: _format_display_value(v))
        .apply(lambda _: res, axis=None)
        .hide(axis="index")
    )


def _style_verdict_dataframe(df, digits=DISPLAY_DECIMALS):
    if df is None or df.empty:
        return df

    verdict_col = df.columns[-1]

    def _row_style(row):
        kind = _study23_status_kind(row.get(verdict_col, ""))
        if kind == "ok":
            return ["background-color: rgba(46, 204, 113, 0.10); color: #1f5131; font-weight: 600;"] * len(row)
        if kind == "warn":
            return ["background-color: rgba(240, 173, 78, 0.12); color: #8a6d1d; font-weight: 600;"] * len(row)
        if kind == "ng":
            return ["background-color: rgba(231, 76, 60, 0.10); color: #7a1e18; font-weight: 600;"] * len(row)
        return [""] * len(row)

    return (
        df.style
        .format(lambda value: _format_display_value(value, digits))
        .apply(_row_style, axis=1)
        .hide(axis="index")
    )


def _study4_aiag_judgement(result):
    if not isinstance(result, dict):
        return pd.DataFrame(columns=["항목", "값", "기준", "판정"])

    summary = result.get("summary")
    if summary is None or summary.empty or "항목" not in summary.columns:
        return pd.DataFrame(columns=["항목", "값", "기준", "판정"])

    return (
        summary.loc[
            summary["항목"].isin(["Intercept", "Slope"]),
            ["항목", "값", "기준", "판정"],
        ]
        .copy()
        .reset_index(drop=True)
    )


def _style_study4_aiag_coefficients(df):
    if df is None or df.empty:
        return df

    term_col, coef_col, se_col, p_col = df.columns[:4]
    ok_css = "background-color: rgba(46, 204, 113, 0.10); color: #1f5131; font-weight: 600;"
    ng_css = "background-color: rgba(231, 76, 60, 0.10); color: #7a1e18; font-weight: 600;"

    def _style_table(frame):
        styles = pd.DataFrame("", index=frame.index, columns=frame.columns)
        for idx, row in frame.iterrows():
            verdict = None
            if idx == 0:
                p_value = pd.to_numeric(pd.Series([row[p_col]]), errors="coerce").iloc[0]
                if not pd.isna(p_value):
                    verdict = p_value > 0.05
            elif idx == 1:
                coef_value = pd.to_numeric(pd.Series([row[coef_col]]), errors="coerce").iloc[0]
                if not pd.isna(coef_value):
                    verdict = abs(coef_value - 1.0) <= 0.05

            if verdict is True:
                styles.loc[idx, :] = ok_css
            elif verdict is False:
                styles.loc[idx, :] = ng_css
        return styles

    return (
        df.style
        .format(
            {
                coef_col: lambda value: _format_display_value(value, digits=5),
                se_col: lambda value: _format_display_value(value, digits=5),
                p_col: lambda value: _format_display_value(value, digits=3),
            }
        )
        .apply(_style_table, axis=None)
        .hide(axis="index")
    )


def _style_study4_aiag_bias(df):
    if df is None or df.empty:
        return df

    ref_col, bias_col, p_col = df.columns[:3]
    ok_css = "background-color: rgba(46, 204, 113, 0.10); color: #1f5131; font-weight: 600;"
    ng_css = "background-color: rgba(231, 76, 60, 0.10); color: #7a1e18; font-weight: 600;"

    def _style_table(frame):
        styles = pd.DataFrame("", index=frame.index, columns=frame.columns)
        for idx, row in frame.iterrows():
            p_value = pd.to_numeric(pd.Series([row[p_col]]), errors="coerce").iloc[0]
            if pd.isna(p_value):
                continue
            styles.loc[idx, :] = ok_css if p_value > 0.05 else ng_css
        return styles

    return (
        df.style
        .format(
            {
                ref_col: lambda value: str(value) if isinstance(value, str) else _format_display_value(value, digits=0),
                bias_col: lambda value: _format_display_value(value, digits=6),
                p_col: lambda value: _format_display_value(value, digits=3),
            }
        )
        .apply(_style_table, axis=None)
        .hide(axis="index")
    )


def _drop_empty_editor_rows(df):
    if df is None:
        return pd.DataFrame()
    return df.dropna(axis=0, how="all").copy()


def _render_consultant_badge():
    st.sidebar.markdown(
        """
        <div class="consultant-badge">
            <div class="consultant-badge-role">❯ 품질관리 기술사</div>
            <div class="consultant-badge-name">김훈희</div>
            <div class="consultant-badge-caption">MSA AI Studio | Measurement System Analysis</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _study6_acceptance_reference():
    return pd.DataFrame(
        [
            {"통계량": "적합",      "유효성 E": "0.9 ~ 1",   "허위 경보율 P(FA)": "0 ~ 0.05",    "누락률 P(miss)": "0 ~ 0.02"},
            {"통계량": "조건부 채택", "유효성 E": "0.8 ~ 0.9", "허위 경보율 P(FA)": "0.05 ~ 0.10", "누락률 P(miss)": "0.02 ~ 0.05"},
            {"통계량": "부적합",     "유효성 E": "0.8 이하",   "허위 경보율 P(FA)": "0.10 이상",   "누락률 P(miss)": "0.05 이상"},
        ]
    )


def _study6_kappa_reference():
    return pd.DataFrame(
        [
            {"평가지침": "우수", "Cohen's κ (vs 기준)": "≥ 0.75 (AIAG) / ≥ 0.90 (미니탭)", "Fleiss' κ (평가자 간)": "≥ 0.75 (AIAG) / ≥ 0.90 (미니탭)"},
            {"평가지침": "보통", "Cohen's κ (vs 기준)": "0.40 ~ 0.75 (AIAG) / 0.70 ~ 0.90 (미니탭)", "Fleiss' κ (평가자 간)": "0.40 ~ 0.75 (AIAG) / 0.70 ~ 0.90 (미니탭)"},
            {"평가지침": "미흡", "Cohen's κ (vs 기준)": "< 0.40 (AIAG) / < 0.70 (미니탭)", "Fleiss' κ (평가자 간)": "< 0.40 (AIAG) / < 0.70 (미니탭)"},
        ]
    )


def _build_study23_type1_context(re_val, ref_val, usl, lsl, u_cal, u_lin, u_msrest, raw_input, bias_dist, re_dist):
    values = _parse_measurement_text(raw_input)
    if len(values) >= 1:
        mean_repeat = float(np.mean(values))
        sigma_repeat = float(np.std(values, ddof=1)) if len(values) >= 2 else 0.0
        abias = mean_repeat - ref_val
        u_bi = calc_u_distribution(abs(abias) * 2, bias_dist)
        u_evr = sigma_repeat
    else:
        mean_repeat = ref_val
        sigma_repeat = 0.0
        abias = 0.0
        u_bi = 0.0
        u_evr = 0.0

    u_re = calc_u_distribution(re_val, re_dist)
    u_ev = max(u_evr, u_re)
    u_ms = float(np.sqrt(u_cal**2 + u_lin**2 + u_bi**2 + u_ev**2 + u_msrest**2))
    u_ms_expanded = 2 * u_ms
    tol = usl - lsl
    return {
        "values": values,
        "mean_repeat": mean_repeat,
        "sigma_repeat": sigma_repeat,
        "abias": abias,
        "u_CAL": float(u_cal),
        "u_BI": float(u_bi),
        "u_EVR": float(u_evr),
        "u_RE": float(u_re),
        "u_EV": float(u_ev),
        "u_MS": u_ms,
        "U_MS": u_ms_expanded,
        "TOL": tol,
    }


def _format_study23_table(df, float_digits=DISPLAY_DECIMALS):
    if df is None or df.empty:
        return "(데이터 없음)"

    display = df.copy().reset_index(drop=False)
    if "index" in display.columns:
        if "항목" in display.columns:
            display = display.drop(columns=["index"])
        else:
            display = display.rename(columns={"index": "항목"})

    def fmt(value):
        if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
            return f"{float(value):.{float_digits}f}"
        if pd.isna(value):
            return ""
        return str(value)

    as_text = display.apply(lambda col: col.map(fmt))
    widths = {col: max(len(str(col)), as_text[col].astype(str).map(len).max()) for col in as_text.columns}
    header = "  ".join(str(col).ljust(widths[col]) for col in as_text.columns)
    lines = [header]
    for _, row in as_text.iterrows():
        lines.append("  ".join(str(row[col]).ljust(widths[col]) for col in as_text.columns))
    return "\n".join(lines)


def _build_study23_report(type1_result, type2_result, run_t1, run_t2, is_expert):
    sections = []
    if run_t1 and type1_result is not None:
        df_summary = type1_result["summary"].copy()
        preview = df_summary[["항목", "값", "기준", "판정"]]
        if not is_expert:
            preview = preview[~preview["항목"].astype(str).str.contains(r"u_|Abias|불확도|편의유래", regex=True, na=False)]
        sections.append("=== Type 1 Gauge Capability ===")
        sections.append(_format_study23_table(preview))

    if run_t2 and type2_result is not None:
        sections.append("")
        if type2_result.get("interaction_significant"):
            sections.append("=== 교호작용 포함 이원분산분석표 Gage R&R Study2 ===")
        elif len(type2_result["anova"].index) == 3:
            sections.append("=== (측정자=1, 자동측정) Gage R&R Study3 ===")
        else:
            sections.append("=== 교호작용 없는 이원분산분석표 Gage R&R Study2 ===")
        sections.append(_format_study23_table(type2_result["anova"].reset_index()))
        sections.append("")
        sections.append("=== Gage R&R 분석표 ===")
        sections.append(_format_study23_table(type2_result["gage"]))
        sections.append("")
        sections.append("=== 판정 결과 ===")
        sections.append(_format_study23_table(_study23_visible_judgement(type2_result, is_expert)))
        if is_expert and not pd.isna(type2_result.get("c_mp", np.nan)):
            c_mp_status = "적합" if float(type2_result["c_mp"]) >= 1.33 else "부적합"
            sections.append("")
            sections.append("=== 추가 통계량 ===")
            sections.append(f"C_MP  {_format_display_value(type2_result['c_mp'])}  >= 1.33  {c_mp_status}")
        if is_expert:
            sections.append("")
            sections.append("[불확도 구성 (0 값 제외)]")
            non_zero = {k: v for k, v in type2_result["u_elements"].items() if abs(v) > 1e-12}
            for key, value in sorted(non_zero.items(), key=lambda item: item[1], reverse=True):
                sections.append(f"{key:<10} {value:.3f}")
            sections.append(f"합성 표준불확도 u = {type2_result['u_mp']:.3f}")
            sections.append(f"확장 불확도(95%) U = {type2_result['u_mp_expanded']:.3f}")

    return "\n".join(sections).strip()


def _build_uncertainty_chart_data(type2_result, type1_context):
    if type2_result is not None:
        label_map = {
            "u_RE": "분해능/계기 해상도",
            "u_BI": "치우침(Bias) 불확도",
            "u_AV": "측정자 간 차이(재현성)",
            "u_CAL": "교정(표준불확도)",
            "u_EVR": "반복성(실험 표준편차)",
            "u_EVO": "반복성(ANOVA 잔차 기반)",
            "u_IA": "부품×측정자 교호작용",
            "u_LIN": "선형성",
            "u_MSREST": "기타 측정시스템 잔여",
            "u_T": "환경(온도 등)",
            "u_STAB": "장기 안정성",
            "u_OBJ": "부품/측정물 영향",
            "u_GV": "게이지 자체 변동",
            "u_REST": "기타 잔여",
        }
        records = [
            {"코드": key, "값": value, "설명": label_map.get(key, key)}
            for key, value in type2_result["u_elements"].items()
            if abs(value) > 1e-12
        ]
    elif type1_context is not None:
        records = [
            {"코드": "u_RE", "값": type1_context["u_RE"], "설명": "분해능/계기 해상도"},
            {"코드": "u_BI", "값": type1_context["u_BI"], "설명": "치우침(Bias) 불확도"},
            {"코드": "u_CAL", "값": type1_context["u_CAL"], "설명": "교정(표준불확도)"},
            {"코드": "u_EVR", "값": type1_context["u_EVR"], "설명": "반복성(실험 표준편차)"},
            {"코드": "u_EV", "값": type1_context["u_EV"], "설명": "측정장치 불확도"},
        ]
    else:
        records = []
    chart_df = pd.DataFrame(records)
    if not chart_df.empty:
        chart_df = chart_df.sort_values("값", ascending=True)
    return chart_df


def _build_study23_payload(tol, report_text, type1_result, type2_result):
    payload = {"tol": tol, "report_text": report_text}
    if type1_result is not None:
        payload["type1_summary"] = type1_result["summary"].to_dict(orient="records")
    if type2_result is not None:
        payload["type2"] = {
            "anova": type2_result["anova"].reset_index().to_dict(orient="records"),
            "gage": type2_result["gage"].to_dict(orient="records"),
            "judgement": type2_result["judgement"].to_dict(orient="records"),
            "u_elements": type2_result["u_elements"],
            "ndc": type2_result["ndc"],
            "q_mp": type2_result["q_mp"],
            "c_mp": type2_result["c_mp"],
            "overall_pass": type2_result["overall_pass"],
            "overall_status": type2_result.get("overall_status"),
            "interaction_significant": type2_result["interaction_significant"],
            "warnings": type2_result["warnings"],
        }
    return payload


def _get_study23_reference_sample():
    type1_values = [
        6.001, 6.002, 6.001, 6.001, 6.002, 6.001, 6.001, 6.000, 5.999, 6.001,
        6.001, 6.000, 6.001, 6.002, 6.002, 6.002, 6.002, 6.002, 6.000, 6.002,
        6.000, 5.999, 6.002, 6.002, 6.001, 6.001, 6.000, 5.999, 5.999, 6.000,
        6.001, 6.001, 6.002, 6.001, 6.001, 6.000, 5.999, 5.999, 6.000, 6.001,
        6.002, 6.001, 6.002, 6.002, 6.001, 6.002, 6.001, 6.001,
    ]
    gage_values = [
        6.029, 6.019, 6.004, 5.982, 6.009, 5.971, 5.995, 6.014, 5.985, 6.024,
        6.030, 6.020, 6.003, 5.982, 6.009, 5.972, 5.997, 6.018, 5.987, 6.028,
        6.033, 6.020, 6.007, 5.985, 6.014, 5.973, 5.997, 6.019, 5.987, 6.029,
        6.032, 6.019, 6.007, 5.986, 6.014, 5.972, 5.996, 6.015, 5.986, 6.025,
        6.031, 6.020, 6.010, 5.984, 6.015, 5.975, 5.995, 6.016, 5.987, 6.026,
        6.030, 6.020, 6.006, 5.984, 6.014, 5.974, 5.994, 6.015, 5.986, 6.025,
    ]
    parts = list(np.tile(np.arange(1, 11), 6))
    operators = list(np.repeat([1, 2, 3], 20))
    return {
        "RE": 0.001,
        "RefV": 6.002,
        "USL": 6.03,
        "LSL": 5.97,
        "u_CAL": 0.001,
        "type1_text": "\n".join(f"{value:.3f}" for value in type1_values),
        "type2_df": pd.DataFrame({
            "부품": parts,
            "측정자": operators,
            "측정값": gage_values,
        }),
    }


def _load_study23_reference_sample():
    sample_data = _get_study23_reference_sample()
    st.session_state["s23_re"] = float(sample_data["RE"])
    st.session_state["s23_refv"] = float(sample_data["RefV"])
    st.session_state["s23_usl"] = float(sample_data["USL"])
    st.session_state["s23_lsl"] = float(sample_data["LSL"])
    st.session_state["s23_ucal"] = float(sample_data["u_CAL"])
    st.session_state["s23_type1_values"] = sample_data["type1_text"]
    st.session_state["s23_sample_type2_df"] = sample_data["type2_df"].copy()
    st.session_state["s23_use_sample_type2"] = True


def _use_uploaded_study23_file():
    st.session_state["s23_use_sample_type2"] = False


def _is_iso_mode(level):
    return "ISO 22514-7" in str(level)


# ─── 에러 메시지 한국어화 매핑 ────────────────────────────
_ERROR_KO_MAP = [
    ("No columns to parse from file",         "파일에서 열을 읽을 수 없습니다. 파일 형식(CSV/Excel)과 인코딩을 확인해 주세요."),
    ("does not exist",                         "파일 또는 경로가 존재하지 않습니다."),
    ("Permission denied",                      "파일 접근 권한이 없습니다."),
    ("UnicodeDecodeError",                     "파일 인코딩 오류입니다. UTF-8 또는 CP949(한글 Windows)로 저장된 파일을 사용해 주세요."),
    ("not enough values to unpack",            "데이터 구조가 맞지 않습니다. 열 수를 확인해 주세요."),
    ("could not convert string to float",      "숫자로 변환할 수 없는 값이 포함되어 있습니다. 측정값 열에 텍스트가 있는지 확인해 주세요."),
    ("index out of range",                     "데이터 범위를 벗어났습니다. 입력 데이터 크기를 확인해 주세요."),
    ("singular matrix",                        "행렬이 특이(singular)합니다. 측정값에 변동이 없거나 데이터가 부족합니다."),
    ("division by zero",                       "0으로 나누기 오류가 발생했습니다. 공차(TOL) 또는 측정 변동이 0인지 확인해 주세요."),
    ("NaN",                                    "계산 중 유효하지 않은 값(NaN)이 발생했습니다. 입력 데이터를 확인해 주세요."),
    ("Expected 2D array",                      "2차원 배열이 필요합니다. 데이터 구조를 확인해 주세요."),
    ("operands could not be broadcast",        "배열 크기가 맞지 않습니다. 데이터 열 수를 확인해 주세요."),
    ("Workbook contains no default named styles", "Excel 파일 형식 오류입니다. 다른 Excel 파일로 시도해 주세요."),
    ("File is not a zip file",                 "파일이 손상되었거나 올바른 Excel 파일이 아닙니다."),
    ("sheet",                                  None),  # 원문 유지
]


def _localize_error(exc: Exception) -> str:
    """Python 예외 메시지를 사용자 친화적 한국어로 변환"""
    msg = str(exc)
    for pattern, ko in _ERROR_KO_MAP:
        if pattern.lower() in msg.lower():
            if ko:
                return ko
            break
    # 이미 한글이 포함된 메시지는 그대로 반환
    return msg


def _get_study4_reference_sample():
    refs = [
        6.19, 6.19, 6.19, 6.19,
        9.17, 9.17, 9.17, 9.17,
        1.99, 1.99, 1.99, 1.99,
        7.77, 7.77, 7.77, 7.77,
        4.00, 4.00, 4.00, 4.00,
        10.77, 10.77, 10.77, 10.77,
        4.78, 4.78, 4.78, 4.78,
        2.99, 2.99, 2.99, 2.99,
        6.98, 6.98, 6.98, 6.98,
        9.98, 9.98, 9.98, 9.98,
    ]
    values = [
        6.31, 6.27, 6.31, 6.28,
        9.27, 9.21, 9.34, 9.23,
        2.21, 2.19, 2.22, 2.20,
        8.00, 7.81, 7.95, 7.84,
        4.27, 4.15, 4.15, 4.15,
        10.93, 10.73, 10.92, 10.89,
        4.95, 4.87, 5.00, 5.00,
        3.24, 3.17, 3.21, 3.21,
        7.14, 7.07, 7.18, 7.20,
        10.23, 10.02, 10.07, 10.17,
    ]
    grr_values = [
        8.120, 7.445, 9.965, 6.140, 5.690, 2.855, 10.685, 6.725, 4.970,
        9.875, 8.435, 6.815, 10.010, 5.960, 5.600, 2.450, 10.595, 6.275, 5.105,
        10.100, 8.480, 7.490, 9.560, 6.365, 5.780, 2.585, 10.775, 6.545, 5.510, 9.875,
        8.200, 7.300, 9.660, 6.095, 5.080, 2.315, 10.450, 6.240, 5.015, 10.080,
        8.290, 7.120, 9.340, 6.185, 5.340, 2.585, 10.840, 6.120, 5.285, 9.800,
        8.245, 7.075, 9.250, 6.185, 5.440, 2.315, 11.050, 6.300, 5.150, 9.970,
        8.525, 7.535, 9.830, 6.140, 5.780, 2.630, 10.865, 6.590, 5.060, 10.190,
        8.435, 7.355, 9.695, 6.140, 5.735, 2.360, 11.000, 6.500, 5.195, 9.785,
        8.345, 7.085, 9.515, 6.050, 5.555, 2.585, 11.180, 6.725, 5.105, 9.965,
    ]
    grr_parts = list(np.tile(np.arange(1, 11), 9))
    grr_ops = list(np.repeat([1, 2, 3], 30))
    return {
        "RE": 0.005,
        "u_CAL": 0.005,
        "USL": 11.0,
        "LSL": 2.0,
        "df": pd.DataFrame({"Reference": refs, "Value": values}),
        "grr_df": pd.DataFrame({"Part": grr_parts, "Operator": grr_ops, "Value": grr_values}),
    }


def _load_study4_reference_sample():
    sample = _get_study4_reference_sample()
    st.session_state["s4_re"] = float(sample["RE"])
    st.session_state["s4_ucal"] = float(sample["u_CAL"])
    st.session_state["s4_usl"] = float(sample["USL"])
    st.session_state["s4_lsl"] = float(sample["LSL"])
    st.session_state["s4_df"] = sample["df"].copy()
    st.session_state["s4_grr_df"] = sample["grr_df"].copy()


def _get_study1_reference_sample():
    values = np.array(
        [
            6.001, 6.002, 6.001, 6.001, 6.002, 6.001, 6.001, 6.000, 5.999, 6.001,
            6.001, 6.000, 6.001, 6.002, 6.002, 6.002, 6.002, 6.002, 6.002, 6.000,
            6.002, 6.000, 5.999, 6.002, 6.002, 6.001, 6.001, 6.000, 5.999, 5.999,
            6.000, 6.001, 6.001, 6.002, 6.001, 6.001, 6.000, 6.000, 5.999, 5.999,
            6.000, 6.001, 6.002, 6.001, 6.002, 6.002, 6.001, 6.002, 6.001, 6.001,
        ],
        dtype=float,
    )
    return {
        "RE": 0.001,
        "RefV": 6.002,
        "USL": 6.03,
        "LSL": 5.97,
        "u_CAL": 0.001,
        "u_LIN": 0.0,
        "u_REST": 0.0,
        "values": values,
        "text": ", ".join(f"{value:.3f}" for value in values),
    }


def _load_study1_reference_sample():
    sample = _get_study1_reference_sample()
    st.session_state["s1_re"] = float(sample["RE"])
    st.session_state["s1_refv"] = float(sample["RefV"])
    st.session_state["s1_usl"] = float(sample["USL"])
    st.session_state["s1_lsl"] = float(sample["LSL"])
    st.session_state["s1_ucal"] = float(sample["u_CAL"])
    st.session_state["s1_ulin"] = float(sample["u_LIN"])
    st.session_state["s1_urest"] = float(sample["u_REST"])
    st.session_state["s1_data"] = sample["text"]

    df_summary, _, _ = run_study1_analysis(
        sample["RE"],
        sample["RefV"],
        sample["USL"],
        sample["LSL"],
        sample["u_CAL"],
        sample["u_LIN"],
        sample["u_REST"],
        sample["values"],
    )
    st.session_state["s1_result"] = df_summary
    st.session_state["s1_raw_X"] = sample["values"].copy()
    st.session_state["s1_ref"] = float(sample["RefV"])
    st.session_state["s1_tol"] = float(sample["USL"] - sample["LSL"])
    st.session_state["s1_result_filter"] = "all"


def _resolve_study6_sample_path():
    downloads_dir = Path.home() / "Downloads"
    candidates = []
    if downloads_dir.exists():
        exact_path = downloads_dir / "계수형.xlsx"
        if exact_path.exists():
            candidates.append(exact_path)
        candidates.extend(sorted(downloads_dir.glob("*계수형*.xlsx")))
    local_copy = Path(__file__).resolve().parent / "study6_sample.xlsx"
    if local_copy.exists():
        candidates.append(local_copy)

    for path in candidates:
        if path.exists():
            return path
    raise FileNotFoundError("Study-6 샘플 파일 `계수형.xlsx`를 찾지 못했습니다.")


def _build_study6_builtin_sample() -> pd.DataFrame:
    """엑셀 파일 없이 내장 샘플 데이터를 생성합니다 (AIAG 예제 기반)."""
    rng = np.random.default_rng(42)
    n = 50
    standard = rng.integers(0, 2, size=n).tolist()
    def _noisy(ref, err_rate=0.08):
        out = ref[:]
        for i in range(len(out)):
            if rng.random() < err_rate:
                out[i] = 1 - out[i]
        return out
    return pd.DataFrame({
        "Sample":    list(range(1, n + 1)),
        "Standard":  standard,
        "Appraiser A": _noisy(standard, 0.06),
        "Appraiser B": _noisy(standard, 0.10),
        "Appraiser C": _noisy(standard, 0.08),
    })


def _load_study6_reference_sample():
    try:
        sample_path = _resolve_study6_sample_path()
        sample_df = pd.read_excel(sample_path)
    except FileNotFoundError:
        sample_df = _build_study6_builtin_sample()
    st.session_state["s6_df"] = normalize_study6_df(sample_df)


def _cleanup_temp_files():
    """앱 시작 시 이전 실행에서 생성된 임시 Excel 파일을 자동 삭제"""
    import glob
    tmp_patterns = [
        Path(__file__).parent / "tmp_*.xlsx",
        Path(__file__).parent / "tmp_*.xls",
    ]
    for pattern in tmp_patterns:
        for f in glob.glob(str(pattern)):
            try:
                os.remove(f)
            except OSError:
                pass


def main():
    _cleanup_temp_files()
    st.sidebar.title("🔬 MSA AI Studio")
    st.sidebar.subheader("Full Suite (AIAG / ISO 22514-7)")
    user_level = st.sidebar.radio("분석 기준", ["AIAG", "ISO 22514-7"])
    st.sidebar.text_input(
        "측정기 번호",
        key="msa_gage_no",
        help="보고서와 결과 이력에 표시할 측정기 식별 번호입니다. 계산에는 직접 사용되지 않습니다.",
    )
    st.sidebar.divider()

    # ── AI 설정 ──────────────────────────────────────────────────────
    _saved_cfg = _load_ai_config()
    _ai_provider_options = {
        "Groq (무료)": "groq",
        "Cerebras (무료)": "cerebras",
        "OpenRouter (무료모델)": "openrouter",
        "Google Gemini (무료)": "gemini",
        "OpenAI": "openai",
        "Claude (Anthropic)": "claude",
        "없음 (기본 엔진)": "",
    }
    _provider_to_label = {v: k for k, v in _ai_provider_options.items()}
    _saved_provider = _saved_cfg.get("provider", "groq")
    _saved_provider_label = _provider_to_label.get(_saved_provider, "Groq (무료)")
    _saved_provider_index = list(_ai_provider_options.keys()).index(_saved_provider_label) if _saved_provider_label in _ai_provider_options else 0

    with st.sidebar.expander("AI 설정", expanded=False):
        _ai_provider_label = st.selectbox(
            "AI 제공사",
            list(_ai_provider_options.keys()),
            index=_saved_provider_index,
            key="ai_provider_select",
        )
        _ai_provider_val = _ai_provider_options[_ai_provider_label]

        _saved_provider_keys = dict(_saved_cfg.get("provider_api_keys", {}))
        _saved_key = str(_saved_provider_keys.get(_ai_provider_val, "")).strip()

        _ai_key_val = ""
        if _ai_provider_val:
            _key_col, _del_col = st.columns([3, 1], gap="small")
            _api_input_widget_key = f"ai_api_key_input_{_ai_provider_val or 'none'}"
            _new_key_input = _key_col.text_input(
                "API Key",
                type="password",
                key=_api_input_widget_key,
                placeholder="저장됨 (변경 시만 입력)" if _saved_key else "API Key 입력",
            )
            if _saved_key and _del_col.button("초기화", key="ai_key_clear_btn"):
                _save_ai_config(
                    _build_ai_provider_config(
                        _saved_cfg,
                        _ai_provider_val,
                        "",
                        _saved_cfg.get("model", ""),
                        clear_key=True,
                    )
                )
                st.rerun()
            if _new_key_input.strip():
                # 새 키 입력 시 즉시 자동 저장
                _ai_key_val = _new_key_input.strip()
                _save_ai_config(
                    _build_ai_provider_config(
                        _saved_cfg,
                        _ai_provider_val,
                        _ai_key_val,
                        _saved_cfg.get("model", ""),
                    )
                )
            else:
                _ai_key_val = _saved_key

        if _ai_provider_val == "groq":
            _model_options = ["llama-3.3-70b-versatile", "llama-3.1-8b-instant", "mixtral-8x7b-32768", "gemma2-9b-it"]
        elif _ai_provider_val == "cerebras":
            _model_options = ["llama-3.3-70b", "llama3.1-8b"]
        elif _ai_provider_val == "openrouter":
            _model_options = [
                "meta-llama/llama-4-scout:free",
                "meta-llama/llama-4-maverick:free",
                "google/gemma-3-12b-it:free",
                "mistralai/mistral-7b-instruct:free",
                "microsoft/phi-4-reasoning:free",
            ]
        elif _ai_provider_val == "gemini":
            _model_options = ["gemini-2.0-flash"]
        elif _ai_provider_val == "openai":
            _model_options = ["gpt-5", "gpt-4o", "gpt-4o-mini"]
        elif _ai_provider_val == "claude":
            _model_options = ["claude-haiku-4-5-20251001", "claude-sonnet-4-6", "claude-opus-4-6"]
        else:
            _model_options = []

        _saved_model = _saved_cfg.get("model", "")
        _model_index = _model_options.index(_saved_model) if _saved_model in _model_options else 0
        if _model_options:
            _ai_model_val = st.selectbox("모델", _model_options, index=_model_index, key="ai_model_select")
            # 모델 변경 시 자동 저장
            if _ai_key_val and _ai_model_val != _saved_model:
                _save_ai_config(
                    _build_ai_provider_config(
                        _saved_cfg,
                        _ai_provider_val,
                        _ai_key_val,
                        _ai_model_val,
                    )
                )
        else:
            _ai_model_val = ""

        # ai_settings 갱신
        _ensure_ai_state()
        st.session_state["ai_settings"]["enabled"] = bool(_ai_provider_val and _ai_key_val)
        st.session_state["ai_settings"]["provider"] = _ai_provider_val
        st.session_state["ai_settings"]["api_key"] = _ai_key_val
        st.session_state["ai_settings"]["model"] = _ai_model_val
        if _ai_provider_val and _ai_key_val:
            st.sidebar.caption("AI 연결됨")
        elif _ai_provider_val and not _ai_key_val:
            st.sidebar.caption("API Key 없음")

    st.sidebar.divider()
    menu = [
        "🏠 Study Guide",
        "🔍 Study-1 (Bias)",
        "📊 Study-2/3 (Precision)",
        "📈 Study-4 (Linearity)",
        "📉 Study-5 (Stability)",
        "✅ Study-6 (Attribute)",
        "📋 통합 보고서 (Study 1~6)",
    ]
    choice = st.sidebar.selectbox("분석 선택", menu)
    _render_consultant_badge()

    if choice == "🏠 Study Guide":       render_home(user_level)
    elif "Study-1" in choice:            render_study1(user_level)
    elif "Study-2/3" in choice:          render_study2_3(user_level)
    elif "Study-4" in choice:            render_study4(user_level)
    elif "Study-5" in choice:            render_study5(user_level)
    elif "Study-6" in choice:            render_study6(user_level)
    elif "통합 보고서" in choice:         render_combined_report(user_level)


def render_home(level):
    st.title("MSA AI Studio 🚀")
    st.info(f"💡 현재 **{level}** 기준으로 설정되어 있습니다.")

    tab_select, tab_detail, tab_compare, tab_data, tab_terms = st.tabs(
        ["📌 Study 선택", "📖 Study별 안내", "⚖️ AIAG vs ISO", "📁 데이터 준비", "📚 용어 사전"]
    )

    # ── 탭 1: Study 선택 플로우 ──────────────────────────────────
    with tab_select:
        st.markdown("### MSA Study 선택 가이드")
        st.caption("아래 흐름도를 따라 현재 상황에 맞는 Study를 선택하세요.")
        components.html("""
<style>
.study-guide-shell {
    display: grid;
    grid-template-columns: minmax(0, 1.5fr) minmax(300px, 0.92fr);
    gap: 18px;
    align-items: start;
    margin: 0.35rem 0 1.25rem;
}
.study-guide-board,
.study-guide-panel {
    border: 1px solid #d9e5f3;
    border-radius: 22px;
    background:
        radial-gradient(circle at top right, rgba(78, 145, 210, 0.10), transparent 32%),
        linear-gradient(135deg, #f8fbff 0%, #ffffff 48%, #fbfdff 100%);
    box-shadow: 0 16px 34px rgba(34, 76, 128, 0.08);
}
.study-guide-board {
    padding: 22px 22px 18px;
}
.study-guide-panel {
    padding: 18px 18px 16px;
}
.study-guide-header {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    gap: 14px;
    margin-bottom: 16px;
}
.study-guide-kicker {
    display: inline-flex;
    align-items: center;
    gap: 0.4rem;
    padding: 0.3rem 0.7rem;
    border-radius: 999px;
    background: #edf4fd;
    color: #315f96;
    font-size: 0.74rem;
    font-weight: 700;
    letter-spacing: 0.04em;
    text-transform: uppercase;
}
.study-guide-lead {
    margin-top: 0.7rem;
    color: #27466f;
    font-size: 1.02rem;
    font-weight: 700;
}
.study-guide-note {
    min-width: 190px;
    padding: 0.8rem 0.9rem;
    border-radius: 16px;
    background: rgba(255, 255, 255, 0.82);
    border: 1px solid #d8e2f0;
    color: #637891;
    font-size: 0.82rem;
    line-height: 1.5;
}
.study-guide-lane {
    display: grid;
    grid-template-columns: minmax(290px, 1fr) 78px minmax(210px, 0.82fr);
    gap: 14px;
    align-items: center;
}
.study-guide-lane + .study-guide-no {
    margin-top: 2px;
}
.guide-question,
.guide-study-card {
    min-height: 90px;
    border-radius: 18px;
    padding: 15px 18px;
    display: flex;
    flex-direction: column;
    justify-content: center;
    position: relative;
}
.guide-question {
    background: linear-gradient(180deg, #f6faff 0%, #eef4fd 100%);
    border: 2px solid #4e91d2;
    color: #174f8a;
    box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.85);
}
.guide-study-card {
    background: rgba(255, 255, 255, 0.92);
    box-shadow: 0 10px 22px rgba(33, 63, 101, 0.08);
}
.guide-study-card strong {
    font-size: 1.04rem;
    line-height: 1.25;
}
.guide-study-card small {
    display: block;
    margin-top: 0.45rem;
    font-size: 0.8rem;
    opacity: 0.95;
}
.guide-study-card.study-s1 { border: 2px solid #f5a623; color: #7a4f00; background: #fff5e7; }
.guide-study-card.study-s2 { border: 2px solid #43a047; color: #1b5e20; background: #edf8ef; }
.guide-study-card.study-s4 { border: 2px solid #8e24aa; color: #4a148c; background: #f8ecfb; }
.guide-study-card.study-s5 { border: 2px solid #e91e63; color: #880e4f; background: #fff0f5; }
.guide-study-card.study-s6 { border: 2px solid #00acc1; color: #006064; background: #ecfbfd; }
.guide-connector {
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 0.28rem;
}
.guide-arrow {
    color: #5b738f;
    font-size: 1.35rem;
    line-height: 1;
}
.guide-yes {
    color: #1f9d55;
    font-size: 0.76rem;
    font-weight: 800;
    letter-spacing: 0.06em;
}
.guide-no {
    margin: 2px 0 8px 12px;
    color: #8a97a8;
    font-size: 0.8rem;
    font-weight: 600;
}
.study-guide-side {
    display: flex;
    flex-direction: column;
    gap: 16px;
}
.guide-panel-title {
    color: #203a5c;
    font-size: 1rem;
    font-weight: 800;
    margin-bottom: 0.9rem;
}
.guide-chip-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 10px;
}
.guide-chip {
    border-radius: 16px;
    padding: 12px 12px 11px;
    background: rgba(255, 255, 255, 0.9);
    border: 1px solid #dce5f1;
    border-left-width: 4px;
}
.guide-chip strong {
    display: block;
    color: #17365c;
    font-size: 0.9rem;
    margin-bottom: 0.2rem;
}
.guide-chip span {
    color: #5f738d;
    font-size: 0.76rem;
    line-height: 1.45;
}
.guide-chip.study-s1 { border-left-color: #f5a623; }
.guide-chip.study-s2 { border-left-color: #43a047; }
.guide-chip.study-s4 { border-left-color: #8e24aa; }
.guide-chip.study-s5 { border-left-color: #e91e63; }
.guide-chip.study-s6 { border-left-color: #00acc1; }
.guide-tip-list {
    display: grid;
    gap: 10px;
}
.guide-tip {
    border-radius: 15px;
    padding: 11px 12px;
    background: rgba(255, 255, 255, 0.9);
    border: 1px solid #dce5f1;
}
.guide-tip b {
    display: block;
    color: #23456e;
    font-size: 0.84rem;
    margin-bottom: 0.2rem;
}
.guide-tip span {
    color: #667b96;
    font-size: 0.77rem;
    line-height: 1.45;
}
@media (max-width: 1180px) {
    .study-guide-shell {
        grid-template-columns: 1fr;
    }
}
@media (max-width: 880px) {
    .study-guide-lane {
        grid-template-columns: 1fr;
        gap: 10px;
    }
    .guide-connector {
        flex-direction: row;
        justify-content: flex-start;
        gap: 0.5rem;
        padding-left: 0.3rem;
    }
    .guide-no {
        margin-left: 0;
    }
    .guide-chip-grid {
        grid-template-columns: 1fr;
    }
    .study-guide-header {
        flex-direction: column;
    }
}
</style>
<div class="study-guide-shell">
<div class="study-guide-board">
<div class="study-guide-header">
<div>
<div class="study-guide-kicker">Decision Flow</div>
<div class="study-guide-lead">질문에 YES가 나오면 해당 Study를 선택합니다.</div>
</div>
<div class="study-guide-note">
위에서 아래로 읽으면 됩니다.<br>
현재 질문에 해당하지 않으면 <b>NO</b>로 내려가 다음 항목을 확인하세요.
</div>
</div>

<div class="study-guide-lane">
<div class="guide-question">신규 장비 도입 / 초기 능력 확인?</div>
<div class="guide-connector"><span class="guide-arrow">→</span><span class="guide-yes">YES</span></div>
<div class="guide-study-card study-s1"><strong>🔍 Study-1</strong><small>편의·반복성</small></div>
</div>
<div class="guide-no">↓ NO</div>

<div class="study-guide-lane">
<div class="guide-question">측정자·설비 간 변동 평가?</div>
<div class="guide-connector"><span class="guide-arrow">→</span><span class="guide-yes">YES</span></div>
<div class="guide-study-card study-s2"><strong>📊 Study-2/3</strong><small>Gage R&amp;R (정밀도)</small></div>
</div>
<div class="guide-no">↓ NO</div>

<div class="study-guide-lane">
<div class="guide-question">측정 범위 전체의 정확도 일관성 확인?</div>
<div class="guide-connector"><span class="guide-arrow">→</span><span class="guide-yes">YES</span></div>
<div class="guide-study-card study-s4"><strong>📈 Study-4</strong><small>선형성</small></div>
</div>
<div class="guide-no">↓ NO</div>

<div class="study-guide-lane">
<div class="guide-question">시간 경과에 따른 안정성 감시?</div>
<div class="guide-connector"><span class="guide-arrow">→</span><span class="guide-yes">YES</span></div>
<div class="guide-study-card study-s5"><strong>📉 Study-5</strong><small>안정성</small></div>
</div>
<div class="guide-no">↓ NO</div>

<div class="study-guide-lane">
<div class="guide-question">계수형(합/불) 판정 검사인가?</div>
<div class="guide-connector"><span class="guide-arrow">→</span><span class="guide-yes">YES</span></div>
<div class="guide-study-card study-s6"><strong>✅ Study-6</strong><small>Attribute 합치도</small></div>
</div>
</div>

<div class="study-guide-side">
<div class="study-guide-panel">
<div class="guide-panel-title">빠른 매칭</div>
<div class="guide-chip-grid">
<div class="guide-chip study-s1"><strong>Study-1</strong><span>신규 장비, 교정 직후, 기준값 대비 편의 확인</span></div>
<div class="guide-chip study-s2"><strong>Study-2/3</strong><span>사람·설비·반복 측정 변동을 함께 보고 싶을 때</span></div>
<div class="guide-chip study-s4"><strong>Study-4</strong><span>측정 범위가 넓고 구간별 bias 추세를 확인할 때</span></div>
<div class="guide-chip study-s5"><strong>Study-5</strong><span>시간 경과에 따른 drift와 안정성 감시가 필요할 때</span></div>
<div class="guide-chip study-s6"><strong>Study-6</strong><span>합/불 판정, 육안 검사, 관능 검사 일치도 평가</span></div>
</div>
</div>

<div class="study-guide-panel">
<div class="guide-panel-title">체크 포인트</div>
<div class="guide-tip-list">
<div class="guide-tip"><b>측정값이 연속형인가?</b><span>연속형이면 Study-1~5, 합/불 판정이면 Study-6부터 봅니다.</span></div>
<div class="guide-tip"><b>시간 축이 중요한가?</b><span>시간 변화가 핵심이면 Study-5가 먼저입니다.</span></div>
<div class="guide-tip"><b>범위 전체가 문제인가?</b><span>특정 기준값이 아니라 측정 범위 전반의 일관성이면 Study-4가 맞습니다.</span></div>
<div class="guide-tip"><b>측정자/설비 차이인가?</b><span>작업자, 설비, 반복 측정의 분산을 나누어 보려면 Study-2/3으로 갑니다.</span></div>
</div>
</div>
</div>
</div>
""", height=980, scrolling=False)

        st.markdown("---")
        st.markdown("### MSA 측정 오차 구성 및 대응방안")
        err_cols = st.columns(5)
        err_items = [
            (
                "🎯", "정확성 (Bias)",
                "측정값과 참값의 차이",
                ["측정기 교정(Calibration) 실시", "기준 시편 확인·교체", "영점(Zero) 조정", "측정 환경(온도·진동) 점검"],
                "#fff3e0", "#f5a623",
            ),
            (
                "🔁", "반복성 (Repeatability)",
                "동일 측정자 반복 측정 시 산포",
                ["측정기 수리 또는 교체", "지그(Fixture) 도입으로 자세 고정", "소모품(프로브·팁) 교체", "측정 방법 SOP 표준화"],
                "#e8f5e9", "#43a047",
            ),
            (
                "👥", "재현성 (Reproducibility)",
                "측정자 간·조건 간 차이",
                ["측정자 훈련 강화", "측정 방법·기준 문서화", "측정 보조 도구(지그) 도입", "측정 환경·조건 표준화"],
                "#e3f2fd", "#1e88e5",
            ),
            (
                "📐", "선형성 (Linearity)",
                "측정 범위에 따른 편의 변화",
                ["보정 곡선(Cal. Curve) 적용", "교정 주기 단축", "사용 범위 제한(축소)", "장비 교체 검토"],
                "#f3e5f5", "#8e24aa",
            ),
            (
                "⏱", "안정성 (Stability)",
                "시간 경과에 따른 일관성",
                ["정기 교정 주기 단축", "온도·습도 환경 관리", "마모 부품 예방 교체", "Drift 보정 절차 수립"],
                "#fce4ec", "#e91e63",
            ),
        ]
        for col, (icon, name, desc, actions, bg, border) in zip(err_cols, err_items):
            li_items = "".join(f"<li>{a}</li>" for a in actions)
            col.markdown(
                f"<div style='background:{bg};border:2px solid {border};border-radius:10px;"
                f"padding:12px 10px;font-size:0.82rem;'>"
                f"<div style='font-size:1.3rem;text-align:center;'>{icon}</div>"
                f"<b style='display:block;text-align:center;font-size:0.85rem;'>{name}</b>"
                f"<div style='color:#555;font-size:0.76rem;text-align:center;margin:4px 0 6px;'>{desc}</div>"
                f"<hr style='margin:4px 0;border:none;border-top:1px solid {border};opacity:0.4;'>"
                f"<div style='font-size:0.74rem;color:#333;'><b>문제 발생 시 대응방안</b>"
                f"<ul style='margin:3px 0 0 14px;padding:0;color:#555;'>{li_items}</ul></div>"
                f"</div>",
                unsafe_allow_html=True
            )

    # ── 탭 2: Study별 상세 안내 ──────────────────────────────────
    with tab_detail:
        studies = [
            {
                "icon": "🔍", "name": "Study-1", "sub": "편의(Bias) 및 반복성 (Repeatability)",
                "color": "#fff3e0", "border": "#f5a623",
                "purpose": "측정기가 참값에 얼마나 가깝게, 얼마나 일관되게 측정하는지 확인",
                "when": "신규 장비 도입, 교정 후 검증, 이상 징후 발생 시",
                "sample": "동일 기준 시편을 **최소 25회** 반복 측정 (권장: 50회, 참조값은 공차 중앙에 가까울수록 좋음)",
                "metrics": [
                    ("Bias / p-value", "%Bias는 참고값, 편의 판정은 **p > 0.05** 또는 CI에 0 포함 여부 중심"),
                    ("반복성 %EV", "반복성은 **≤ 10% 양호 / 10~30% 조건부 / >30% 개선**"),
                    ("Cg", "게이지 반복성 능력지수 — **≥ 1.33** 적합 (공차 대비 반복성)"),
                    ("Cgk", "편의 보정 게이지 능력지수 — **≥ 1.33** 적합 (반복성 + 편의 동시 고려)"),
                    ("Q_MS", "측정 시스템 불확도 능력지수 — **≤ 15%** (ISO 모드)"),
                    ("C_MS", "능력 지수 — **≥ 1.33** (ISO 모드)"),
                ],
                "tip": "Bias는 %값만으로 단정하지 말고 p-value/신뢰구간을 함께 보세요. 반복성(%EV)이 큰 상태에서는 bias 해석이 왜곡될 수 있습니다.",
            },
            {
                "icon": "📊", "name": "Study-2/3", "sub": "Gage R&R (정밀도)",
                "color": "#e8f5e9", "border": "#43a047",
                "purpose": "측정자 간·반복 측정 간 변동이 공차 대비 허용 수준인지 평가",
                "when": "양산 전 측정 시스템 승인, 측정자 교육 효과 검증, 정기 능력 평가",
                "sample": "**Study-2(교차형): 부품 10개 × 측정자 2~3명 × 2~3회 반복**\n**Study-3(자동/단일 측정자): 부품 10개 × 측정자 1명 × 2~3회 반복**",
                "metrics": [
                    ("%GRR (%SV 기준)", "총변동 대비 게이지 변동 — **≤ 10% 양호 / 10~30% 조건부 / >30% 개선**"),
                    ("%GRR (%공차 기준)", "공차 대비 게이지 변동 — **≤ 10% 양호 / 10~30% 조건부 / >30% 개선**"),
                    ("NDC", "구별 가능한 범주 수 — **≥ 5** 권장"),
                    ("Q_MP / C_MP", "측정 프로세스 능력 — Q_MP ≤ 30%, C_MP ≥ 1.33 (ISO 모드)"),
                ],
                "tip": "교호작용(부품×측정자) p-value ≤ 0.05이면 특정 측정자가 특정 부품을 다르게 측정한다는 의미입니다. 원인을 파악하세요.",
            },
            {
                "icon": "📈", "name": "Study-4", "sub": "선형성 (Linearity)",
                "color": "#f3e5f5", "border": "#8e24aa",
                "purpose": "측정 범위 전체에서 편의가 일정한지, 범위에 따라 달라지는지 확인",
                "when": "넓은 측정 범위를 가진 장비, 범위별 정확도 차이가 의심될 때",
                "sample": "**5개 이상 기준값 × 각 12회** 반복 (기준값은 측정 범위를 균등 분할)",
                "metrics": [
                    ("Slope p-value", "Bias 대 Reference 기울기 유의성 — **p > 0.05** 적합"),
                    ("Average bias p-value", "평균 치우침 유의성 — **p > 0.05** 적합"),
                    ("Q_MS / C_MS", "측정 시스템 능력 — Q_MS ≤ 15%, C_MS ≥ 1.33"),
                    ("Q_MP / C_MP", "측정 프로세스 능력 (GRR 데이터 추가 시) — Q_MP ≤ 30%"),
                ],
                "tip": "Bias 대 Reference 회귀에서 slope가 0에 가깝고 평균 bias도 0에 가까우면 양호합니다. 기울기가 유의하면 구간별 bias를 따로 봐야 합니다.",
            },
            {
                "icon": "📉", "name": "Study-5", "sub": "안정성 (Stability)",
                "color": "#fce4ec", "border": "#e91e63",
                "purpose": "시간이 지남에 따라 측정 시스템의 성능이 일정하게 유지되는지 감시",
                "when": "장기 사용 중인 장비 정기 점검, 환경 변화(온도·습도) 영향 평가",
                "sample": "**동일 기준 시편(master part) 1개**를 일정 주기별로 측정하여 **최소 20~25개 시점(subgroup)** 확보\n각 시점마다 **3~5회 반복 측정** (최소 2회 가능, 온도·습도 함께 기록 권장)",
                "metrics": [
                    ("X-bar 관리도", "평균의 이상 신호 탐지 — 관리 한계 이탈 없음"),
                    ("R 관리도", "산포의 이상 신호 탐지 — 관리 한계 이탈 없음"),
                    ("특이 신호", "현재 화면은 관리 한계 이탈 중심 확인, 추세/런 규칙은 보조 해석"),
                ],
                "tip": "관리도에서 이상 신호가 나타나면 즉시 측정기 점검 및 교정이 필요합니다. 최소 분기 1회 수행을 권장합니다.",
            },
            {
                "icon": "✅", "name": "Study-6", "sub": "계수형 합치도 (Attribute Agreement Analysis)",
                "color": "#e0f7fa", "border": "#00acc1",
                "purpose": "합/불 판정 검사자(또는 방법)가 기준과 얼마나 일치하는지 평가",
                "when": "육안 검사, 관능 검사, 기준 대비 합/불 판정 시스템 검증",
                "sample": "**최소 50 샘플 × 평가자 3명 이상 권장** (기준값 Standard 포함)\n현재 화면은 평가 결과 열 기준 비교 중심이며, 반복 trial 분석은 전용 입력 구조가 더 적합",
                "metrics": [
                    ("유효성 E", "기준과 일치한 비율 — **≥ 0.9** 적합 / 0.8~0.9 조건부"),
                    ("허위 경보율 P(FA)", "양품을 불량 판정한 비율 — **< 0.05** 적합"),
                    ("누락률 P(miss)", "불량을 양품 판정한 비율 — **< 0.02** 적합 (AIAG 4판)"),
                    ("Kappa", "우연 일치 보정 합치도 — **≥ 0.75** 우수 (AIAG 기준)"),
                ],
                "tip": "P(miss)(불량 누락)는 P(FA)(허위 경보)보다 엄격하게 관리해야 합니다. within-appraiser 반복합치도까지 보려면 trial 구조를 별도로 수집하는 편이 좋습니다.",
            },
        ]

        for s in studies:
            with st.expander(f"{s['icon']} **{s['name']}** — {s['sub']}", expanded=False):
                c1, c2 = st.columns([1, 1], gap="large")
                with c1:
                    st.markdown(f"**목적**\n\n{s['purpose']}")
                    st.markdown(f"**수행 시점**\n\n{s['when']}")
                    st.markdown(f"**필요 샘플**\n\n{s['sample']}")
                with c2:
                    st.markdown("**핵심 출력 지표 및 판정 기준**")
                    for metric, desc in s["metrics"]:
                        st.markdown(f"- `{metric}` : {desc}")
                st.markdown(
                    f"<div style='background:#fffde7;border-left:4px solid #f9a825;"
                    f"padding:8px 12px;border-radius:4px;font-size:0.85rem;margin-top:8px;'>"
                    f"💡 <b>실무 팁:</b> {s['tip']}</div>",
                    unsafe_allow_html=True
                )

    # ── 탭 3: AIAG vs ISO 비교 ──────────────────────────────────
    with tab_compare:
        st.markdown("### AIAG MSA 4판 vs ISO 22514-7:2021 비교")
        compare_df = pd.DataFrame([
            {"항목": "적용 관점", "AIAG MSA 4판": "자동차 Core Tools 기반 제조 현장 MSA", "ISO 22514-7": "산업 일반의 측정시스템/측정프로세스 능력 평가"},
            {"항목": "편의(Bias) 표현", "AIAG MSA 4판": "평균 bias, t/p-value (%Bias 병행 가능)", "ISO 22514-7": "bias 표준불확도 성분으로 반영"},
            {"항목": "반복성 표현", "AIAG MSA 4판": "EV (반복성, 6σ)", "ISO 22514-7": "반복성 표준불확도 성분"},
            {"항목": "재현성 표현", "AIAG MSA 4판": "AV (재현성, 6σ)", "ISO 22514-7": "작업자/환경 등 추가 표준불확도 성분"},
            {"항목": "종합 판정 지표", "AIAG MSA 4판": "%GRR, NDC", "ISO 22514-7": "Q_MS, C_MS, Q_MP, C_MP"},
            {"항목": "선형성 판정", "AIAG MSA 4판": "회귀 기반 선형성 평가", "ISO 22514-7": "u_LIN 반영 후 최종 지표로 평가"},
            {"항목": "합격 기준 (공차 기준)", "AIAG MSA 4판": "%GRR ≤ 10% 양호 / 10~30% 조건부", "ISO 22514-7": "Q_MP ≤ 30%, Q_MS ≤ 15%"},
            {"항목": "합격 기준 (능력 기준)", "AIAG MSA 4판": "NDC ≥ 5", "ISO 22514-7": "C_MP ≥ 1.33, C_MS ≥ 1.33"},
            {"항목": "교정 불확도 반영", "AIAG MSA 4판": "별도 핵심 입력 아님", "ISO 22514-7": "u_CAL 명시적 반영"},
            {"항목": "분해능 기준", "AIAG MSA 4판": "%RE < 5% (공차 대비)", "ISO 22514-7": "분해능 요구사항 검토 + u_RE 반영"},
        ])
        st.dataframe(compare_df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("### 어떤 기준을 선택해야 하나요?")
        col_a, col_b = st.columns(2)
        col_a.markdown(
            "<div style='background:#fff3e0;border:2px solid #f5a623;border-radius:10px;padding:16px;'>"
            "<b>AIAG MSA 4판 선택</b><br><br>"
            "✔ 자동차·자동차 부품 업체<br>"
            "✔ IATF 16949 인증 요구<br>"
            "✔ 고객사가 AIAG 양식 요구<br>"
            "✔ 상대적으로 단순한 분석 원할 때"
            "</div>", unsafe_allow_html=True
        )
        col_b.markdown(
            "<div style='background:#e8f5e9;border:2px solid #43a047;border-radius:10px;padding:16px;'>"
            "<b>ISO 22514-7 선택</b><br><br>"
            "✔ 계측·계량 분야, 범용 제조업<br>"
            "✔ 교정 불확도를 명시적으로 반영<br>"
            "✔ Q_MP, C_MP 지수 요구<br>"
            "✔ 불확도 전파 이론 기반 분석 원할 때"
            "</div>", unsafe_allow_html=True
        )

    # ── 탭 4: 데이터 준비 ──────────────────────────────────────
    with tab_data:
        st.markdown("### Study별 데이터 형식")

        data_studies = [
            ("🔍 Study-1", "편의·반복성",
             pd.DataFrame({"측정값": [10.002, 10.001, 9.999, 10.003, 10.000]}),
             "단일 열 숫자 데이터 (참조값·공차는 화면에서 별도 입력)"),
            ("📊 Study-2/3", "Gage R&R",
             pd.DataFrame({"Part": [1,1,1,2,2,2], "Operator": ["A","A","B","A","A","B"], "Value": [10.01,10.02,10.00,9.98,9.99,9.97]}),
             "Part / Operator / Value (또는 Gdata) 3개 열 필수"),
            ("📈 Study-4", "선형성",
             pd.DataFrame({"Reference": [5.0,5.0,10.0,10.0], "Data": [5.02,5.01,10.03,10.04]}),
             "Reference(참조값) / Data(측정값) 2개 열"),
            ("📉 Study-5", "안정성",
             pd.DataFrame({"Trial 1": [10.01,10.02,10.00], "Trial 2": [10.00,10.03,9.99], "Trial 3": [10.02,10.01,10.01]}),
             "각 열이 반복 측정값, 각 행이 부분군"),
            ("✅ Study-6", "Attribute",
             pd.DataFrame({"Sample": [1,2,3,4], "Standard": ["G","NG","G","NG"], "평가자1": ["G","NG","G","G"], "평가자2": ["G","NG","NG","NG"]}),
             "Sample / Standard(기준) / 평가자열... — G/NG, 0/1, OK/NG, 양품/불량 모두 허용"),
        ]

        for icon_name, sub, df_ex, note in data_studies:
            st.markdown(f"**{icon_name}** — {sub}")
            st.dataframe(df_ex, use_container_width=False, hide_index=True)
            st.caption(f"※ {note}")
            st.markdown("")

        st.markdown("---")
        st.markdown("### 공통 유의사항")
        st.markdown("""
- Excel(`.xlsx`, `.xls`) 및 CSV(`.csv`) 파일 모두 지원합니다.
- 첫 행은 반드시 **열 이름(헤더)**이어야 합니다.
- 숫자 열에 텍스트가 섞이면 해당 행은 자동 제외됩니다.
- Study-6 판정값: `G / NG`, `0 / 1`, `OK / NG`, `PASS / FAIL`, `양품 / 불량` 형식 모두 인식합니다.
- `u_CAL` 입력 시 교정성적서의 **확장불확도 U**가 아닌 **표준불확도 u = U/k**를 입력하세요.
""")

    # ── 탭 5: 용어 사전 ──────────────────────────────────────
    with tab_terms:
        st.markdown("### MSA 주요 용어 사전")
        terms_data = [
            ("RE", "Resolution / 분해능", "측정기가 나타낼 수 있는 최소 눈금 단위. 예: 0.001 mm"),
            ("Bias", "편의 / 치우침", "반복 측정 평균값과 참값(RefV)의 차이"),
            ("%Bias", "편의율", "공차 대비 편의 비율. 10% 이하 권장"),
            ("Cg", "게이지 반복성 능력지수", "반복성만 고려한 게이지 능력. Cg = 0.2×T / (6×σg). ≥1.33 적합"),
            ("Cgk", "편의 보정 게이지 능력지수", "편의까지 반영한 게이지 능력. Cgk = (0.1×T − |x̄−RefV|) / (3×σg). ≥1.33 적합"),
            ("EV", "Equipment Variation / 반복성", "동일 측정자가 동일 부품을 반복 측정할 때의 변동 (6σ 기준)"),
            ("AV", "Appraiser Variation / 재현성", "측정자 간 평균 차이로 인한 변동 (6σ 기준)"),
            ("GRR", "Gage Repeatability & Reproducibility", "반복성(EV)과 재현성(AV)의 합성 — 부품-대-부품 변동을 제외한 측정 시스템 변동"),
            ("%GRR", "GRR 비율", "총변동(SV) 또는 공차 대비 GRR 비율. %SV = 6σ_GRR / 6σ_Total × 100. ≤10% 적합, ≤30% 조건부"),
            ("IA", "Interaction / 교호작용", "부품×측정자 교호작용. 특정 측정자가 특정 부품을 다르게 측정할 때 발생. p ≤ 0.05이면 유의"),
            ("NDC", "Number of Distinct Categories", "측정 시스템이 구별할 수 있는 범주 수. NDC = 1.41 × σ_part / σ_GRR. ≥5 적합"),
            ("u_CAL", "교정 표준불확도", "교정성적서의 확장불확도 U를 포함인수 k로 나눈 값 (u = U/k)"),
            ("u_BI", "편의 불확도", "측정 치우침에서 기인하는 표준불확도 성분"),
            ("u_EVR", "반복성 불확도 (선형성)", "선형성 데이터의 순수 오차에서 추출한 반복성 불확도"),
            ("u_EVO", "반복성 불확도 (GRR)", "Gage R&R 데이터의 ANOVA 반복성에서 추출한 불확도"),
            ("u_AV", "재현성 불확도", "측정자 간 차이에서 기인하는 불확도 성분"),
            ("u_IA", "교호작용 불확도", "부품×측정자 교호작용에서 기인하는 불확도 성분 (교호작용 유의 시 적용)"),
            ("u_EV", "유효 반복성 불확도", "u_EVR, u_RE, u_EVO 중 최댓값 채택. u_EV = max(u_EVR, u_RE, u_EVO)"),
            ("u_LIN", "선형성 불확도", "적합결여(Lack-of-Fit)에서 추출한 선형성 불확도"),
            ("u_MS", "측정 시스템 합성 불확도", "선형성·분해능·교정 등 시스템 성분의 합성"),
            ("u_MP", "측정 프로세스 합성 불확도", "u_MS + 반복성·재현성까지 포함한 현장 전체 불확도"),
            ("U_MS / U_MP", "확장 불확도 (k=2)", "합성 불확도의 2배. 약 95% 신뢰구간에 해당"),
            ("Q_MS", "측정 시스템 능력 지수 (%)", "2×U_MS / 공차 × 100. ≤15% 적합"),
            ("C_MS", "측정 시스템 능력 지수", "0.3×공차 / (6×u_MS). ≥1.33 적합"),
            ("Q_MP", "측정 프로세스 능력 지수 (%)", "2×U_MP / 공차 × 100. ≤30% 적합"),
            ("C_MP", "측정 프로세스 능력 지수", "0.4×공차 / (2×U_MP). ≥1.33 적합"),
            ("유효성 E", "Effectiveness", "Study-6: 평가자가 기준과 일치하게 판정한 비율. ≥0.9 적합"),
            ("P(FA)", "False Alarm Rate / 허위 경보율", "양품을 불량으로 오판한 비율. <0.05 적합 (AIAG 4판)"),
            ("P(miss)", "Miss Rate / 누락률", "불량을 양품으로 오판한 비율. <0.02 적합 (AIAG 4판)"),
            ("Kappa", "Cohen's Kappa", "우연 일치를 보정한 합치도 지수. ≥0.75 우수 (AIAG), ≥0.9 우수 (미니탭)"),
            ("Slope", "회귀 기울기", "Study-4: 1에 가까울수록 선형성 양호. |Slope-1| ≤ 0.05 권장"),
            ("Intercept", "회귀 절편", "Study-4: 0에 가까울수록 편의 없음. p > 0.05 권장"),
            ("UCL/LCL", "관리 상·하한선", "Study-5: ±3σ 기반 관리 한계선. 이탈 시 이상 원인 조사"),
        ]
        terms_df = pd.DataFrame(terms_data, columns=["용어", "한글명", "설명"])
        st.dataframe(terms_df, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────
# Study-1: Bias & Repeatability
# ─────────────────────────────────────────
_STUDY1_FIELD_HELP = {
    "re": "측정기의 최소 눈금 또는 분해능 값입니다. 예: 소수 셋째 자리까지 읽으면 0.001",
    "refv": "반복 측정의 기준이 되는 참조값입니다. 기준 시편값이나 참값을 입력하세요.",
    "usl": "허용 가능한 최대값입니다. 도면이나 규격서의 상한값을 입력하세요.",
    "lsl": "허용 가능한 최소값입니다. 도면이나 규격서의 하한값을 입력하세요.",
    "u_cal": "교정성적서의 확장불확도 U가 아니라 표준불확도 u를 입력합니다. 예: U=0.0002, k=2이면 u_CAL=0.0001",
    "u_lin": "측정 범위 위치에 따라 값이 달라지는 선형성 영향입니다. 별도 데이터가 없으면 0으로 둘 수 있습니다.",
    "u_rest": "교정과 선형성 외에 추가 반영할 기타 표준불확도입니다. 없으면 0으로 두면 됩니다.",
}


def render_study1(level):
    is_expert = _is_iso_mode(level)
    sample_cfg = _get_study1_reference_sample()
    st.header("🔍 Study-1: 편의성 및 반복성 (측정기 능력)")

    with st.container(border=True):
        st.caption("공정 및 불확도 설정" if is_expert else "공정 설정")
        if is_expert:
            st.caption("각 입력 항목의 ⓘ 아이콘에 마우스를 올리면 쉬운 설명을 볼 수 있습니다.")
            c = st.columns(7)
            re_val  = c[0].number_input("RE (분해능)",    value=float(st.session_state.get("s1_re", sample_cfg["RE"])), format="%.4f", key="s1_re", help=_STUDY1_FIELD_HELP["re"])
            ref_val = c[1].number_input("RefV (참조값)",  value=float(st.session_state.get("s1_refv", sample_cfg["RefV"])), format="%.4f", key="s1_refv", help=_STUDY1_FIELD_HELP["refv"])
            usl     = c[2].number_input("USL (상한)",     value=float(st.session_state.get("s1_usl", sample_cfg["USL"])), format="%.4f", key="s1_usl", help=_STUDY1_FIELD_HELP["usl"])
            lsl     = c[3].number_input("LSL (하한)",     value=float(st.session_state.get("s1_lsl", sample_cfg["LSL"])), format="%.4f", key="s1_lsl", help=_STUDY1_FIELD_HELP["lsl"])
            u_cal   = c[4].number_input("u_CAL (교정)",   value=float(st.session_state.get("s1_ucal", sample_cfg["u_CAL"])), format="%.5f", key="s1_ucal", help=_STUDY1_FIELD_HELP["u_cal"])
            u_lin   = c[5].number_input("u_LIN (선형성)", value=float(st.session_state.get("s1_ulin", sample_cfg["u_LIN"])), format="%.5f", key="s1_ulin", help=_STUDY1_FIELD_HELP["u_lin"])
            u_rest  = c[6].number_input("u_REST (기타)",  value=float(st.session_state.get("s1_urest", sample_cfg["u_REST"])), format="%.5f", key="s1_urest", help=_STUDY1_FIELD_HELP["u_rest"])
        else:
            c = st.columns(4)
            re_val  = c[0].number_input("분해능 (RE)",       value=float(st.session_state.get("s1_re", sample_cfg["RE"])), format="%.4f", key="s1_re", help=_STUDY1_FIELD_HELP["re"])
            ref_val = c[1].number_input("참조값 (RefV)",     value=float(st.session_state.get("s1_refv", sample_cfg["RefV"])), format="%.4f", key="s1_refv", help=_STUDY1_FIELD_HELP["refv"])
            usl     = c[2].number_input("상한 공차 (USL)",   value=float(st.session_state.get("s1_usl", sample_cfg["USL"])), format="%.4f", key="s1_usl", help=_STUDY1_FIELD_HELP["usl"])
            lsl     = c[3].number_input("하한 공차 (LSL)",   value=float(st.session_state.get("s1_lsl", sample_cfg["LSL"])), format="%.4f", key="s1_lsl", help=_STUDY1_FIELD_HELP["lsl"])
            u_cal = float(st.session_state.get("s1_ucal", 0.0001))
            u_lin = float(st.session_state.get("s1_ulin", 0.0))
            u_rest = float(st.session_state.get("s1_urest", 0.0))

    if is_expert:
        with st.expander("불확도 입력이 처음이라면"):
            st.markdown(
                "\n".join(
                    [
                        "- `u_CAL`: 교정성적서의 `확장불확도 U`가 아니라 `표준불확도 u`를 넣습니다. 예: `U=0.0002`, `k=2`이면 `u_CAL=0.0001`",
                        "- `u_LIN`: 측정 위치에 따라 값이 조금씩 달라지는 영향입니다. 자료가 없으면 `0`으로 둘 수 있습니다.",
                        "- `u_REST`: 위 두 항목 외에 추가 반영할 기타 불확도입니다. 없으면 `0`으로 둡니다.",
                        "- `RE`: 측정기의 최소 눈금값입니다. 예: `0.001 mm` 단위로 읽으면 `RE=0.001`",
                    ]
                )
            )

    # ── 버튼 행 ──
    b0, b1, b2, b3, b4 = st.columns([1.5, 1.5, 1.5, 1.5, 1.5])

    run_btn = b0.button("🚀 Study-1 분석 실행", type="primary", use_container_width=True)

    # 파일 불러오기: 드래그 안내 숨기고 Browse 버튼만 노출
    with b1:
        up_file = st.file_uploader(
            "📂 파일 불러오기",
            type=["csv", "xlsx", "xls"],
            label_visibility="collapsed",
            key="s1_upload"
        )
    if up_file:
        try:
            _store_uploaded_study1_values(up_file)
        except Exception as e:
            st.error(_localize_error(e))

    b2.button("📂 샘플 데이터 로드", on_click=_load_study1_reference_sample, use_container_width=True)

    # Excel 보고서: 결과 있으면 다운로드 활성
    with b3:
        s1_res = st.session_state.get('s1_result')
        if s1_res is not None:
            export_standard = st.session_state.get("s1_export_standard", "ISO 22514-7" if is_expert else "AIAG")
            params = _merge_report_metadata(st.session_state.get(
                "s1_export_params",
                {
                    'Standard': export_standard,
                    'RE (분해능)': re_val,
                    'RefV (참조값)': ref_val,
                    'USL (상한)': usl,
                    'LSL (하한)': lsl,
                    'TOL': usl - lsl,
                    'u_CAL': u_cal,
                    'u_LIN': u_lin,
                    'u_REST': u_rest,
                },
            ))
            excel_data = create_study1_excel(
                s1_res,
                st.session_state['s1_raw_X'],
                params,
                standard=export_standard,
            )
            if excel_data:
                standard_slug = "ISO22514-7" if "ISO 22514-7" in str(export_standard) else "AIAG"
                fname = f"MSA_Study1_{standard_slug}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button(
                    "💾 Excel 보고서", excel_data, fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.button("💾 Excel 보고서", disabled=True, use_container_width=True,
                          help="openpyxl 패키지가 필요합니다.")
        else:
            st.button("💾 Excel 보고서", disabled=True, use_container_width=True,
                      help="분석 실행 후 활성화됩니다.")

    # PDF 보고서 다운로드
    with b4:
        s1_res_pdf = st.session_state.get('s1_result')
        if s1_res_pdf is not None:
            try:
                _s1_pdf_standard = st.session_state.get("s1_export_standard", "ISO 22514-7" if is_expert else "AIAG")
                _s1_pdf_params = _merge_report_metadata(st.session_state.get("s1_export_params", {}))
                _s1_pdf_data = create_study1_pdf(
                    s1_res_pdf,
                    st.session_state.get('s1_raw_X', []),
                    _s1_pdf_params,
                    standard=_s1_pdf_standard,
                )
                st.download_button(
                    "📄 PDF 보고서",
                    data=_s1_pdf_data,
                    file_name=f"MSA_Study1_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception:
                st.button("📄 PDF 보고서", disabled=True, use_container_width=True, help="PDF 생성 실패")
        else:
            st.button("📄 PDF 보고서", disabled=True, use_container_width=True, help="분석 실행 후 활성화됩니다.")

    st.divider()

    study1_note_lines = [
        "Cg = (0.2 × TOL) / (6σ): 반복성 기준 측정 시스템 능력 지표 (1.33 이상 합격)",
        "Cgk = (0.1 × TOL - |X̄-참값|) / (3σ): 반복성과 치우침 반영 지표 (1.33 이상 합격)",
        "C_MS = 0.3 TOL / (6 × u_MS): 불확도 기반 측정 시스템 능력 지표 (1.33 이상 합격)",
        "Q_MS = 2×U_MS / TOL×100: 확장불확도 대비 공차폭 비율 (15% 이내 합격)",
        "%EV = AIAG 기준 30% 이내면 반복성은 합격",
        "p-value < 0.05 이면 치우침 유의",
        "u_CAL = 교정성적서 불확도 U / 2. 예) 교정성적서 불확도 U = 0.0002 이면 u_CAL = 0.0001",
        "Abias = |X̄-참값|×2: 측정기 평균과 참값 간 거리",
        "u_BI = Abias / √12: 편향 유래 표준 불확도",
        "u_EV = max(u_EVR, u_RE): 반복성 또는 분해능 중 큰 값",
        "U_MS = 2 × u_MS: 확장 불확도 (약 95% 신뢰 수준)",
    ]
    notes_html = "<div class='study1-note-list'>" + "".join(
        f"<div class='study1-note-item'><span class='study1-note-bullet'>■</span>{line}</div>"
        for line in study1_note_lines
    ) + "</div>"

    m1, m2 = st.columns([1.35, 1], gap="large")
    with m1:
        with st.container(border=True):
            st.markdown("<div class='study1-panel-title'>측정 시스템 주요 지표 설명 및 측정데이터 입력</div>", unsafe_allow_html=True)
            st.markdown(notes_html, unsafe_allow_html=True)
            st.markdown("<div class='study1-input-label'>측정데이터 입력 (콤마, 줄바꿈으로 구분)</div>", unsafe_allow_html=True)
            raw_input = st.text_area(
                "측정데이터 입력",
                value=st.session_state.get('s1_data', ""),
                height=220,
                label_visibility="collapsed"
            )

    if run_btn and raw_input:
        try:
            X = np.array([float(x) for x in re.findall(r'[-+]?\d*\.?\d+', raw_input)])
            # 이상값 사전 경고
            _s1_outlier = detect_outliers(X)
            if _s1_outlier["has_outliers"]:
                st.warning(
                    f"⚠️ 이상값 감지 ({_s1_outlier['method'].upper()} 기준): "
                    f"측정값 {_s1_outlier['outlier_values']} — "
                    f"허용 범위 [{_s1_outlier['bounds'][0]:.4f}, {_s1_outlier['bounds'][1]:.4f}]. "
                    "이상값을 확인 후 분석을 진행하세요."
                )
            df_sum, _, _ = run_study1_analysis(re_val, ref_val, usl, lsl, u_cal, u_lin, u_rest, X)
            st.session_state['s1_result'] = df_sum
            st.session_state['s1_raw_X'] = X
            st.session_state['s1_ref']   = ref_val
            st.session_state['s1_tol']   = usl - lsl
            st.session_state['s1_export_standard'] = "ISO 22514-7" if is_expert else "AIAG"
            st.session_state['s1_export_params'] = {
                'Standard': "ISO 22514-7" if is_expert else "AIAG",
                'RE (분해능)': re_val,
                'RefV (참조값)': ref_val,
                'USL (상한)': usl,
                'LSL (하한)': lsl,
                'TOL': usl - lsl,
                'u_CAL': u_cal,
                'u_LIN': u_lin,
                'u_REST': u_rest,
            }
        except Exception as e:
            st.error(_localize_error(e))

    with m2:
        with st.container(border=True):
            st.markdown("<div class='study1-panel-title'>분석 결과 미리보기</div>", unsafe_allow_html=True)
            if "s1_result_filter" not in st.session_state:
                st.session_state["s1_result_filter"] = "all"

            if st.session_state.get('s1_result') is not None:
                df_all = st.session_state['s1_result'].copy()
                df_base = df_all.copy()
                if not is_expert:
                    df_base = df_base[~df_base['항목'].str.contains(r'C_MS|Q_MS', regex=True, na=False)]

                judged_rows = df_base[df_base["판정"].isin(["OK", "NG"])].copy()
                ok_count = int((judged_rows["판정"] == "OK").sum())
                ng_count = int((judged_rows["판정"] == "NG").sum())
                is_pass = ng_count == 0 and not judged_rows.empty
                verdict_text = "측정시스템 판정: 합격" if is_pass else "측정시스템 판정: 불합격"
                verdict_detail = f"OK {ok_count}건 / NG {ng_count}건"
                verdict_box = "status-ok-box" if is_pass else "status-ng-box"
                st.markdown(
                    f"<div class='{verdict_box}'>{verdict_text} ({verdict_detail})</div>",
                    unsafe_allow_html=True
                )

                df_disp = df_base.copy()

                toolbar = st.columns([1.6, 0.8, 0.8, 0.95])
                search_query = toolbar[0].text_input("검색:", key="s1_result_search")
                if toolbar[1].button(
                    "전체",
                    key="s1_filter_all",
                    use_container_width=True,
                    type="primary" if st.session_state["s1_result_filter"] == "all" else "secondary",
                ):
                    st.session_state["s1_result_filter"] = "all"
                if toolbar[2].button(
                    "핵심지표",
                    key="s1_filter_core",
                    use_container_width=True,
                    type="primary" if st.session_state["s1_result_filter"] == "core" else "secondary",
                ):
                    st.session_state["s1_result_filter"] = "core"
                if toolbar[3].button(
                    "불합격만",
                    key="s1_filter_ng",
                    use_container_width=True,
                    type="primary" if st.session_state["s1_result_filter"] == "ng" else "secondary",
                ):
                    st.session_state["s1_result_filter"] = "ng"

                filter_mode = st.session_state["s1_result_filter"]
                if filter_mode == "core":
                    df_disp = df_disp[df_disp["항목"].str.contains(r"^Cg|^Cgk|^C_MS|^Q_MS|^%EV|^RE \(%\)", regex=True, na=False)]
                elif filter_mode == "ng":
                    df_disp = df_disp[df_disp["판정"] == "NG"]

                if search_query:
                    search_mask = df_disp.apply(
                        lambda row: row.astype(str).str.contains(search_query, case=False, na=False).any(),
                        axis=1
                    )
                    df_disp = df_disp[search_mask]

                def _fmt_metric(row):
                    value = row.get("값")
                    item = str(row.get("항목", ""))
                    if isinstance(value, (int, float, np.integer, np.floating)) and not pd.isna(value):
                        digits = 8 if (item.startswith("u_") or item in {"U_MS", "Abias"}) else 3
                        return f"{float(value):.{digits}f}"
                    return value

                def _row_style(row):
                    if row.get("판정") == "OK":
                        return ["background-color:rgba(151, 187, 132, 0.18)"] * len(row)
                    if row.get("판정") == "NG":
                        return ["background-color:rgba(231, 76, 60, 0.10)"] * len(row)
                    return [""] * len(row)

                df_display = df_disp.copy()
                df_display["값"] = df_display.apply(_fmt_metric, axis=1)

                st.dataframe(
                    df_display.style.apply(_row_style, axis=1),
                    use_container_width=True,
                    height=390
                )
            else:
                st.info("분석 실행 후 결과가 표시됩니다.")

    if st.session_state.get('s1_result') is not None:
        X   = st.session_state['s1_raw_X']
        ref = st.session_state['s1_ref']
        tol = st.session_state['s1_tol']
        with st.container(border=True):
            st.markdown("<div class='study1-panel-title'>그래프</div>", unsafe_allow_html=True)
            upper_limit = ref + 0.1 * tol
            lower_limit = ref - 0.1 * tol
            out_of_control = [(value > upper_limit) or (value < lower_limit) for value in X]
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                y=X,
                mode='lines+markers',
                name="측정값",
                line=dict(color='#2f6fb0', width=2),
                marker=dict(
                    size=[10 if is_out else 8 for is_out in out_of_control],
                    color=['#d9534f' if is_out else '#2f6fb0' for is_out in out_of_control],
                ),
            ))
            fig.add_hline(y=ref, line_color="#4f7f33", line_width=1.6,
                          annotation_text=f"참조값 = {ref:.3f}", annotation_position="right")
            fig.add_hline(y=upper_limit, line_dash="dash", line_color="#d9534f",
                          annotation_text=f"참조값 +0.1*공차 = {upper_limit:.3f}", annotation_position="right")
            fig.add_hline(y=lower_limit, line_dash="dash", line_color="#d9534f",
                          annotation_text=f"참조값 -0.1*공차 = {lower_limit:.3f}", annotation_position="right")
            fig.update_layout(
                title="측정값 Run Chart",
                height=360,
                margin=dict(l=20, r=160, t=55, b=20),
                xaxis_title="표본",
                yaxis_title="측정값",
                yaxis_tickformat=".3f",
            )
            fig.update_yaxes(hoverformat=".3f")
            st.plotly_chart(fig, use_container_width=True)

        study1_ai_payload = build_ai_study1_payload(
            st.session_state.get("s1_result"),
            _merge_report_metadata(st.session_state.get("s1_export_params", {})),
            standard=st.session_state.get("s1_export_standard", "ISO 22514-7" if is_expert else "AIAG"),
        )
        _render_ai_explainer_card("study1", study1_ai_payload)


# ─────────────────────────────────────────
# Study-2/3: Precision (Gage R&R)
# ─────────────────────────────────────────
_STUDY23_FIELD_HELP = {
    "re": "측정기의 최소 눈금 또는 분해능 값입니다. 예: 소수 셋째 자리까지 읽으면 0.001",
    "refv": "Type I 반복측정의 기준 참조값입니다. 기준 시편값이나 참값을 입력하세요.",
    "usl": "허용 가능한 최대값입니다. 도면이나 규격서의 상한값을 입력하세요.",
    "lsl": "허용 가능한 최소값입니다. 도면이나 규격서의 하한값을 입력하세요.",
    "u_cal": "교정성적서의 확장불확도 U가 아니라 표준불확도 u를 입력합니다. 예: U=0.0002, k=2이면 u_CAL=0.0001",
    "bias_dist": "Bias를 표준불확도로 환산할 때 가정하는 분포입니다. 잘 모르면 기본값인 균등분포를 사용하세요.",
    "re_dist": "분해능 RE를 표준불확도로 환산할 때 가정하는 분포입니다. 잘 모르면 균등분포를 사용하세요.",
    "type1_values": "한 기준값을 반복 측정한 값을 넣습니다. 콤마 또는 줄바꿈으로 구분할 수 있습니다.",
    "tol_mode": "USL/LSL을 직접 넣거나 전체 공차폭 TOL만 넣을 수 있습니다.",
    "tol": "전체 허용 공차폭입니다. USL-LSL과 같은 값입니다.",
    "u_lin": "측정 범위 위치에 따라 값이 달라지는 선형성 영향입니다. 별도 자료가 없으면 0으로 둘 수 있습니다.",
    "u_msrest": "선형성 외 측정시스템 잔여 영향입니다. 별도 근거가 없으면 0으로 둘 수 있습니다.",
    "u_t": "온도 변화 영향에 대한 표준불확도입니다. 고려하지 않으면 0으로 둡니다.",
    "u_stab": "시간 경과에 따른 안정성 영향의 표준불확도입니다. 자료가 없으면 0으로 둡니다.",
    "u_obj": "측정 대상물 자체 특성에서 오는 표준불확도입니다. 없으면 0으로 둘 수 있습니다.",
    "u_gv": "기준기나 참조기에서 오는 표준불확도입니다. 없으면 0으로 둘 수 있습니다.",
    "u_rest": "기타 미분류 표준불확도입니다. 따로 반영할 항목이 없으면 0으로 둡니다.",
}


def render_study2_3(level):
    is_expert = _is_iso_mode(level)
    st.header("📊 Study-2/3: Gage R&R 분석 (정밀도)")

    sample_data = _get_study23_reference_sample()
    st.session_state.setdefault("s23_bias_dist", "균등분포")
    st.session_state.setdefault("s23_re_dist", "균등분포")
    default_type1 = st.session_state.get(
        "s23_type1_values",
        sample_data["type1_text"]
    )
    top_cols = st.columns([1.55, 1.0], gap="large")

    with top_cols[0]:
        if is_expert:
            st.caption("각 입력 항목의 ⓘ 아이콘에 마우스를 올리면 쉬운 설명을 볼 수 있습니다.")
        if is_expert:
            config_cols = st.columns([1.2, 0.95, 1.05], gap="small")
            type1_col, tol_col, typeb_col = config_cols
        else:
            config_cols = st.columns([1.25, 0.95], gap="small")
            type1_col, tol_col = config_cols
            typeb_col = None

        with type1_col:
            with st.container(border=True):
                panel_title = "Type I 불확도 입력" if is_expert else "Type I 입력"
                st.markdown(f"<div class='study23-panel-title'>{panel_title}</div>", unsafe_allow_html=True)
                type1_metrics = st.columns(2, gap="small")
                re_val = type1_metrics[0].number_input("RE", value=float(sample_data["RE"]), format="%.4f", key="s23_re", help=_STUDY23_FIELD_HELP["re"])
                ref_val = type1_metrics[1].number_input("RefV", value=float(sample_data["RefV"]), format="%.4f", key="s23_refv", help=_STUDY23_FIELD_HELP["refv"])
                usl_direct = type1_metrics[0].number_input("USL", value=float(sample_data["USL"]), format="%.4f", key="s23_usl", help=_STUDY23_FIELD_HELP["usl"])
                lsl_direct = type1_metrics[1].number_input("LSL", value=float(sample_data["LSL"]), format="%.4f", key="s23_lsl", help=_STUDY23_FIELD_HELP["lsl"])
                if is_expert:
                    u_cal = st.number_input("u_CAL", value=float(sample_data["u_CAL"]), format="%.5f", key="s23_ucal", help=_STUDY23_FIELD_HELP["u_cal"])
                    dist_cols = st.columns(2, gap="small")
                    bias_dist = dist_cols[0].selectbox("Bias 분포", ["균등분포", "정규분포", "삼각분포", "U분포"], key="s23_bias_dist", help=_STUDY23_FIELD_HELP["bias_dist"])
                    re_dist = dist_cols[1].selectbox("u_RE 분포", ["균등분포", "정규분포", "삼각분포", "U분포"], key="s23_re_dist", help=_STUDY23_FIELD_HELP["re_dist"])
                else:
                    u_cal = float(st.session_state.get("s23_ucal", sample_data["u_CAL"]))
                    bias_dist = st.session_state.get("s23_bias_dist", "균등분포")
                    re_dist = st.session_state.get("s23_re_dist", "균등분포")
                raw_type1 = st.text_area("Type I 값", value=default_type1, height=150, key="s23_type1_values", help=_STUDY23_FIELD_HELP["type1_values"])

        with tol_col:
            with st.container(border=True):
                st.markdown("<div class='study23-panel-title'>공차 / 입력 모드</div>", unsafe_allow_html=True)
                tol_mode = st.radio(
                    "입력 모드",
                    ["USL/LSL 직접입력", "공차범위(TOL) 입력"],
                    key="s23_tol_mode",
                    help=_STUDY23_FIELD_HELP["tol_mode"],
                )
                if tol_mode == "공차범위(TOL) 입력":
                    tol = st.number_input("TOL(공차폭)", value=0.0600, format="%.6f", key="s23_tol", help=_STUDY23_FIELD_HELP["tol"])
                    usl = ref_val + tol / 2.0
                    lsl = ref_val - tol / 2.0
                    st.caption(f"계산된 USL = {usl:.3f}")
                    st.caption(f"계산된 LSL = {lsl:.3f}")
                else:
                    usl = usl_direct
                    lsl = lsl_direct
                    tol = usl - lsl
                    st.number_input("TOL(공차폭)", value=float(tol), format="%.6f", disabled=True, key="s23_tol_preview", help=_STUDY23_FIELD_HELP["tol"])
                st.markdown("<div class='study23-report-note'>공차는 Type I/II 공통 기준으로 사용됩니다.</div>", unsafe_allow_html=True)

        if is_expert and typeb_col is not None:
            with typeb_col:
                with st.container(border=True):
                    st.markdown("<div class='study23-panel-title'>추가 불확도 (Type B)</div>", unsafe_allow_html=True)
                    typeb_cols = st.columns(2, gap="small")
                    u_lin = typeb_cols[0].number_input("u_LIN", value=0.0, format="%.5f", key="s23_ulin", help=_STUDY23_FIELD_HELP["u_lin"])
                    u_msrest = typeb_cols[1].number_input("u_MSREST", value=0.0, format="%.5f", key="s23_umsrest", help=_STUDY23_FIELD_HELP["u_msrest"])
                    u_t = typeb_cols[0].number_input("u_T", value=0.0, format="%.5f", key="s23_ut", help=_STUDY23_FIELD_HELP["u_t"])
                    u_stab = typeb_cols[1].number_input("u_STAB", value=0.0, format="%.5f", key="s23_ustab", help=_STUDY23_FIELD_HELP["u_stab"])
                    u_obj = typeb_cols[0].number_input("u_OBJ", value=0.0, format="%.5f", key="s23_uobj", help=_STUDY23_FIELD_HELP["u_obj"])
                    u_gv = typeb_cols[1].number_input("u_GV", value=0.0, format="%.5f", key="s23_ugv", help=_STUDY23_FIELD_HELP["u_gv"])
                    u_rest = st.number_input("u_REST", value=0.0, format="%.5f", key="s23_urest", help=_STUDY23_FIELD_HELP["u_rest"])
        else:
            u_lin = 0.0
            u_msrest = 0.0
            u_t = 0.0
            u_stab = 0.0
            u_obj = 0.0
            u_gv = 0.0
            u_rest = 0.0

        if is_expert:
            with st.expander("Study-2/3 불확도 입력이 처음이라면"):
                st.markdown(
                    "\n".join(
                        [
                            "- `u_CAL`: 교정성적서의 `확장불확도 U`가 아니라 `표준불확도 u`를 넣습니다. 예: `U=0.0002`, `k=2`이면 `u_CAL=0.0001`",
                            "- `Bias 분포`, `u_RE 분포`: 값을 표준불확도로 바꿀 때 쓰는 가정입니다. 잘 모르면 기본값 `균등분포`를 사용하면 됩니다.",
                            "- `u_LIN`, `u_MSREST`, `u_T`, `u_STAB`, `u_OBJ`, `u_GV`, `u_REST`: 추가 반영할 불확도 항목입니다. 근거 데이터가 없으면 `0`으로 두면 됩니다.",
                            "- `RE`: 측정기의 최소 눈금값입니다. 예: `0.001 mm` 단위로 읽으면 `RE=0.001`",
                        ]
                    )
                )

    with top_cols[1]:
        side_cols = st.columns([1.45, 0.9], gap="small")
        with side_cols[0]:
            with st.container(border=True):
                st.markdown("<div class='study23-panel-title'>Type II CSV/Excel</div>", unsafe_allow_html=True)
                sample_df = st.session_state.get("s23_sample_type2_df")
                sample_selected = bool(st.session_state.get("s23_use_sample_type2")) and sample_df is not None
                file = st.file_uploader(
                    "Gage R&R 파일",
                    type=["xlsx", "xls", "csv"],
                    key="s23_upload",
                    on_change=_use_uploaded_study23_file,
                )
                source_label = (
                    "Figure 6 추출 샘플데이터 (60행)"
                    if sample_selected
                    else (file.name if file is not None else ("Figure 6 추출 샘플데이터 (60행)" if sample_df is not None else ""))
                )
                st.text_input(
                    "파일 경로/이름",
                    value=source_label,
                    disabled=True,
                    key="s23_file_name"
                )
                st.markdown("- 파일 형식: [부품, 측정자, 측정값] 3개 열", unsafe_allow_html=False)
                if sample_selected or (sample_df is not None and file is None):
                    st.caption("Figure 6 그림에서 추출한 샘플 Gage R&R 데이터가 로드되어 있습니다.")

        with side_cols[1]:
            with st.container(border=True):
                st.markdown("<div class='study23-panel-title'>실행 범위</div>", unsafe_allow_html=True)
                run_scope = st.radio(
                    "실행 모드",
                    ["Type 1만", "Type 2만", "동시 실행"],
                    index=1,
                    key="s23_run_scope",
                    label_visibility="collapsed",
                )
                st.caption("권장: 업로드 분석은 Type 2만 실행")
                if not is_expert:
                    st.caption("AIAG 모드는 핵심 항목만 표시됩니다.")

    run_t1 = run_scope in ("Type 1만", "동시 실행")
    run_t2 = run_scope in ("Type 2만", "동시 실행")

    stored_type1_result = st.session_state.get("s23_type1_result")
    stored_type2_result = st.session_state.get("s23_type2_result")
    report_text = st.session_state.get("s23_report_text", "")
    payload_json = st.session_state.get("s23_payload_json", "")
    export_params = _merge_report_metadata(st.session_state.get("s23_export_params", {}))
    export_is_expert = bool(st.session_state.get("s23_export_is_expert", is_expert))
    export_standard_slug = "ISO22514-7" if "ISO 22514-7" in str(export_params.get("standard", "")) else "AIAG"
    excel_data = None
    if stored_type1_result is not None or stored_type2_result is not None:
        excel_data = create_study23_excel(
            params=export_params,
            type1_result=stored_type1_result,
            type2_result=stored_type2_result,
            report_text=report_text,
            raw_type1_values=st.session_state.get("s23_type1_values_used", []),
            raw_type2_df=st.session_state.get("s23_type2_input_df"),
            is_expert=export_is_expert,
        )

    actions = st.columns([1.1, 1.0, 1.15, 1.05, 1.0], gap="small")
    run_btn = actions[0].button("분석 실행", type="primary", use_container_width=True, key="s23_run")
    actions[1].download_button(
        "💾 Excel 보고서",
        data=excel_data or b"",
        file_name=f"MSA_Study23_{export_standard_slug}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=not bool(excel_data),
        use_container_width=True,
    )
    actions[2].download_button(
        "데이터 저장(JSON)",
        data=payload_json or "",
        file_name=f"Study23_result_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
        disabled=not bool(payload_json),
        use_container_width=True,
    )
    # Study-2/3 PDF 보고서
    _s23_t2_res = stored_type2_result
    if _s23_t2_res is not None:
        try:
            _s23_pdf_data = create_study2_pdf(
                _s23_t2_res,
                standard=export_params.get("standard", "ISO 22514-7"),
                metadata=_get_report_metadata(),
            )
            actions[3].download_button(
                "📄 PDF 보고서",
                data=_s23_pdf_data,
                file_name=f"MSA_Study23_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except Exception:
            actions[3].button("📄 PDF 보고서", disabled=True, use_container_width=True, help="PDF 생성 실패")
    else:
        actions[3].button("📄 PDF 보고서", disabled=True, use_container_width=True, help="분석 실행 후 활성화됩니다.")
    actions[4].button(
        "샘플데이터 로드",
        use_container_width=True,
        key="s23_type1_sample",
        on_click=_load_study23_reference_sample,
    )

    if run_btn:
        try:
            if not run_t1 and not run_t2:
                raise ValueError("실행 모드를 하나 이상 선택하세요.")
            if usl <= lsl:
                raise ValueError("USL > LSL 이어야 합니다.")

            type1_context = _build_study23_type1_context(
                re_val, ref_val, usl, lsl, u_cal, u_lin, u_msrest, raw_type1, bias_dist, re_dist
            )
            type1_result = None
            if run_t1:
                if len(type1_context["values"]) < 1:
                    raise ValueError("Type 1 실행에는 Type I 값이 필요합니다.")
                type1_summary, _, _ = run_study1_analysis(
                    re_val,
                    ref_val,
                    usl,
                    lsl,
                    u_cal,
                    u_lin,
                    u_msrest,
                    type1_context["values"],
                    bias_dist=bias_dist,
                    ure_dist=re_dist,
                )
                type1_result = {"summary": type1_summary}

            type2_result = None
            if run_t2:
                if sample_selected:
                    df_type2 = sample_df.copy()
                elif file is not None:
                    df_type2 = load_and_normalize_data(file, normalize_study2_df)
                elif sample_df is not None:
                    df_type2 = sample_df.copy()
                else:
                    raise ValueError("Type 2 실행에는 CSV, Excel 또는 샘플 데이터가 필요합니다.")
                # 이상값 / 중복 데이터 사전 경고
                _s23_vals = df_type2["측정값"].dropna().values if "측정값" in df_type2.columns else np.array([])
                if len(_s23_vals) > 3:
                    _s23_out = detect_outliers(_s23_vals)
                    if _s23_out["has_outliers"]:
                        st.warning(
                            f"⚠️ Study-2 측정값 이상값 감지: {_s23_out['outlier_values']} "
                            f"(범위 [{_s23_out['bounds'][0]:.4f}, {_s23_out['bounds'][1]:.4f}])"
                        )
                _s23_dup = check_duplicate_measurements(df_type2, group_cols=list(df_type2.columns))
                if _s23_dup["has_duplicates"]:
                    st.warning(f"⚠️ 완전 중복 행 {_s23_dup['duplicate_count']}개 감지 — 데이터를 확인해 주세요.")
                type2_result = run_study2_analysis(
                    df_type2,
                    tol,
                    {
                        "u_CAL": u_cal,
                        "u_BI": type1_context["u_BI"],
                        "u_EVR": type1_context["u_EVR"],
                        "u_RE": type1_context["u_RE"],
                        "u_LIN": u_lin,
                        "u_MSREST": u_msrest,
                        "u_T": u_t,
                        "u_STAB": u_stab,
                        "u_OBJ": u_obj,
                        "u_GV": u_gv,
                        "u_REST": u_rest,
                    },
                )

            report_text = _build_study23_report(type1_result, type2_result, run_t1, run_t2, is_expert)
            chart_unc = _build_uncertainty_chart_data(type2_result, type1_context if run_t1 else None) if is_expert else pd.DataFrame()
            export_params = {
                "standard": "ISO 22514-7" if is_expert else "AIAG",
                "re": re_val,
                "refv": ref_val,
                "usl": usl,
                "lsl": lsl,
                "tol": tol,
                "u_cal": u_cal,
                "bias_dist": bias_dist,
                "re_dist": re_dist,
                "u_lin": u_lin,
                "u_msrest": u_msrest,
                "u_t": u_t,
                "u_stab": u_stab,
                "u_obj": u_obj,
                "u_gv": u_gv,
                "u_rest": u_rest,
                "type2_source": source_label,
            }
            st.session_state["s23_type1_result"] = type1_result
            st.session_state["s23_type2_result"] = type2_result
            st.session_state["s23_uncertainty_chart"] = chart_unc
            st.session_state["s23_report_text"] = report_text
            st.session_state["s23_export_params"] = export_params
            st.session_state["s23_export_is_expert"] = is_expert
            st.session_state["s23_type1_values_used"] = type1_context["values"].tolist() if type1_result is not None else []
            st.session_state["s23_type2_input_df"] = df_type2.copy() if type2_result is not None else None
            st.session_state["s23_payload_json"] = json.dumps(
                _build_study23_payload(tol, report_text, type1_result, type2_result),
                ensure_ascii=False,
                indent=2,
            )
            st.rerun()
        except Exception as e:
            st.error(_localize_error(e))

    type1_result = st.session_state.get("s23_type1_result")
    type2_result = st.session_state.get("s23_type2_result")
    chart_unc = st.session_state.get("s23_uncertainty_chart", pd.DataFrame())
    report_text = st.session_state.get("s23_report_text", "")

    if is_expert:
        bottom_cols = st.columns([0.9, 1.0, 1.3], gap="small")
        with bottom_cols[0]:
            with st.container(border=True):
                st.markdown("<div class='study23-panel-title'>불확도 시각화 (0 제외)</div>", unsafe_allow_html=True)
                if isinstance(chart_unc, pd.DataFrame) and not chart_unc.empty:
                    fig_unc = px.bar(
                        chart_unc,
                        x="값",
                        y="코드",
                        orientation="h",
                        text="값",
                        color="값",
                        color_continuous_scale="Blues",
                        hover_data=["설명"],
                    )
                    fig_unc.update_traces(texttemplate="%{text:.3f}", textposition="inside")
                    fig_unc.update_layout(
                        coloraxis_showscale=False,
                        height=460,
                        xaxis_title="불확도 값",
                        yaxis_title="",
                        margin=dict(l=10, r=10, t=10, b=10),
                    )
                    fig_unc.update_xaxes(hoverformat=".3f")
                    st.plotly_chart(fig_unc, use_container_width=True)
                else:
                    st.info("분석 실행 후 불확도 구성요소가 표시됩니다.")
        gage_col = bottom_cols[1]
        result_col = bottom_cols[2]
    else:
        bottom_cols = st.columns([0.95, 1.25], gap="small")
        gage_col = bottom_cols[0]
        result_col = bottom_cols[1]

    with gage_col:
        with st.container(border=True):
            st.markdown("<div class='study23-panel-title'>Gage R&R 시각화</div>", unsafe_allow_html=True)
            if type2_result is not None:
                chart_gage = type2_result["gage"].copy()
                chart_gage = chart_gage[~chart_gage["출처"].str.startswith("    ")]
                fig_gage = px.bar(
                    chart_gage,
                    x="출처",
                    y="SD",
                    text="SD",
                    color="출처",
                    color_discrete_sequence=["#3f7fba", "#a7bdd9", "#f3a43b", "#5aa96b", "#2d8f2d"],
                )
                fig_gage.update_traces(texttemplate="%{text:.3f}", textposition="outside")
                fig_gage.update_layout(
                    showlegend=False,
                    height=460,
                    xaxis_title="",
                    yaxis_title="Std. Dev (σ)",
                    yaxis_tickformat=".3f",
                    margin=dict(l=10, r=10, t=30, b=10),
                    title="Gage R&R 및 변동 요소",
                )
                fig_gage.update_yaxes(hoverformat=".3f")
                st.plotly_chart(fig_gage, use_container_width=True)
            else:
                st.info("Type 2 실행 후 Gage R&R 차트가 표시됩니다.")

    with result_col:
        with st.container(border=True):
            st.markdown("<div class='study23-panel-title'>결과 / 표시</div>", unsafe_allow_html=True)
            if type2_result is not None:
                visible_judgement = _study23_visible_judgement(type2_result, is_expert)
                overall = _study23_overall_status(visible_judgement, fallback_pass=type2_result.get("overall_pass", False))
                box_cls = "status-ok-box" if overall == "합격" else ("status-warn-box" if overall == "조건부 채택" else "status-ng-box")
                st.markdown(
                    f"<div class='{box_cls}'>Study-2/3 전체 판정: {overall}</div>",
                    unsafe_allow_html=True
                )
                for warning in type2_result.get("warnings", []):
                    st.warning(warning)
                metric_cards = _study23_metric_cards(visible_judgement, type2_result=type2_result, is_expert=is_expert)
                if metric_cards:
                    for idx in range(0, len(metric_cards), 2):
                        card_group = metric_cards[idx:idx + 2]
                        metric_cols = st.columns(len(card_group), gap="small")
                        for col, card in zip(metric_cols, card_group):
                            with col:
                                st.markdown(_study23_metric_card_html(card), unsafe_allow_html=True)

                if is_expert:
                    tab_labels = ["핵심 판정", "Gage 표", "ANOVA", "리포트"]
                    summary_tab, gage_tab, anova_tab, report_tab = st.tabs(tab_labels)
                else:
                    tab_labels = ["핵심 판정", "Gage 표", "리포트"]
                    summary_tab, gage_tab, report_tab = st.tabs(tab_labels)
                    anova_tab = None

                with summary_tab:
                    st.dataframe(
                        _style_study23_judgement(visible_judgement),
                        use_container_width=True,
                        height=215,
                    )

                with gage_tab:
                    st.dataframe(
                        _style_study23_gage(type2_result["gage"]),
                        use_container_width=True,
                        height=260,
                    )

                if anova_tab is not None:
                    anova_display = type2_result["anova"].reset_index().rename(columns={"index": "항목"})
                    with anova_tab:
                        st.dataframe(
                            _style_study23_anova(anova_display),
                            use_container_width=True,
                            height=240,
                        )

                with report_tab:
                    st.text_area(
                        "결과/표시",
                        value=report_text,
                        height=245 if anova_tab is not None else 305,
                        label_visibility="collapsed"
                    )
            elif report_text:
                st.text_area(
                    "결과/표시",
                    value=report_text,
                    height=320,
                    label_visibility="collapsed"
                )
            else:
                st.info("Type 2 실행 후 결과 패널이 활성화됩니다.")

    if type1_result is not None or type2_result is not None:
        study23_ai_payload = build_ai_study23_payload(
            type1_result,
            type2_result,
            export_params,
            preferred="type2" if type2_result is not None else "type1",
        )
        _render_ai_explainer_card("study23", study23_ai_payload)


# ─────────────────────────────────────────
# Study-4: Linearity
# ─────────────────────────────────────────
_STUDY4_FIELD_HELP = {
    "re": "측정기의 최소 눈금 또는 분해능 값입니다. 예: 소수 셋째 자리까지 읽으면 0.001",
    "u_cal": "교정성적서의 확장불확도 U가 아니라 표준불확도 u를 입력합니다. 예: U=0.0002, k=2이면 u_CAL=0.0001",
    "usl": "허용 가능한 최대값입니다. 도면이나 규격서의 상한값을 입력하세요.",
    "lsl": "허용 가능한 최소값입니다. 도면이나 규격서의 하한값을 입력하세요.",
    "linearity_file": "Reference와 Data 또는 LData 두 열이 있는 선형성 데이터 파일입니다. 참조값과 실제 측정데이터를 넣습니다.",
    "grr_file": "Part, Operator, Gdata 또는 Value 세 열이 있는 Gage R&R 데이터 파일입니다. ISO 모드에서 Q_MP, C_MP 계산에 활용됩니다.",
}


def render_study4(level):
    is_expert = _is_iso_mode(level)
    st.header("📏 Study-4: Linearity Study (Type 4)")
    sample_cfg = _get_study4_reference_sample()

    if is_expert:
        with st.container(border=True):
            st.caption("기본 설정")
            st.caption("각 입력 항목의 ⓘ 아이콘에 마우스를 올리면 쉬운 설명을 볼 수 있습니다.")
            setup_cols = st.columns(4)
            re_val = setup_cols[0].number_input("RE (분해능)", value=float(st.session_state.get("s4_re", sample_cfg["RE"])), format="%.4f", key="s4_re", help=_STUDY4_FIELD_HELP["re"])
            u_cal = setup_cols[1].number_input("u_CAL", value=float(st.session_state.get("s4_ucal", sample_cfg["u_CAL"])), format="%.4f", key="s4_ucal", help=_STUDY4_FIELD_HELP["u_cal"])
            usl = setup_cols[2].number_input("USL", value=float(st.session_state.get("s4_usl", sample_cfg["USL"])), format="%.3f", key="s4_usl", help=_STUDY4_FIELD_HELP["usl"])
            lsl = setup_cols[3].number_input("LSL", value=float(st.session_state.get("s4_lsl", sample_cfg["LSL"])), format="%.3f", key="s4_lsl", help=_STUDY4_FIELD_HELP["lsl"])
    else:
        re_val = float(st.session_state.get("s4_re", sample_cfg["RE"]))
        u_cal = float(st.session_state.get("s4_ucal", sample_cfg["u_CAL"]))
        usl = float(st.session_state.get("s4_usl", sample_cfg["USL"]))
        lsl = float(st.session_state.get("s4_lsl", sample_cfg["LSL"]))
        st.caption("AIAG Study-4는 기본 설정값 없이 Reference/Data 기반 선형성과 치우침을 중심으로 분석합니다.")

    if is_expert:
        with st.expander("Study-4 입력이 처음이라면"):
            st.markdown(
                "\n".join(
                    [
                        "- `u_CAL`: 교정성적서의 `확장불확도 U`가 아니라 `표준불확도 u`를 넣습니다. 예: `U=0.0002`, `k=2`이면 `u_CAL=0.0001`",
                        "- `RE`: 측정기의 최소 눈금값입니다. 예: `0.001 mm` 단위로 읽으면 `RE=0.001`",
                        "- `Linearity Study 데이터`: `Reference`는 참조값, `Data` 또는 `LData`는 실제 측정데이터입니다.",
                        "- `Gage R&R 데이터`: `Part`, `Operator`, `Gdata` 형식을 사용할 수 있으며 ISO 모드에서 함께 넣으면 `Q_MP`, `C_MP` 계산에 사용됩니다.",
                    ]
                )
            )

    input_cols = st.columns([1, 1], gap="large")
    with input_cols[0]:
        st.markdown("<div class='study23-panel-title'>Linearity Study 데이터</div>", unsafe_allow_html=True)
        st.markdown("- 파일 형식: `Reference`, `Data` 또는 `LData` 2개 열")
        up_linearity = st.file_uploader(
            "Linearity Study CSV/Excel",
            type=["csv", "xlsx", "xls"],
            key="s4_upload",
            label_visibility="collapsed",
            help=_STUDY4_FIELD_HELP["linearity_file"],
        )
        if up_linearity:
            try:
                _store_uploaded_data_once(up_linearity, "s4_upload_sig", "s4_df", normalize_study4_df)
            except Exception as e:
                st.error(_localize_error(e))
        linearity_df = st.data_editor(
            st.session_state.get("s4_df", pd.DataFrame(columns=["Reference", "Value"])).rename(columns={"Value": "Data"}),
            num_rows="dynamic",
            use_container_width=True,
            height=320,
        )

    with input_cols[1]:
        st.markdown("<div class='study23-panel-title'>Gage R&amp;R 데이터</div>", unsafe_allow_html=True)
        st.markdown("- 파일 형식: `Part`, `Operator`, `Gdata` 또는 `Value` 3개 열")
        up_grr = st.file_uploader(
            "Gage R&R CSV/Excel",
            type=["csv", "xlsx", "xls"],
            key="s4_grr_upload",
            label_visibility="collapsed",
            help=_STUDY4_FIELD_HELP["grr_file"],
        )
        if up_grr:
            try:
                _store_uploaded_data_once(up_grr, "s4_grr_upload_sig", "s4_grr_df", normalize_study2_df)
            except Exception as e:
                st.error(_localize_error(e))
        if is_expert:
            st.caption("ISO 22514-7 모드에서는 Gage R&R 데이터를 함께 넣으면 Q_MP, C_MP를 계산합니다.")
        else:
            st.caption("AIAG 모드에서는 선형성 중심 결과만 표시합니다.")
        grr_df_input = st.data_editor(
            st.session_state.get("s4_grr_df", pd.DataFrame(columns=["Part", "Operator", "Value"])).rename(columns={"Value": "Gdata"}),
            num_rows="dynamic",
            use_container_width=True,
            height=320,
        )

    current_standard = "ISO 22514-7" if is_expert else "AIAG"
    stored_standard = st.session_state.get("s4_export_standard")
    stored_result = st.session_state.get("s4_result")
    stored_linearity_df = st.session_state.get("s4_linearity_input_df")
    stored_grr_df = st.session_state.get("s4_grr_input_df")
    has_stored_result = stored_result is not None and stored_standard == current_standard

    excel_data = None
    if has_stored_result:
        excel_data = create_study4_excel(
            params=_merge_report_metadata(st.session_state.get("s4_export_params", {})),
            result=stored_result,
            raw_linearity_df=stored_linearity_df,
            raw_grr_df=stored_grr_df,
            standard=stored_standard,
        )

    action_cols = st.columns([1, 1, 1, 1], gap="small")
    action_cols[0].button("샘플데이터 로드", on_click=_load_study4_reference_sample, use_container_width=True)
    action_cols[1].download_button(
        "💾 Excel 보고서",
        data=excel_data or b"",
        file_name=f"MSA_Study4_{'ISO22514-7' if is_expert else 'AIAG'}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=not bool(excel_data),
        use_container_width=True,
    )
    # Study-4 PDF 보고서
    if has_stored_result and stored_result is not None:
        try:
            _s4_pdf_data = create_study4_pdf(
                stored_result,
                _merge_report_metadata(st.session_state.get("s4_export_params", {})),
                standard=current_standard,
            )
            action_cols[2].download_button(
                "📄 PDF 보고서",
                data=_s4_pdf_data,
                file_name=f"MSA_Study4_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except Exception:
            action_cols[2].button("📄 PDF 보고서", disabled=True, use_container_width=True, help="PDF 생성 실패")
    else:
        action_cols[2].button("📄 PDF 보고서", disabled=True, use_container_width=True, help="분석 실행 후 활성화됩니다.")
    run_btn = action_cols[3].button("분석 실행", type="primary", use_container_width=True)

    if run_btn:
        try:
            linearity_df = _drop_empty_editor_rows(linearity_df)
            if linearity_df.empty:
                raise ValueError("Study-4 선형성 데이터를 먼저 입력해 주세요.")
            linearity_df = normalize_study4_df(linearity_df)

            grr_df = _drop_empty_editor_rows(grr_df_input)
            if grr_df.empty:
                grr_df = None
            else:
                grr_df = normalize_study2_df(grr_df)

            analysis_grr_df = grr_df if is_expert else None
            res = run_study4_analysis(linearity_df, usl, lsl, u_cal, re_val, grr_df=analysis_grr_df)
            st.session_state["s4_result"] = res
            st.session_state["s4_export_standard"] = current_standard
            st.session_state["s4_export_params"] = {
                "standard": current_standard,
                "re": re_val,
                "u_cal": u_cal,
                "usl": usl,
                "lsl": lsl,
                "tol": usl - lsl,
            }
            st.session_state["s4_linearity_input_df"] = linearity_df.copy()
            st.session_state["s4_grr_input_df"] = None if grr_df is None else grr_df.copy()
            st.rerun()
        except Exception as e:
            st.error(_localize_error(e))
            return
    else:
        if not has_stored_result:
            return
        res = stored_result
        linearity_df = stored_linearity_df.copy() if isinstance(stored_linearity_df, pd.DataFrame) else pd.DataFrame()
        grr_df = stored_grr_df.copy() if isinstance(stored_grr_df, pd.DataFrame) and not stored_grr_df.empty else None

    try:
        linearity_pass = bool(res.get("linearity_pass", False))
        process_pass = bool(res.get("process_pass", True))
        overall_pass = linearity_pass if not is_expert else (linearity_pass and process_pass)

        if not is_expert:
            aiag = res["aiag_linearity"]
            aiag_judgement = _study4_aiag_judgement(res)
            aiag_judged_rows = aiag_judgement[aiag_judgement["판정"].isin(["OK", "NG"])].copy()
            aiag_pass = not aiag_judged_rows.empty and aiag_judged_rows["판정"].eq("OK").all()
            intercept_p = aiag["coefficients"].iloc[0]["P"]
            slope_value = aiag["coefficients"].iloc[1]["계수"]
            slope_delta = abs(float(slope_value) - 1.0) if not pd.isna(slope_value) else np.nan

            st.markdown("---")
            verdict_cols = st.columns([1.15, 0.85], gap="large")
            with verdict_cols[0]:
                box_cls = "status-ok-box" if aiag_pass else "status-ng-box"
                headline = "Study-4 AIAG 판정: 합격" if aiag_pass else "Study-4 AIAG 판정: 불합격"
                st.markdown(f"<div class='{box_cls}'>{headline}</div>", unsafe_allow_html=True)
                if not aiag_judgement.empty:
                    st.markdown("**판정 기준**")
                    st.dataframe(_style_verdict_dataframe(aiag_judgement, digits=4), use_container_width=True, height=120)
            with verdict_cols[1]:
                metric_cols = st.columns(2)
                metric_cols[0].metric("Intercept P", _format_display_value(intercept_p, digits=3))
                metric_cols[1].metric("Slope Δ", _format_display_value(slope_delta, digits=5))
                st.caption("기준: Intercept p > 0.05, Slope는 1에 근접")

            aiag_cols = st.columns([1.15, 0.85], gap="large")

            with aiag_cols[0]:
                pred = aiag["prediction"]
                mean_bias_points = aiag["mean_bias_points"]
                fig_aiag = go.Figure()
                fig_aiag.add_trace(
                    go.Scatter(
                        x=pred["Reference"],
                        y=pred["fit"],
                        mode="lines",
                        name="회귀",
                        line=dict(color="#b22222", width=2),
                    )
                )
                fig_aiag.add_trace(
                    go.Scatter(
                        x=pred["Reference"],
                        y=pred["ci_high"],
                        mode="lines",
                        name="95% CI",
                        line=dict(color="#2e8b57", width=1.2, dash="dash"),
                    )
                )
                fig_aiag.add_trace(
                    go.Scatter(
                        x=pred["Reference"],
                        y=pred["ci_low"],
                        mode="lines",
                        showlegend=False,
                        line=dict(color="#2e8b57", width=1.2, dash="dash"),
                    )
                )
                fig_aiag.add_trace(
                    go.Scatter(
                        x=res["raw_with_bias"]["Reference"],
                        y=res["raw_with_bias"]["Bias"],
                        mode="markers",
                        name="데이터",
                        marker=dict(color="#1f5aa6", size=7),
                    )
                )
                fig_aiag.add_trace(
                    go.Scatter(
                        x=mean_bias_points["Reference"],
                        y=mean_bias_points["Bias"],
                        mode="markers",
                        name="평균 치우침",
                        marker=dict(color="#8b1e1e", size=9, symbol="square"),
                    )
                )
                fig_aiag.add_hline(y=0, line_dash="dash", line_color="#8c8c8c", line_width=1)
                fig_aiag.update_layout(
                    height=430,
                    margin=dict(l=10, r=10, t=10, b=10),
                    xaxis_title="기준 값",
                    yaxis_title="치우침",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
                )
                fig_aiag.update_xaxes(tickformat=".3f", hoverformat=".3f")
                fig_aiag.update_yaxes(tickformat=".3f", hoverformat=".3f")
                st.plotly_chart(fig_aiag, use_container_width=True)

            with aiag_cols[1]:
                st.markdown("**Gage 선형성**")
                st.dataframe(_style_study4_aiag_coefficients(aiag["coefficients"]), use_container_width=True, hide_index=True)
                stat_cols = st.columns(2)
                stat_cols[0].markdown(f"**S**  {_format_display_value(aiag['s_value'], digits=5)}")
                stat_cols[1].markdown(f"**R-제곱**  {_format_display_value(aiag['r_squared_pct'], digits=1, suffix='%')}")
                st.markdown("**Gage 치우침**")
                st.dataframe(_style_study4_aiag_bias(aiag["bias_summary"]), use_container_width=True, hide_index=True)

            with st.expander("원본 데이터"):
                st.markdown("**Linearity Study 원본**")
                st.dataframe(
                    _format_display_df(linearity_df.rename(columns={"Value": "Data"}), digits=6),
                    use_container_width=True,
                    hide_index=True,
                    height=220,
                )
            study4_ai_payload = build_ai_study4_payload(
                res,
                _merge_report_metadata(st.session_state.get("s4_export_params", {})),
                standard=current_standard,
            )
            _render_ai_explainer_card("study4", study4_ai_payload)
            return

        st.markdown("---")
        result_cols = st.columns([1.1, 1.2], gap="large")

        with result_cols[0]:
            st.markdown("<div class='study23-panel-title'>결과 / 요약</div>", unsafe_allow_html=True)
            if is_expert:
                box_cls = "status-ok-box" if overall_pass else "status-ng-box"
                headline = "Study-4 전체 판정: 합격" if overall_pass else "Study-4 전체 판정: 불합격"
                st.markdown(f"<div class='{box_cls}'>{headline}</div>", unsafe_allow_html=True)

                linearity_display = res["linearity_capability"].copy()
                linearity_display["값"] = linearity_display["값"].map(lambda value: _truncate_decimal(value, 2))
                metric_cols = st.columns(4 if "q_lmp" in res else 2)
                metric_cols[0].metric("Q_MS (<= 15%)", _format_display_value(_truncate_decimal(res["q_ms"], 2), digits=2, suffix="%"))
                metric_cols[1].metric("C_MS (>= 1.33)", _format_display_value(_truncate_decimal(res["c_ms"], 2), digits=2))
                if "q_lmp" in res:
                    process_display = res["capability_summary"].copy()
                    process_display["값"] = process_display["값"].map(lambda value: _truncate_decimal(value, 2))
                    metric_cols[2].metric("Q_MP (<= 30%)", _format_display_value(_truncate_decimal(res["q_lmp"], 2), digits=2, suffix="%"))
                    metric_cols[3].metric("C_MP (>= 1.33)", _format_display_value(_truncate_decimal(res["c_mp"], 2), digits=2))
                else:
                    process_display = None

                criteria_df = linearity_display.copy()
                criteria_df.insert(0, "구분", "Linearity")
                if process_display is not None:
                    process_df = process_display.copy()
                    process_df.insert(0, "구분", "Process")
                    criteria_df = pd.concat([criteria_df, process_df], ignore_index=True)
                st.markdown("**판정 기준**")
                st.dataframe(_style_verdict_dataframe(criteria_df, digits=2), use_container_width=True, height=170 if len(criteria_df) > 2 else 120)
            else:
                st.info("AIAG Study-4는 Q_MS, C_MS를 표시하지 않고 회귀/ANOVA와 선형성 데이터를 중심으로 확인합니다.")

            tab_names = ["회귀/ANOVA", "불확도", "원본 데이터"]
            if is_expert and "gage_result" in res:
                tab_names = ["회귀/ANOVA", "불확도", "프로세스 능력", "원본 데이터"]
            rendered_tabs = st.tabs(tab_names)

            with rendered_tabs[0]:
                st.markdown("**회귀 계수**")
                st.dataframe(_format_display_df(res["regression_table"], digits=6), use_container_width=True, hide_index=True)
                st.markdown("**Analysis of Variance**")
                st.dataframe(_format_display_df(res["anova_table"], digits=6), use_container_width=True, hide_index=True)

            if is_expert and "gage_result" in res:
                with rendered_tabs[1]:
                    unc_left, unc_right = st.columns([2.1, 1], gap="large")
                    with unc_left:
                        st.markdown("**Uncertainty Components**")
                        st.dataframe(_format_display_df(res["uncertainty_components"], digits=6), use_container_width=True, hide_index=True, height=320)
                    with unc_right:
                        st.markdown("**Resolution Analysis**")
                        st.dataframe(_format_display_df(res["resolution_summary"], digits=6), use_container_width=True, hide_index=True, height=210)
                        st.markdown("**Measurement System Capability**")
                        linearity_display = res["linearity_capability"].copy()
                        linearity_display["값"] = linearity_display["값"].map(lambda value: _truncate_decimal(value, 2))
                        st.dataframe(_style_verdict_dataframe(linearity_display, digits=2), use_container_width=True, height=120)

                with rendered_tabs[2]:
                    proc_left, proc_right = st.columns([1.2, 1], gap="large")
                    with proc_left:
                        process_display = res["capability_summary"].copy()
                        process_display["값"] = process_display["값"].map(lambda value: _truncate_decimal(value, 2))
                        st.dataframe(_style_verdict_dataframe(process_display, digits=2), use_container_width=True, height=120)
                        st.dataframe(_style_study23_gage(res["gage_result"]["gage"]), use_container_width=True, height=260)
                    with proc_right:
                        components_df = res["process_components"].copy()
                        item_col, value_col = components_df.columns[:2]

                        _S4_SOURCE = {
                            "u_BI":   "선형성",  "u_EVR": "선형성",
                            "u_LIN":  "선형성",  "u_RE":  "설정값",
                            "u_CAL":  "설정값",  "u_REST":"기타",
                            "u_EVO":  "Gage R&R","u_AV":  "Gage R&R",
                            "u_IA":   "Gage R&R","u_T":   "기타",
                            "u_STAB": "기타",    "u_OBJ": "기타",
                            "u_GV":   "기타",    "u_MP":  "합성",
                            "U_MP":   "합성",
                        }
                        _S4_COLOR = {
                            "선형성":  "background-color:#e8f4f8;",
                            "Gage R&R":"background-color:#fff3e0;",
                            "설정값":  "background-color:#f3f3f3;",
                            "기타":    "",
                            "합성":    "background-color:#e8f5e9; font-weight:700;",
                        }
                        components_df.insert(
                            1, "출처",
                            components_df[item_col].map(lambda k: _S4_SOURCE.get(k, "기타"))
                        )

                        def _style_s4_components(df):
                            styles = pd.DataFrame("", index=df.index, columns=df.columns)
                            for i, src in enumerate(df["출처"]):
                                css = _S4_COLOR.get(src, "")
                                if css:
                                    styles.iloc[i] = css
                            return styles

                        u_evr_v = res["linearity_components"].get("u_EVR", 0)
                        u_re_v  = res["linearity_components"].get("u_RE",  0)
                        u_evo_v = res.get("gage_result", {}).get("u_elements", {}).get("u_EVO", 0)
                        u_ev_src = max(
                            [("u_EVR", u_evr_v), ("u_RE", u_re_v), ("u_EVO", u_evo_v)],
                            key=lambda t: t[1]
                        )[0]

                        styled = (
                            components_df.style
                            .format({value_col: lambda v: _format_display_value(v, digits=6)})
                            .apply(_style_s4_components, axis=None)
                            .hide(axis="index")
                        )
                        st.dataframe(styled, use_container_width=True, hide_index=True, height=420)
                        st.caption(
                            f"🔵 선형성 데이터  🟠 Gage R&R 데이터  ⚪ 설정값  🟢 합성 결과\n"
                            f"u_EV = max(u_EVR, u_RE, u_EVO) → 채택: **{u_ev_src}** "
                            f"({_format_display_value(max(u_evr_v, u_re_v, u_evo_v), digits=6)})"
                        )
                        fig_components = px.bar(
                            components_df[components_df[item_col].isin(
                                ["u_BI","u_EVR","u_LIN","u_RE","u_CAL","u_EVO","u_AV","u_IA"]
                            )],
                            x=value_col, y=item_col, orientation="h", text=value_col,
                            color="출처",
                            color_discrete_map={
                                "선형성":  "#4e91d2",
                                "Gage R&R":"#f5a623",
                                "설정값":  "#aaaaaa",
                            },
                            title="불확도 성분별 기여",
                        )
                        fig_components.update_traces(texttemplate="%{x:.4f}", textposition="outside")
                        fig_components.update_layout(xaxis_tickformat=".4f", yaxis_title="", legend_title="출처")
                        fig_components.update_xaxes(hoverformat=".4f")
                        st.plotly_chart(fig_components, use_container_width=True)

                with rendered_tabs[3]:
                    st.markdown("**Linearity Study 원본**")
                    st.dataframe(_format_display_df(linearity_df.rename(columns={"Value": "Data"}), digits=6), use_container_width=True, hide_index=True, height=220)
                    st.markdown("**Gage R&R 원본**")
                    st.dataframe(_format_display_df(grr_df.rename(columns={"Value": "Gdata"}), digits=6), use_container_width=True, hide_index=True, height=220)
            else:
                with rendered_tabs[1]:
                    unc_left, unc_right = st.columns([2.1, 1], gap="large")
                    with unc_left:
                        st.markdown("**Uncertainty Components**")
                        st.dataframe(_format_display_df(res["uncertainty_components"], digits=6), use_container_width=True, hide_index=True, height=320)
                    with unc_right:
                        st.markdown("**Resolution Analysis**")
                        st.dataframe(_format_display_df(res["resolution_summary"], digits=6), use_container_width=True, hide_index=True, height=210)
                        st.caption("AIAG 모드에서는 Q_MS, C_MS를 표시하지 않습니다.")
                with rendered_tabs[2]:
                    st.dataframe(_format_display_df(linearity_df.rename(columns={"Value": "Data"}), digits=6), use_container_width=True, hide_index=True, height=360)

        with result_cols[1]:
            st.markdown("<div class='study23-panel-title'>시각화</div>", unsafe_allow_html=True)
            fig_value = px.scatter(
                res["raw_with_bias"],
                x="Reference",
                y="Value",
                trendline="ols",
                title="Reference vs Data",
            )
            fig_value.update_layout(xaxis_tickformat=".3f", yaxis_tickformat=".3f")
            fig_value.update_xaxes(hoverformat=".3f")
            fig_value.update_yaxes(hoverformat=".3f")
            st.plotly_chart(fig_value, use_container_width=True)

            fig_bias = px.bar(
                res["raw_with_bias"].groupby("Reference", as_index=False)["Bias"].mean().sort_values("Reference"),
                x="Reference",
                y="Bias",
                text="Bias",
                title="Reference별 평균 Bias",
            )
            fig_bias.update_traces(texttemplate="%{y:.3f}", textposition="outside")
            fig_bias.update_layout(xaxis_tickformat=".3f", yaxis_tickformat=".3f")
            fig_bias.update_xaxes(hoverformat=".3f")
            fig_bias.update_yaxes(hoverformat=".3f")
            st.plotly_chart(fig_bias, use_container_width=True)

            if is_expert and "gage_result" in res:
                gage_df = res["gage_result"]["gage"].copy()
                source_col, sd_col = gage_df.columns[:2]
                gage_chart_df = gage_df.loc[
                    gage_df[source_col].astype(str).str.contains("Gage|반복성|재현성|부품|총 변동", regex=True),
                    [source_col, sd_col],
                ].copy()
                fig_gage = px.bar(
                    gage_chart_df,
                    x=source_col,
                    y=sd_col,
                    text=sd_col,
                    title="Gage R&R 및 변동 요소",
                )
                fig_gage.update_traces(texttemplate="%{y:.3f}", textposition="outside")
                fig_gage.update_layout(yaxis_tickformat=".3f", xaxis_title="")
                fig_gage.update_yaxes(hoverformat=".3f")
                st.plotly_chart(fig_gage, use_container_width=True)

        study4_ai_payload = build_ai_study4_payload(
            res,
            _merge_report_metadata(st.session_state.get("s4_export_params", {})),
            standard=current_standard,
        )
        _render_ai_explainer_card("study4", study4_ai_payload)
    except Exception as e:
        st.error(_localize_error(e))


# ─────────────────────────────────────────
# Study-5: Stability
# ─────────────────────────────────────────
def render_study5(level):
    is_expert = _is_iso_mode(level)
    st.header("📉 Study-5: Stability (안정성 관리도)")
    col1, col2 = st.columns([1, 1])
    with col1:
        st.subheader("데이터 입력")
        st.markdown("- 파일 형식: 각 행은 부분군, 각 열은 반복 측정값")
        up_file = st.file_uploader("Stability 데이터 업로드 (Excel/CSV)", type=["csv", "xlsx", "xls"], key="s5_upload")
        if up_file:
            try:
                _store_uploaded_data_once(up_file, "s5_upload_sig", "s5_df", normalize_study5_df)
            except Exception as e:
                st.error(_localize_error(e))
        if st.button("📂 샘플 데이터 로드"):
            data = np.random.normal(5.0, 0.005, (20, 3))
            st.session_state['s5_df'] = pd.DataFrame(data, columns=["Trial 1", "Trial 2", "Trial 3"])
            st.rerun()
        df      = st.data_editor(st.session_state.get('s5_df', pd.DataFrame(columns=["Trial 1", "Trial 2", "Trial 3"])),
                                 num_rows="dynamic", use_container_width=True)
        _s5_btn_cols = st.columns(2, gap="small")
        run_btn = _s5_btn_cols[0].button("🚀 Study-5 분석 실행", type="primary", use_container_width=True)
        # Study-5 PDF 버튼
        _s5_stored_res = st.session_state.get("s5_result")
        if _s5_stored_res is not None:
            try:
                _s5_pdf_data = create_study5_pdf(_s5_stored_res, metadata=_get_report_metadata())
                _s5_btn_cols[1].download_button(
                    "📄 PDF 보고서",
                    data=_s5_pdf_data,
                    file_name=f"MSA_Study5_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception:
                _s5_btn_cols[1].button("📄 PDF 보고서", disabled=True, use_container_width=True)
        else:
            _s5_btn_cols[1].button("📄 PDF 보고서", disabled=True, use_container_width=True, help="분석 실행 후 활성화됩니다.")
    if run_btn and not df.empty:
        try:
            df = normalize_study5_df(df)
            res = run_study5_analysis(df)
            st.session_state["s5_result"] = res
            with col2:
                st.subheader("시스템 안정성 판정")
                box_cls = "status-ok-box" if res['is_stable'] else "status-ng-box"
                st.markdown(
                    f"<div class='{box_cls}'>상태: {'안정 (Stable)' if res['is_stable'] else '불안정 (Unstable)'}</div>",
                    unsafe_allow_html=True
                )
                if is_expert:
                    st.dataframe(_format_display_df(res['summary']), use_container_width=True)
            chart_cols = st.columns(2, gap="large")

            with chart_cols[0]:
                st.subheader("X-bar Control Chart")
                fig_x = go.Figure()
                fig_x.add_trace(
                    go.Scatter(
                        x=list(range(1, len(res["x_bars"]) + 1)),
                        y=res["x_bars"],
                        mode="lines+markers",
                        name="Mean",
                        marker=dict(
                            color=["#e74c3c" if is_out else "#1f77b4" for is_out in res["out_x"]],
                            size=9,
                        ),
                    )
                )
                fig_x.add_hline(y=res["summary"].iloc[2, 1], line_dash="dash", line_color="red", annotation_text="UCL")
                fig_x.add_hline(y=res["summary"].iloc[3, 1], line_dash="dash", line_color="red", annotation_text="LCL")
                fig_x.update_layout(yaxis_tickformat=".3f", xaxis_title="Subgroup")
                fig_x.update_yaxes(hoverformat=".3f")
                st.plotly_chart(fig_x, use_container_width=True)

            with chart_cols[1]:
                st.subheader("R Control Chart")
                fig_r = go.Figure()
                fig_r.add_trace(
                    go.Scatter(
                        x=list(range(1, len(res["ranges"]) + 1)),
                        y=res["ranges"],
                        mode="lines+markers",
                        name="Range",
                        marker=dict(
                            color=["#e74c3c" if is_out else "#2ca02c" for is_out in res["out_r"]],
                            size=9,
                        ),
                    )
                )
                fig_r.add_hline(y=res["summary"].iloc[4, 1], line_dash="dash", line_color="red", annotation_text="UCL")
                fig_r.add_hline(y=res["summary"].iloc[5, 1], line_dash="dash", line_color="red", annotation_text="LCL")
                fig_r.update_layout(yaxis_tickformat=".3f", xaxis_title="Subgroup")
                fig_r.update_yaxes(hoverformat=".3f")
                st.plotly_chart(fig_r, use_container_width=True)
        except Exception as e:
            st.error(_localize_error(e))


# ─────────────────────────────────────────
# Study-6: Attribute
# ─────────────────────────────────────────
def render_study6(level):
    is_expert = _is_iso_mode(level)
    st.header("✅ Study-6: Attribute (계수형 합치도)")
    guide_left, guide_right = st.columns([1.4, 1], gap="large")
    with guide_left:
        st.markdown("<div class='study23-panel-title'>판정 기준표</div>", unsafe_allow_html=True)
        st.markdown("""
<table class="ref-table">
<thead><tr>
  <th>통계량</th>
  <th>유효성 E <span class="tip-icon" data-tip="검사자가 기준과 일치하게 판정한 비율">?</span></th>
  <th>허위 경보율 P(FA) <span class="tip-icon" data-tip="양품을 불량으로 오판한 비율 — AIAG MSA 4판">?</span></th>
  <th>누락률 P(miss) <span class="tip-icon" data-tip="불량을 양품으로 오판한 비율 — AIAG MSA 4판">?</span></th>
</tr></thead>
<tbody>
  <tr><td>적합</td><td>0.9 ~ 1</td><td>0 ~ 0.05</td><td>0 ~ 0.02</td></tr>
  <tr><td>조건부 채택</td><td>0.8 ~ 0.9</td><td>0.05 ~ 0.10</td><td>0.02 ~ 0.05</td></tr>
  <tr><td>부적합</td><td>0.8 이하</td><td>0.10 이상</td><td>0.05 이상</td></tr>
</tbody>
</table>""", unsafe_allow_html=True)
    with guide_right:
        st.markdown("<div class='study23-panel-title'>Kappa 해석 기준</div>", unsafe_allow_html=True)
        st.dataframe(_study6_kappa_reference(), use_container_width=True, hide_index=True)
        with st.expander("ℹ️ Fleiss' Kappa 란?"):
            st.markdown("""
**Fleiss' Kappa (κ)** 는 **3명 이상 평가자들 사이의 상호 일치도**를 측정하는 통계량입니다.

| 항목 | 설명 |
|---|---|
| **Cohen's κ** | 평가자 1명 ↔ 기준값 (1:1 비교) |
| **Fleiss' κ** | 모든 평가자들 간 동시 일치도 |

**계산 공식**

$$\\kappa = \\frac{\\bar{P} - \\bar{P}_e}{1 - \\bar{P}_e}$$

- **P̄** : 실제 관측된 평균 일치 비율
- **P̄ₑ** : 우연히 일치할 기대 확률 = p₀² + p₁²
- **p₀, p₁** : 전체에서 0(양품) / 1(불량) 판정 비율

**해석 원칙**
- Cohen's κ(정확도) ↑ + Fleiss' κ(재현성) ↑ → 측정시스템 신뢰
- Fleiss' κ가 낮으면 **평가자 간 교육/기준 통일** 필요
- Cohen's κ가 낮으면 **평가자 vs 기준 불일치** → 판정 기준 재검토 필요
""")

    col1, col2 = st.columns([1, 1])
    with col1:
        st.subheader("데이터 입력")
        st.markdown("- 파일 형식: [Sample, Standard, Appraiser1, Appraiser2, ...]")
        up_file = st.file_uploader("Attribute 데이터 업로드 (Excel/CSV)", type=["csv", "xlsx", "xls"], key="s6_upload")
        if up_file:
            try:
                _store_uploaded_data_once(up_file, "s6_upload_sig", "s6_df", normalize_study6_df)
            except Exception as e:
                st.error(_localize_error(e))
        st.button("📂 샘플 데이터 로드", on_click=_load_study6_reference_sample)

        # ── 열 이름 수정 ──
        _s6_df_cur = st.session_state.get('s6_df', pd.DataFrame(columns=["Sample", "Standard", "Appraiser1"]))
        with st.expander("✏️ 열 이름 수정"):
            _s6_cols = list(_s6_df_cur.columns)
            _s6_ncols = min(len(_s6_cols), 5)
            _s6_input_cols = st.columns(_s6_ncols) if _s6_ncols > 0 else []
            _s6_new_names = {}
            for _i, _col in enumerate(_s6_cols):
                with _s6_input_cols[_i % _s6_ncols]:
                    _s6_new_names[_col] = st.text_input(
                        f"열 {_i+1}", value=_col, key=f"s6_colname_{_i}",
                        label_visibility="collapsed",
                        placeholder=_col,
                    )
            if st.button("적용", key="s6_col_rename_apply"):
                _s6_rename_map = {o: n.strip() for o, n in _s6_new_names.items()
                                  if n.strip() and n.strip() != o}
                if _s6_rename_map:
                    if 's6_df' in st.session_state:
                        st.session_state['s6_df'] = st.session_state['s6_df'].rename(columns=_s6_rename_map)
                    else:
                        _new_cols = [_s6_new_names.get(c, c).strip() or c for c in _s6_cols]
                        st.session_state['s6_df'] = pd.DataFrame(columns=_new_cols)
                    st.rerun()

        df = st.data_editor(st.session_state.get('s6_df', pd.DataFrame(columns=["Sample", "Standard", "Appraiser1"])),
                            num_rows="dynamic", use_container_width=True)
        run_btn = st.button("🚀 Study-6 분석 실행", type="primary", use_container_width=True)
    if run_btn and not df.empty:
        try:
            df = normalize_study6_df(df)
            res_df = run_study6_analysis(df)
            res_df = res_df.rename(columns={"Kappa": "Cohen's κ", "AIAG 판정": "AIAG 판정(κ)", "미니탭 판정": "미니탭 판정(κ)"})
            fk = calc_fleiss_kappa(df)
            with col2:
                st.subheader("계수형 합치도 분석 결과")
                summary_cols = st.columns(4, gap="small")
                summary_cols[0].metric("평균 유효성(E)", _format_display_value(res_df["유효성(E)"].mean()))
                summary_cols[1].metric("평균 P(FA)", _format_display_value(res_df["P(FA)"].mean()))
                summary_cols[2].metric("평균 P(miss)", _format_display_value(res_df["P(miss)"].mean()))
                summary_cols[3].metric("평균 Cohen's κ", _format_display_value(res_df["Cohen's κ"].mean()))

                display_cols = (
                    ["평가자", "유효성(E)", "유효성 판정", "P(FA)", "P(FA) 판정", "P(miss)", "P(miss) 판정", "Cohen's κ", "AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"]
                    if is_expert
                    else ["평가자", "유효성(E)", "P(FA)", "P(miss)", "Cohen's κ", "종합 판정"]
                )
                st.dataframe(_style_s6_result(res_df[display_cols]), use_container_width=True, hide_index=True)

                # ── Fleiss' Kappa 결과 ──
                st.markdown("---")
                st.markdown("**Fleiss' Kappa — 평가자 간 상호 일치도**")
                fk_val = fk["fleiss_kappa"]
                if not np.isnan(fk_val):
                    fk_cols = st.columns(4, gap="small")
                    fk_cols[0].metric("Fleiss' κ", f"{fk_val:.4f}")
                    fk_cols[1].metric("AIAG 판정", fk["aiag_grade"])
                    fk_cols[2].metric("미니탭 판정", fk["minitab_grade"])
                    fk_cols[3].metric("평가자 수 / 부품 수", f"{fk['n_raters']} / {fk['n_subjects']}")
                    if is_expert:
                        st.caption(
                            f"P̄ (관측 일치) = {fk['p_bar']:.4f} │ "
                            f"P̄ₑ (기대 일치) = {fk['p_e_bar']:.4f} │ "
                            f"p₀ = {fk['p0']:.4f} │ p₁ = {fk['p1']:.4f}"
                        )
                    # 해석 메시지
                    cohen_mean = res_df["Cohen's κ"].mean()
                    if fk_val >= 0.75 and cohen_mean >= 0.75:
                        st.success("Fleiss' κ ≥ 0.75 & Cohen's κ ≥ 0.75 — 평가자 간 재현성 및 정확도 모두 양호합니다.")
                    elif fk_val < 0.4:
                        st.error("Fleiss' κ < 0.40 — 평가자 간 일치도가 낮습니다. 판정 기준 교육 및 통일이 필요합니다.")
                    elif fk_val < 0.75:
                        st.warning("Fleiss' κ 0.40 ~ 0.75 — 평가자 간 일치도가 보통 수준입니다. 판정 기준 재검토를 권장합니다.")
                else:
                    st.info(fk["aiag_grade"])

                # ── 보고서 다운로드 ──
                excel_data = create_study6_excel(res_df, fk, df, metadata=_get_report_metadata())
                fname = f"MSA_Study6_Attribute_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                _s6_dl_cols = st.columns(2, gap="small")
                if excel_data:
                    _s6_dl_cols[0].download_button(
                        "💾 Excel 보고서", excel_data, fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                else:
                    _s6_dl_cols[0].button("💾 Excel 보고서", disabled=True, use_container_width=True,
                                          help="openpyxl 패키지가 필요합니다.")
                try:
                    _s6_pdf_data = create_study6_pdf(res_df, fk, metadata=_get_report_metadata())
                    _s6_dl_cols[1].download_button(
                        "📄 PDF 보고서",
                        data=_s6_pdf_data,
                        file_name=f"MSA_Study6_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )
                except Exception:
                    _s6_dl_cols[1].button("📄 PDF 보고서", disabled=True, use_container_width=True, help="PDF 생성 실패")

                detail_tab, chart_tab = st.tabs(["계산식", "차트"])
                with detail_tab:
                    detail_cols = ["평가자", "유효성 계산", "P(FA) 계산", "P(miss) 계산", "AIAG 판정(κ)", "미니탭 판정(κ)", "종합 판정"]
                    st.dataframe(_style_s6_result(res_df[detail_cols]), use_container_width=True, hide_index=True)

                with chart_tab:
                    chart_df = res_df[["평가자", "유효성(E)", "P(FA)", "P(miss)", "Cohen's κ"]].melt(
                        id_vars="평가자",
                        var_name="지표",
                        value_name="값",
                    )
                    chart_df["값"] = pd.to_numeric(chart_df["값"], errors="coerce")
                    chart_df = chart_df.dropna(subset=["값"])
                    fig_attr = px.bar(
                        chart_df, x="평가자", y="값", color="지표",
                        barmode="group",
                        text=chart_df["값"].map(lambda v: f"{v:.3f}"),
                    )
                    fig_attr.update_layout(yaxis_tickformat=".3f")
                    fig_attr.update_yaxes(hoverformat=".3f")
                    fig_attr.update_traces(textposition="outside")
                    st.plotly_chart(fig_attr, use_container_width=True)
        except Exception as e:
            st.error(_localize_error(e))


if __name__ == "__main__":
    main()
