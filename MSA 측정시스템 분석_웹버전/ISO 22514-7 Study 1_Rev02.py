import numpy as np
import scipy.stats as stats
import matplotlib.pyplot as plt
import pandas as pd
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
from matplotlib import font_manager, rc
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import datetime
import pyautogui
import time
import os
import re
import platform


# ===== 1. 폰트 설정 (OS 호환성 개선) =====
def set_font():
    """운영체제에 따라 한글 폰트를 자동 설정합니다."""
    system_name = platform.system()
    try:
        if system_name == "Windows":
            font_path = "C:/Windows/Fonts/malgun.ttf"
            if os.path.exists(font_path):
                font_name = font_manager.FontProperties(fname=font_path).get_name()
                rc('font', family=font_name)
            else:
                # 맑은 고딕이 없을 경우 기본 폰트 사용
                rc('font', family="Malgun Gothic")
        elif system_name == "Darwin":  # Mac
            rc('font', family="AppleGothic")
        else:  # Linux etc.
            rc('font', family="NanumGothic")
    except Exception:
        # 폰트 설정 실패 시 기본값 유지 (에러 방지)
        pass

    plt.rcParams['axes.unicode_minus'] = False


set_font()

# ===== GUI 색상/스타일 =====
COLOR_BG_MAIN = "#f4f7fb"
COLOR_PANEL = "#ffffff"
COLOR_ACCENT = "#4f81bd"
COLOR_ACCENT2 = "#dbe5f1"
COLOR_TEXT = "#212121"
COLOR_FRAME = "#e3e6ea"
COLOR_TITLE = "#2a3a59"

# ===== Tkinter GUI 선언 =====
root = tk.Tk()
root.title("VDA 5 & ISO 22514-7 Gage R&R Study1 (불확도 계산 포함) - Rev.03")
root.geometry("1500x860")
root.configure(bg=COLOR_BG_MAIN)

# ===== 스타일 =====
style = ttk.Style()
style.theme_use("clam")
style.configure("TFrame", background=COLOR_BG_MAIN)
style.configure("TLabel", background=COLOR_BG_MAIN, foreground=COLOR_TEXT, font=("맑은 고딕", 11))
style.configure("TLabelframe", background=COLOR_BG_MAIN, foreground=COLOR_TITLE, font=("맑은 고딕", 12, "bold"))
style.configure("TLabelframe.Label", foreground=COLOR_ACCENT, background=COLOR_BG_MAIN, font=("맑은 고딕", 12, "bold"))
style.configure("TEntry", fieldbackground=COLOR_PANEL, background=COLOR_PANEL, foreground=COLOR_TEXT,
                font=("맑은 고딕", 11), relief="flat")
style.configure("TButton",
                background=COLOR_ACCENT, foreground="#fff", font=("맑은 고딕", 11, "bold"),
                borderwidth=0, padding=6, relief="flat")
style.map("TButton", background=[("active", "#275682"), ("pressed", "#375382")])
style.configure("Treeview", background=COLOR_PANEL, foreground=COLOR_TEXT, fieldbackground=COLOR_PANEL,
                bordercolor=COLOR_FRAME, font=("맑은 고딕", 11))
style.configure("Treeview.Heading", background=COLOR_ACCENT2, foreground=COLOR_TITLE,
                font=("맑은 고딕", 11, "bold"), relief="flat")
style.configure("Status.TFrame", background="#edeef0", relief="flat")
style.configure("Status.TLabel", background="#edeef0", foreground="#2d4359", font=("맑은 고딕", 12, "italic"))

# ===== 전역 변수 =====
chart_canvas = None
current_fig = None  # 현재 그려진 Figure 객체 보관용 (메모리 관리)
_last_df_summary = None  # 표시용(문자)
_last_judgement = None
_rows_cache = []  # 미리보기 렌더링용 캐시(필터/검색에 사용)
_filter_mode = tk.StringVar(value="ALL")  # ALL | CORE | NG
_search_var = tk.StringVar()

# 핵심 지표 목록
CORE_ITEMS = {
    "Cg (반복성 1.33 이상)",
    "Cgk (1.33 이상)",
    "C_MS (1.33 이상)",
    "Q_MS (15% 이내)",
    "%EV (30% 이내)",
    "RE (%) (5% 이내)",
}

# 각 항목의 기준과 판정 방식 (기준 문자열, 판정 함수)
CRITERIA = {
    "Cg (반복성 1.33 이상)": ("≥ 1.33", lambda v: (not np.isnan(v)) and v >= 1.33),
    "Cgk (1.33 이상)": ("≥ 1.33", lambda v: (not np.isnan(v)) and v >= 1.33),
    "C_MS (1.33 이상)": ("≥ 1.33", lambda v: (not np.isnan(v)) and v >= 1.33),
    "Q_MS (15% 이내)": ("≤ 15.00", lambda v: (not np.isnan(v)) and v <= 15.0),
    "%EV (30% 이내)": ("≤ 30.00%", lambda v: (not np.isnan(v)) and v <= 30.0),
    "RE (%) (5% 이내)": ("≤ 5.00%", lambda v: (not np.isnan(v)) and v <= 5.0),
}

# 상세 설명(더블클릭 팝업)
EXPLAINS = {
    "Cg (반복성 1.33 이상)": "Cg = (0.2×TOL) / (6σ): 반복성(gage repeatability)만 보는 지표. 1.33 이상 권장.",
    "Cgk (1.33 이상)": "Cgk = (0.1×TOL − |X̄−참값|) / (3σ): 반복성과 치우침을 함께 반영.",
    "C_MS (1.33 이상)": "C_MS = 0.3×TOL / (6×u_MS): 불확도 기반 측정능력. 1.33 이상 권장.",
    "Q_MS (15% 이내)": "Q_MS = (2×U_MS / TOL)×100: 측정확장불확도 대비 공차폭 비율.",
    "%EV (30% 이내)": "%EV = (6σ / TOL)×100: 반복성 비율.",
    "RE (%) (5% 이내)": "RE% = RE / TOL × 100: 분해능이 공차폭에서 차지하는 비율.",
    "p-value": "치우침 검정: p<0.05이면 평균이 참값과 유의하게 다를 수 있음.",
    "정규성 p-value": "Shapiro-Wilk 정규성 검정 p값: p<0.05면 정규성 가정이 약함.",
}


# ===== 유틸 =====
def parse_float_safe(s):
    try:
        return float(s)
    except Exception:
        return np.nan


def get_float_from_entry(entry_widget, field_name):
    """Entry 위젯에서 안전하게 float 변환, 실패 시 에러 발생"""
    val = entry_widget.get().strip()
    try:
        return float(val)
    except ValueError:
        raise ValueError(f"'{field_name}' 항목에 유효한 숫자를 입력해주세요.\n입력값: {val}")


def make_enriched_rows(df_summary):
    """ df_summary(항목, 값: 문자열) → rows list[dict]: 항목, 값, 기준, 판정, tag, is_core """
    rows = []
    for _, r in df_summary.iterrows():
        name = str(r["항목"])
        val_str = str(r["값"])
        v = parse_float_safe(val_str)
        if name in CRITERIA:
            crit_text, ok_fn = CRITERIA[name]
            is_ok = ok_fn(v)
            judge = "OK" if is_ok else "NG"
            tag = "ok" if is_ok else "ng"
        else:
            crit_text = "-"
            judge = "-"
            tag = "na"
        rows.append({
            "항목": name,
            "값": val_str,
            "기준": crit_text,
            "판정": judge,
            "tag": tag,
            "is_core": (name in CORE_ITEMS)
        })
    return rows


def filter_and_search_rows():
    """ _rows_cache에 기반해 현재 필터/검색 적용 결과 반환 """
    mode = _filter_mode.get()
    q = _search_var.get().strip().lower()
    out = []
    for row in _rows_cache:
        # 필터
        if mode == "CORE" and not row["is_core"]:
            continue
        if mode == "NG" and row["판정"] != "NG":
            continue
        # 검색
        if q:
            hay = f"{row['항목']} {row['값']} {row['기준']} {row['판정']}".lower()
            if q not in hay:
                continue
        out.append(row)
    return out


def auto_set_column_widths(tree):
    tree.update_idletasks()
    cols = tree["columns"]
    for col in cols:
        # 헤더 너비와 데이터 너비 중 최대값 계산
        # 데이터가 너무 많으면 전체 스캔 시 느려질 수 있으므로 상위 100개만 샘플링하거나 전체를 하되 예외처리
        try:
            items = tree.get_children("")
            maxw = tree.heading(col, option="text").__len__() * 10
            for item in items:
                val = str(tree.set(item, col))
                if len(val) * 8 > maxw:
                    maxw = len(val) * 8
            # 최대/최소 폭 제한
            tree.column(col, width=min(400, max(100, int(maxw) + 20)))
        except Exception:
            pass


# ===== 함수 =====
def show_help(title, message):
    messagebox.showinfo(title, message)


def get_judgement(df_summary):
    # 판정 조건: Cg≥1.33, Cgk≥1.33, C_MS≥1.33, Q_MS≤15, %EV≤30, RE(%)≤5
    try:
        val = lambda name: float(df_summary.loc[df_summary["항목"] == name, "값"].values[0])
        conds = []
        conds.append(val("Cg (반복성 1.33 이상)") >= 1.33)
        conds.append(val("Cgk (1.33 이상)") >= 1.33)
        conds.append(val("C_MS (1.33 이상)") >= 1.33)
        conds.append(val("Q_MS (15% 이내)") <= 15)
        conds.append(val("%EV (30% 이내)") <= 30)
        conds.append(val("RE (%) (5% 이내)") <= 5)
        return "합격" if all(conds) else "불합격"
    except Exception as e:
        # print("판정 에러:", e) # 디버깅용 출력 제거
        return "판정불가"


def update_result_table(df_summary):
    """미리보기 향상 버전: 4열 + 필터/검색 + 색상 태깅"""
    global _rows_cache, _last_judgement
    _rows_cache = make_enriched_rows(df_summary)

    # 판정 라벨
    judgement = get_judgement(df_summary)
    _last_judgement = judgement
    lbl_judge["text"] = f"측정 시스템 {judgement}"
    lbl_judge["foreground"] = "#2ca02c" if judgement == "합격" else ("#e74c3c" if judgement == "불합격" else "#8e44ad")

    # 렌더
    render_rows()


def render_rows(*args):
    """필터/검색 적용하여 Treeview 렌더"""
    rows = filter_and_search_rows()
    tree_result.delete(*tree_result.get_children())
    for i, row in enumerate(rows):
        tags = (row["tag"], "odd" if i % 2 else "even")
        tree_result.insert("", "end", values=(row["항목"], row["값"], row["기준"], row["판정"]), tags=tags)
    auto_set_column_widths(tree_result)


def on_tree_double_click(event):
    item = tree_result.identify_row(event.y)
    if not item:
        return
    vals = tree_result.item(item, "values")
    if not vals:
        return
    name, vstr, crit, judge = vals
    hint = EXPLAINS.get(name, "")
    msg = f"항목: {name}\n값: {vstr}\n기준: {crit}\n판정: {judge}"
    if hint:
        msg += f"\n\n설명: {hint}"
    messagebox.showinfo("상세 정보", msg)


def copy_selection():
    items = tree_result.selection()
    if not items:
        return
    lines = []
    cols = tree_result["columns"]
    header = "\t".join(cols)
    lines.append(header)
    for it in items:
        vals = tree_result.item(it, "values")
        lines.append("\t".join(map(str, vals)))
    root.clipboard_clear()
    root.clipboard_append("\n".join(lines))
    root.update()


def copy_all():
    items = tree_result.get_children("")
    if not items:
        return
    lines = []
    cols = tree_result["columns"]
    header = "\t".join(cols)
    lines.append(header)
    for it in items:
        vals = tree_result.item(it, "values")
        lines.append("\t".join(map(str, vals)))
    root.clipboard_clear()
    root.clipboard_append("\n".join(lines))
    root.update()


def export_visible_csv():
    items = tree_result.get_children("")
    if not items:
        messagebox.showerror("오류", "내보낼 데이터가 없습니다.")
        return
    data = []
    cols = tree_result["columns"]
    for it in items:
        vals = tree_result.item(it, "values")
        data.append(dict(zip(cols, vals)))
    df = pd.DataFrame(data)
    path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
    if not path:
        return
    df.to_csv(path, index=False, encoding="utf-8-sig")
    messagebox.showinfo("완료", f"표시중 데이터 CSV 저장:\n{path}")


def on_tree_right_click(event):
    iid = tree_result.identify_row(event.y)
    if iid:
        tree_result.selection_set(iid)
    menu = tk.Menu(root, tearoff=0)
    menu.add_command(label="선택행 복사", command=copy_selection)
    menu.add_command(label="전체 복사", command=copy_all)
    menu.add_separator()
    menu.add_command(label="표시행 CSV로 내보내기", command=export_visible_csv)
    menu.tk_popup(event.x_root, event.y_root)


def run_analysis():
    global chart_canvas, _last_df_summary, _last_judgement, current_fig
    try:
        # 입력값 검증 및 변환
        RE_val = get_float_from_entry(entry_dict['RE'], "RE (분해능)")
        RefV_val = get_float_from_entry(entry_dict['RefV'], "RefV (참조값)")
        USL_val = get_float_from_entry(entry_dict['USL'], "USL (상한공차)")
        LSL_val = get_float_from_entry(entry_dict['LSL'], "LSL (하한공차)")
        u_CAL_val = get_float_from_entry(entry_dict['u_CAL'], "u_CAL (교정불확도)")
        u_LIN = get_float_from_entry(entry_dict['u_LIN'], "u_LIN (선형성불확도)")
        u_MS_REST = get_float_from_entry(entry_dict['u_MS_REST'], "u_MS_REST (기타불확도)")

        # 데이터 파싱 (유연하게 숫자만 추출)
        raw_text = text_X.get("1.0", tk.END)
        tokens = re.findall(r'[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?', raw_text)
        if not tokens:
            raise ValueError("측정 데이터가 비어있습니다. 데이터를 입력해주세요.")

        X_array = np.array(list(map(float, tokens)), dtype=float)
        N = len(X_array)
        if N < 2:
            raise ValueError("측정 데이터는 최소 2개 이상이어야 합니다.")

        TOL = round(USL_val - LSL_val, 6)
        if TOL <= 0:
            raise ValueError(f"공차(TOL)가 {TOL}입니다. USL이 LSL보다 커야 합니다.")

        # 통계 계산
        mean_x = np.mean(X_array)
        std_x = np.std(X_array, ddof=1)

        # 불확도 계산
        PntRE = RE_val / TOL * 100
        Abias = abs(mean_x - RefV_val) * 2
        u_BI = Abias / np.sqrt(12)
        u_EVR = std_x
        u_RE = RE_val / np.sqrt(12)
        u_EV = max(u_EVR, u_RE)
        u_MS = np.sqrt(u_CAL_val ** 2 + u_LIN ** 2 + u_BI ** 2 + u_EV ** 2 + u_MS_REST ** 2)
        U_MS = 2 * u_MS
        Q_MS = 2 * U_MS / TOL * 100

        # Cg / Cgk / C_MS
        C_MS = 0.3 * TOL / (6 * u_MS) if u_MS > 0 else np.nan
        Cg = (0.2 * TOL) / (6 * std_x) if std_x > 0 else np.inf
        # Cgk 계산 시 분모가 0인 경우 예외처리
        if std_x > 0:
            Cgk = (0.1 * TOL - abs(mean_x - RefV_val)) / (3 * std_x)
        else:
            Cgk = np.inf if abs(mean_x - RefV_val) == 0 else np.nan

        # T-Test & CI
        t_stat, p_val = stats.ttest_1samp(X_array, RefV_val) if std_x > 0 else (np.nan, np.nan)
        se = std_x / np.sqrt(N) if std_x > 0 else np.nan
        if std_x > 0 and N > 1:
            tcrit = stats.t.ppf(0.975, N - 1)
            diff = mean_x - RefV_val
            ci_low, ci_high = diff - tcrit * se, diff + tcrit * se
        else:
            ci_low, ci_high = np.nan, np.nan
        percent_EV = (6 * std_x) / TOL * 100 if TOL > 0 else np.nan

        # 결과 DataFrame(표시용 포맷)
        df_summary = pd.DataFrame({
            "항목": [
                "참조값 (RefV)", "표준편차", "평균", "USL", "LSL", "TOL",
                "Cg (반복성 1.33 이상)", "Cgk (1.33 이상)",
                "C_MS (1.33 이상)", "Q_MS (15% 이내)", "%EV (30% 이내)", "RE (%) (5% 이내)",
                "u_CAL", "u_LIN", "u_MS_REST", "Abias", "u_BI", "u_EVR", "u_RE", "u_EV",
                "u_MS", "U_MS", "치우침", "표준오차", "t 통계량", "p-value", "95% CI 하한", "95% CI 상한"
            ],
            "값": [
                f"{RefV_val}", f"{std_x:.5f}", f"{mean_x:.5f}", f"{USL_val}", f"{LSL_val}", f"{TOL:.3f}",
                f"{Cg:.3f}", f"{Cgk:.3f}", f"{C_MS:.3f}", f"{Q_MS:.2f}", f"{percent_EV:.2f}", f"{PntRE:.2f}",
                f"{u_CAL_val:.5f}", f"{u_LIN:.5f}", f"{u_MS_REST:.5f}",
                f"{Abias:.5f}", f"{u_BI:.5f}", f"{u_EVR:.5f}", f"{u_RE:.5f}", f"{u_EV:.5f}",
                f"{u_MS:.7f}", f"{U_MS:.7f}", f"{mean_x - RefV_val:.5f}", f"{se if not np.isnan(se) else 0:.5f}",
                f"{t_stat:.3f}", f"{p_val:.5f}", f"{ci_low:.5f}", f"{ci_high:.5f}"
            ]
        })
        _last_df_summary = df_summary.copy()
        update_result_table(df_summary)

        # ===== 그래프 생성 (메모리 누수 방지 로직 적용) =====
        upper_limit = RefV_val + 0.1 * TOL
        lower_limit = RefV_val - 0.1 * TOL

        # 기존 Figure가 있다면 닫아서 메모리 해제
        if current_fig:
            plt.close(current_fig)

        # 새로운 Figure 생성
        fig, ax = plt.subplots(figsize=(15, 3.2))
        current_fig = fig  # 참조 저장

        ax.plot(range(1, N + 1), X_array, marker='o', linestyle='-', color='#3979ad')
        ax.axhline(upper_limit, color='#e06666', linestyle='--', label=f'참조값 +0.1*공차 = {upper_limit:.3f}')
        ax.axhline(RefV_val, color='#38761d', linestyle='-', label=f'참조값 = {RefV_val:.3f}')
        ax.axhline(lower_limit, color='#e06666', linestyle='--', label=f'참조값 -0.1*공차 = {lower_limit:.3f}')
        ax.set_xlabel('표본')
        ax.set_ylabel('측정값')
        ax.set_title('측정값 Run Chart')
        ax.legend(loc='upper left', bbox_to_anchor=(1, 1), fontsize=9)

        # Y축 여유 패딩
        y_min = min(np.min(X_array), lower_limit)
        y_max = max(np.max(X_array), upper_limit)
        pad = max(0.05 * (upper_limit - lower_limit), 0.1 * (y_max - y_min))
        if pad == 0: pad = 0.1
        ax.set_ylim(y_min - pad, y_max + pad)
        fig.tight_layout()

        # Canvas 갱신
        if chart_canvas:
            chart_canvas.get_tk_widget().destroy()

        chart_canvas = FigureCanvasTkAgg(fig, master=frame_plot)
        chart_canvas.draw()
        chart_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=3, pady=3)

    except ValueError as ve:
        messagebox.showerror("입력 오류", str(ve))
    except Exception as e:
        messagebox.showerror("시스템 오류", f"분석 중 예상치 못한 오류가 발생했습니다:\n{e}")


def save_excel_report():
    global _last_df_summary, _last_judgement
    try:
        if _last_df_summary is None:
            messagebox.showerror("엑셀 저장 오류", "먼저 [계산 실행 및 그래프 생성]을 실행하세요.")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return

        # 데이터프레임 복사 및 판정 결과 추가
        df = _last_df_summary.copy()

        # 마지막 행에 판정 결과 추가
        # pandas 버전에 따라 append가 deprecated일 수 있으므로 concat 사용 권장하나
        # 간단한 리스트 추가 방식 유지 (loc 이용)
        df.loc[len(df)] = ["판정 결과", _last_judgement]

        df.to_excel(file_path, index=False)

        # 스타일링
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 헤더 스타일
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True, name="맑은 고딕", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 데이터 셀 스타일
        normal_font = Font(name="맑은 고딕", size=10)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")
                cell.font = normal_font

        # 조건부 서식 (NG 항목 빨간색 표시)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            cell_item = ws[f"A{row}"]
            cell_val = ws[f"B{row}"]
            항목 = cell_item.value

            # 값이 숫자가 아니면 패스 (맨 마지막 판정 행 등)
            try:
                값 = float(cell_val.value)
            except:
                continue

            if 항목:
                if (항목.startswith("Cg") or 항목.startswith("Cgk") or 항목.startswith("C_MS")) and 값 < 1.33:
                    cell_val.fill = red_fill
                elif 항목.startswith("Q_MS") and 값 > 15:
                    cell_val.fill = red_fill
                elif 항목.startswith("%EV") and 값 > 30:
                    cell_val.fill = red_fill
                elif 항목.startswith("RE (%)") and 값 > 5:
                    cell_val.fill = red_fill

        # 컬럼 너비 자동 조정
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(12, max_len + 2)

        wb.save(file_path)
        messagebox.showinfo("완료", f"엑셀 보고서 저장 완료:\n{file_path}")
    except Exception as e:
        messagebox.showerror("엑셀 저장 오류", str(e))


def capture_and_save():
    try:
        root.state('zoomed')
        root.update()
        time.sleep(0.5)  # 대기 시간 약간 단축

        # winfo_rootx, y는 윈도우 내부 좌표이므로 타이틀바 등을 포함하려면 조정이 필요할 수 있음
        # 하지만 pyautogui는 전체 화면 기준이므로 현재 로직 유지
        x = root.winfo_rootx()
        y = root.winfo_rooty()
        w = root.winfo_width()
        h = root.winfo_height()

        screenshot = pyautogui.screenshot(region=(x, y, w, h))
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG 파일", "*.png")])
        if file_path:
            screenshot.save(file_path)
            messagebox.showinfo("저장 완료", f"캡처화면 저장 완료:\n{file_path}")
    except Exception as e:
        messagebox.showerror("캡처 오류", f"화면 캡처 중 오류가 발생했습니다.\n{e}")


def load_measurement_from_file():
    try:
        path = filedialog.askopenfilename(
            title="측정 데이터 파일 선택",
            filetypes=[("데이터 파일", "*.csv;*.xlsx;*.xls"), ("CSV", "*.csv"), ("Excel", "*.xlsx;*.xls")]
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        if ext in [".xlsx", ".xls"]:
            df = pd.read_excel(path)
        elif ext == ".csv":
            df = pd.read_csv(path)
        else:
            messagebox.showerror("오류", "지원하지 않는 파일 형식입니다.")
            return

        df_num = df.apply(pd.to_numeric, errors='coerce')
        # 유효한 숫자 데이터가 2개 이상 있는 컬럼만 추출
        numeric_cols = df_num.columns[df_num.notna().sum() >= 2].tolist()

        if not numeric_cols:
            messagebox.showerror("오류", "수치형 데이터 컬럼을 찾을 수 없습니다.")
            return

        # 가장 데이터가 많은 컬럼을 기본으로 선택
        best_col = max(numeric_cols, key=lambda c: df_num[c].notna().sum())
        series = df_num[best_col].dropna().astype(float).values

        if series.size < 2:
            messagebox.showerror("오류", "유효한 수치 데이터가 2개 미만입니다.")
            return

        text_X.delete("1.0", tk.END)
        out = []
        for i, v in enumerate(series, start=1):
            out.append(f"{v}")
            if i % 10 == 0:
                out.append("\n")
            else:
                out.append(", ")
        text_X.insert(tk.END, "".join(out).strip(", \n"))
        messagebox.showinfo("완료", f"데이터 로드 완료\n파일: {os.path.basename(path)}\n선택된 컬럼: {best_col}")
    except Exception as e:
        messagebox.showerror("오류", f"데이터 불러오기 실패:\n{e}")


# ===== 입력 프레임 (입력값 + 버튼) =====
frame_inputs = ttk.LabelFrame(root, text="입력 값", padding=10)
frame_inputs.grid(row=0, column=0, sticky="nw", padx=7, pady=3)

fields = [
    ("RE", "0.001", "분해능", "측정기의 최소 표시 단위"),
    ("RefV", "6.002", "참조값", "기준 참값"),
    ("USL", "6.03", "상한공차", "허용 최대값"),
    ("LSL", "5.97", "하한공차", "허용 최소값"),
    ("u_CAL", "0.001", "교정 불확도", "교정서 불확도의 절반"),
    ("u_LIN", "0", "선형성 불확도", "선형성 오차"),
    ("u_MS_REST", "0", "기타 불확도", "기타 원인 불확도")
]

entry_dict = {}
for idx, (key, default, desc, help_text) in enumerate(fields):
    col = idx % 4
    row = idx // 4
    ttk.Label(frame_inputs, text=f"{key}:").grid(row=row * 2, column=col * 3, sticky="e", padx=(5, 2), pady=(3, 0))
    entry = ttk.Entry(frame_inputs, width=12)
    entry.insert(0, default)
    entry.grid(row=row * 2, column=col * 3 + 1, sticky="w", padx=2, pady=(3, 0))
    entry_dict[key] = entry
    # 툴팁처럼 설명 라벨 추가
    lbl_desc = ttk.Label(frame_inputs, text=f"({desc})", foreground="gray")
    lbl_desc.grid(row=row * 2, column=col * 3 + 2, sticky="w", padx=2, pady=(3, 0))

btn_frame = ttk.Frame(frame_inputs)
btn_frame.grid(row=4, column=0, columnspan=12, pady=(10, 2), sticky="we")
ttk.Button(btn_frame, text="📝 계산 실행 및 그래프 생성", command=run_analysis, width=28).pack(side="left", padx=3)
ttk.Button(btn_frame, text="📂 데이터 불러오기 (CSV/엑셀)", command=load_measurement_from_file, width=26).pack(side="left",
                                                                                                     padx=3)
ttk.Button(btn_frame, text="💾 Excel 보고서 저장", command=save_excel_report, width=22).pack(side="left", padx=3)
ttk.Button(btn_frame, text="🖼️ 화면캡처 저장", command=capture_and_save, width=20).pack(side="left", padx=3)

# 판정 레이블
lbl_judge = ttk.Label(frame_inputs, text="", font=("맑은 고딕", 14, "bold"))
lbl_judge.grid(row=3, column=0, columnspan=12, pady=(10, 5), sticky="w")

# ===== 설명 + 데이터 입력 =====
frame_desc_data = ttk.LabelFrame(root, text="측정 시스템 주요 지표 설명 및 측정데이터 입력", padding=10)
frame_desc_data.grid(row=1, column=0, sticky="nsew", padx=7, pady=3)

desc_text = (
    "■ Cg = (0.2 × TOL) / (6×σ): 반복성 기준 측정 시스템 능력 지표 (1.33 이상 합격)\n"
    "■ Cgk = (0.1 × TOL - |X̄-참값|) / (3×σ): 반복성과 치우침 반영 지표 (1.33 이상 합격)\n"
    "■ C_MS = 0.3 TOL / (6 × u_MS): 불확도 기반 측정 능력 지표 (1.33 이상 합격)\n"
    "■ Q_MS = 2×U_MS / TOL×100: 확장불확도 대비 공차폭 비율 (15% 이하 합격)\n"
    "■ %EV = AIAG 기준 30% 이내이면 반복성은 합격\n"
    "■ p-value < 0.05 이면 치우침 유의\n"
    "■ u_CAL = 교정성적서 불확도 U / 2 예) 교정성적서 불확도 U = 0.0002 이면 u_CAL = 0.0001\n"
    "■ Abias = |X̄-참값|×2: 측정치 평균과 참값 간 거리\n"
    "■ u_BI = Abias / √12: 편향 유래 표준 불확도\n"
    "■ u_EV = max(u_EVR, u_RE): 반복성 or 분해능 중 큰 값\n"
    "■ U_MS = 2 × u_MS: 확장 불확도 (약 95% 신뢰 수준)"
)
ttk.Label(frame_desc_data, text=desc_text, justify="left", foreground="#b37e00", font=("맑은 고딕", 10)).pack(anchor="nw")

frame_data_input = ttk.LabelFrame(frame_desc_data, text="측정데이터 입력 (콤마, 줄바꿈으로 구분)", padding=8)
frame_data_input.pack(fill="both", expand=True, pady=(10, 0))
text_X = tk.Text(frame_data_input, height=6, font=("Consolas", 11))
text_X.pack(fill="both", expand=True, padx=3, pady=3)
text_X.insert(tk.END, """6.001, 6.002, 6.001, 6.001, 6.002, 6.001, 6.001, 6.000, 5.999, 6.001,
6.001, 6.000, 6.001, 6.002, 6.002, 6.002, 6.002, 6.002, 6.002, 6.000, 6.002, 6.000, 5.999,
6.002, 6.002, 6.001, 6.001, 6.000, 5.999, 5.999, 6.000, 6.001, 6.001, 6.002, 6.001, 6.001,
6.000, 6.000, 5.999, 5.999, 6.000, 6.001, 6.002, 6.001, 6.002, 6.002, 6.001, 6.002, 6.001, 6.001""")

# ===== 결과 미리보기 =====
frame_result = ttk.LabelFrame(root, text="분석 결과 미리보기", padding=10)
frame_result.grid(row=1, column=1, sticky="nsew", padx=7, pady=3)

# --- 툴바: 필터 + 검색 ---
toolbar = ttk.Frame(frame_result)
toolbar.pack(fill="x", pady=(0, 6))
btn_all = ttk.Button(toolbar, text="전체", width=8, command=lambda: (_filter_mode.set("ALL"), render_rows()))
btn_core = ttk.Button(toolbar, text="핵심지표", width=8, command=lambda: (_filter_mode.set("CORE"), render_rows()))
btn_ng = ttk.Button(toolbar, text="불합격만", width=10, command=lambda: (_filter_mode.set("NG"), render_rows()))
ttk.Label(toolbar, text="   검색:").pack(side="left", padx=(6, 2))
ent_search = ttk.Entry(toolbar, textvariable=_search_var, width=24)
ent_search.pack(side="left", padx=(0, 6))
btn_all.pack(side="left", padx=(0, 4));
btn_core.pack(side="left", padx=4);
btn_ng.pack(side="left", padx=4)
_search_var.trace_add("write", lambda *_: render_rows())

# --- Treeview (4열) ---
tree_result = ttk.Treeview(frame_result, columns=("항목", "값", "기준", "판정"), show="headings", height=26)
for col, w, anchor in [("항목", 280, "w"), ("값", 140, "e"), ("기준", 140, "center"), ("판정", 100, "center")]:
    tree_result.heading(col, text=col)
    tree_result.column(col, width=w, minwidth=80, anchor=anchor)
tree_result.pack(fill="both", expand=True, padx=3, pady=3)

# 색상 태그
tree_result.tag_configure("ok", background="#eaf7ea", foreground="#1b5e20")
tree_result.tag_configure("ng", background="#fde2e2", foreground="#b71c1c")
tree_result.tag_configure("na", background="#f5f6f7", foreground="#37474f")
tree_result.tag_configure("odd", background="")  # 얼룩무늬용 (기본)
tree_result.tag_configure("even", background="")  # 얼룩무늬용

# 이벤트
tree_result.bind("<Double-1>", on_tree_double_click)
tree_result.bind("<Button-3>", on_tree_right_click)

# 그래프 프레임
frame_plot = ttk.LabelFrame(root, text="그래프", padding=10)
frame_plot.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=7, pady=5)

# 상태바
frame_status = ttk.Frame(root, style="Status.TFrame", padding=(5, 1))
frame_status.grid(row=3, column=0, columnspan=2, sticky="ew")
lbl_sig = ttk.Label(frame_status, text="품질관리기술사 김훈희", font=("맑은 고딕", 10, "bold"))
lbl_sig.pack(side="left", padx=6)
current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
lbl_time = ttk.Label(frame_status, text=current_time, style="Status.TLabel")
lbl_time.pack(side="right")

# 그리드 최적화
root.grid_columnconfigure(0, weight=3, minsize=520)
root.grid_columnconfigure(1, weight=4, minsize=520)
root.grid_rowconfigure(0, weight=0)
root.grid_rowconfigure(1, weight=3, minsize=500)
root.grid_rowconfigure(2, weight=2, minsize=240)
root.grid_rowconfigure(3, weight=0)

root.mainloop()